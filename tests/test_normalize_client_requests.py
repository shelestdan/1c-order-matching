from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from document_text_extractor import ExtractionLine, ExtractionResult
import normalize_client_requests as ncr
from process_1c_orders import load_order_lines


class NormalizeDocumentRequestTest(unittest.TestCase):
    def test_detect_header_from_rows_merges_two_line_header(self) -> None:
        rows = [
            ["Позиция", "Наименование", "", ""],
            ["", "материала", "Ед. изм.", "Количество"],
            ["1", "Фланец Ду40", "шт", "4"],
            ["2", "Тройник Ду32", "шт", "2"],
        ]

        header_index, columns, headers = ncr.detect_header_from_rows(rows)

        self.assertEqual(header_index, 1)
        self.assertEqual(columns["position"], 1)
        self.assertEqual(columns["name"], 2)
        self.assertEqual(columns["unit"], 3)
        self.assertEqual(columns["qty"], 4)
        self.assertIn("материала", headers[1])

    def test_merge_wrapped_ocr_lines_combines_multiline_table_item(self) -> None:
        merged = ncr.merge_wrapped_ocr_lines(
            "Кран шаровой приварной\n"
            "Ду20 Ру40 раб.среда вода, 150* полнопроходной шт 480\n"
            "Фланец 80-16-01-1-В-Ст 20 ГОСТ 33259-2015 шт 50"
        )

        lines = merged.splitlines()
        self.assertEqual(len(lines), 2)
        self.assertIn("Кран шаровой приварной Ду20 Ру40", lines[0])
        parsed, issue = ncr.parse_freeform_line(lines[0], "demo.png", "ocr:1")
        self.assertIsNotNone(parsed)
        self.assertIsNone(issue)
        self.assertEqual(parsed.quantity, 480.0)

    def test_ocr_table_rows_fall_back_to_freeform_when_header_is_missing(self) -> None:
        extraction = ExtractionResult(
            file="/tmp/request.jpg",
            kind="image",
            extraction_mode="paddleocr_contrast",
            page_count=1,
            text=(
                "Чугунная задвижка Ду 50мм | Ci МЗВЗОч39р Ду50 Ру16 Т D040.023 10 | шт | 6\n"
                "Счетчик воды турбинный ДУ 50 мм | ДЕКАСТ СТВХ-50 78-50-01 | шт | 3"
            ),
            lines=[
                ExtractionLine(page=1, index=1, text="Чугунная задвижка Ду 50мм | Ci МЗВЗОч39р Ду50 Ру16 Т D040.023 10 | шт | 6", confidence=0.84, source="paddleocr"),
                ExtractionLine(page=1, index=2, text="Счетчик воды турбинный ДУ 50 мм | ДЕКАСТ СТВХ-50 78-50-01 | шт | 3", confidence=0.81, source="paddleocr"),
            ],
            warnings=[],
        )

        with mock.patch.object(ncr, "extract_document_text", return_value=extraction):
            parsed, issues, metadata = ncr.normalize_document_request(Path("/tmp/request.jpg"))

        self.assertEqual(len(parsed), 2)
        self.assertEqual(len(issues), 0)
        self.assertEqual(metadata.extraction_mode, "paddleocr_contrast")
        self.assertEqual(parsed[0].quantity, 6.0)
        self.assertEqual(parsed[1].quantity, 3.0)
        self.assertIn("Чугунная задвижка", parsed[0].name)
        self.assertIn("Счетчик воды", parsed[1].name)

    def test_ocr_table_rows_keep_delimited_mode_when_header_is_present(self) -> None:
        extraction = ExtractionResult(
            file="/tmp/request.jpg",
            kind="image",
            extraction_mode="paddleocr_rgb",
            page_count=1,
            text=(
                "Наименование материала | Марка, размер | Ед. изм | Всего кол-во\n"
                "Фланец плоский Ду50 50 мм | DN, пл Ст20 Ду50 Ру16 приварной | шт | 16"
            ),
            lines=[
                ExtractionLine(page=1, index=1, text="Наименование материала | Марка, размер | Ед. изм | Всего кол-во", confidence=0.88, source="paddleocr"),
                ExtractionLine(page=1, index=2, text="Фланец плоский Ду50 50 мм | DN, пл Ст20 Ду50 Ру16 приварной | шт | 16", confidence=0.83, source="paddleocr"),
            ],
            warnings=[],
        )

        with mock.patch.object(ncr, "extract_document_text", return_value=extraction):
            parsed, issues, _ = ncr.normalize_document_request(Path("/tmp/request.jpg"))

        self.assertEqual(len(parsed), 1)
        self.assertEqual(len(issues), 0)
        self.assertEqual(parsed[0].quantity, 16.0)
        self.assertEqual(parsed[0].unit, "шт")
        self.assertIn("Фланец плоский", parsed[0].name)

    def test_load_order_lines_skips_normalized_service_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "normalized.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Заявка"
            ws.append(["Позиция", "Наименование", "Ед. изм.", "Количество"])
            ws.append(["1", "Фланец Ду40", "шт", 2])
            issues = wb.create_sheet("Проблемы")
            issues.append(["Файл", "Источник", "Исходный текст", "Причина"])
            issues.append(["demo.xlsx", "ocr:1", "мусор", "нет количества"])
            summary = wb.create_sheet("Сводка")
            summary.append(["Параметр", "Значение"])
            summary.append(["Распознано строк", 1])
            wb.save(path)

            lines = load_order_lines(path)

            self.assertEqual(len(lines), 1)
            self.assertEqual(lines[0].sheet_name, "Заявка")
            self.assertEqual(lines[0].name, "Фланец Ду40")


if __name__ == "__main__":
    unittest.main()
