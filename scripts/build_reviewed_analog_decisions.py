#!/usr/bin/env python3
"""Build a reviewed analog decision artifact from a color-marked Excel workbook."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from process_1c_orders import build_search_text, normalize_candidate_code


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert a reviewed analog workbook into a JSON artifact.")
    parser.add_argument("--input", required=True, help="Path to the reviewed Excel workbook.")
    parser.add_argument("--output", required=True, help="Where to write the JSON artifact.")
    parser.add_argument(
        "--golden-output",
        default="",
        help="Optional path for a grouped golden set artifact.",
    )
    return parser.parse_args()


def is_rejected_fill(cell) -> bool:
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return False
    for raw_color in (fill.fgColor.rgb, fill.start_color.rgb, fill.end_color.rgb):
        color = (raw_color or "").upper()
        if color.endswith("FF0000"):
            return True
    return False


def extract_decisions(workbook_path: Path) -> list[dict[str, object]]:
    wb = load_workbook(workbook_path)
    decisions: list[dict[str, object]] = []
    for ws in wb.worksheets:
        headers = [ws.cell(row=1, column=index).value for index in range(1, ws.max_column + 1)]
        if "Наименование заявки" not in headers:
            continue
        query_column = headers.index("Наименование заявки") + 1
        analog_columns = [index + 1 for index, value in enumerate(headers) if str(value or "").startswith("Аналог ")]
        for row_index in range(2, ws.max_row + 1):
            raw_query = str(ws.cell(row=row_index, column=query_column).value or "").strip()
            if not raw_query:
                continue
            query_key = build_search_text(raw_query)
            for column_index in analog_columns:
                cell = ws.cell(row=row_index, column=column_index)
                raw_value = str(cell.value or "").strip()
                if not raw_value:
                    continue
                candidate_code = normalize_candidate_code(raw_value.split("|", 1)[0])
                if not candidate_code:
                    continue
                decision = "rejected" if is_rejected_fill(cell) else "approved"
                decisions.append(
                    {
                        "sheet": ws.title,
                        "source_row": row_index,
                        "query": raw_query,
                        "query_key": query_key,
                        "candidate_code": candidate_code,
                        "decision": decision,
                    }
                )
    return decisions


def deduplicate(decisions: list[dict[str, object]]) -> list[dict[str, object]]:
    collapsed: dict[tuple[str, str], dict[str, object]] = {}
    for item in decisions:
        key = (str(item["query_key"]), str(item["candidate_code"]))
        current = collapsed.get(key)
        if current is None:
            collapsed[key] = item
            continue
        if current["decision"] == "rejected":
            continue
        if item["decision"] == "rejected":
            collapsed[key] = item
    return sorted(
        collapsed.values(),
        key=lambda item: (str(item["query_key"]), str(item["candidate_code"])),
    )


def build_golden_set(decisions: list[dict[str, object]], source_workbook: Path) -> dict[str, object]:
    grouped: dict[str, dict[str, object]] = {}
    for item in decisions:
        query_key = str(item["query_key"])
        bucket = grouped.setdefault(
            query_key,
            {
                "query": str(item["query"]),
                "query_key": query_key,
                "approved_candidate_codes": [],
                "rejected_candidate_codes": [],
                "sources": [],
            },
        )
        source_ref = f"{item['sheet']}:{item['source_row']}"
        if source_ref not in bucket["sources"]:
            bucket["sources"].append(source_ref)
        list_key = "approved_candidate_codes" if item["decision"] == "approved" else "rejected_candidate_codes"
        candidate_code = str(item["candidate_code"])
        if candidate_code not in bucket[list_key]:
            bucket[list_key].append(candidate_code)

    cases = sorted(grouped.values(), key=lambda item: (str(item["query_key"]), str(item["query"])))
    approved_total = sum(len(case["approved_candidate_codes"]) for case in cases)
    rejected_total = sum(len(case["rejected_candidate_codes"]) for case in cases)
    return {
        "version": 1,
        "metadata": {
            "version": "1.0.0",
            "generated_on": date.today().isoformat(),
            "source_workbook": str(source_workbook),
            "description": "Сгруппированный golden set по ручной разметке аналогов",
        },
        "case_count": len(cases),
        "approved_candidate_total": approved_total,
        "rejected_candidate_total": rejected_total,
        "cases": cases,
    }


def main() -> int:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    raw_decisions = extract_decisions(input_path)
    decisions = deduplicate(raw_decisions)
    stats: defaultdict[str, int] = defaultdict(int)
    for item in decisions:
        stats[str(item["decision"])] += 1

    payload = {
        "version": 1,
        "source_workbook": str(input_path),
        "decision_count": len(decisions),
        "stats": dict(stats),
        "decisions": decisions,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Saved {len(decisions)} decisions to {output_path}")
    if args.golden_output:
        golden_output_path = Path(args.golden_output).expanduser().resolve()
        golden_output_path.parent.mkdir(parents=True, exist_ok=True)
        golden_payload = build_golden_set(decisions, input_path)
        golden_output_path.write_text(json.dumps(golden_payload, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Saved {golden_payload['case_count']} golden cases to {golden_output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
