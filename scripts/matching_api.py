#!/usr/bin/env python3
"""FastAPI backend for the 1C order matching web service.

Endpoints:
  POST /api/login          — simple shared-password auth
  POST /api/upload         — upload client file, run pipeline
  GET  /api/jobs           — list completed jobs
  GET  /api/jobs/{id}      — get job results (matches, analogs, not_found)
  POST /api/jobs/{id}/approve  — approve specific analog rows
  POST /api/jobs/{id}/export   — generate final 1C XLSX with approved analogs
"""

from __future__ import annotations

import copy
import hashlib
import json
import math
import os
import queue
import re
import secrets
import shutil
import tempfile
import threading
import time
import uuid
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, HTTPException, Header, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

from learning_store import LearningStore
from normalize_client_requests import ensure_output_dir, normalize_request_file
from process_1c_orders import (
    Candidate,
    DEFAULT_MANUAL_SELECTION_MEMORY_PATH,
    DEFAULT_REVIEWED_ANALOG_DECISIONS_PATH,
    LEGACY_MANUAL_SELECTION_MEMORY_PATH,
    OrderLine,
    PARSER_HINTS_PATH,
    STATUS_NOT_FOUND,
    StockMatcher,
    SUBSTITUTION_POLICY_PATH,
    augment_search_text_with_dimension_tags,
    build_search_text,
    build_structural_query_keys,
    clone_stock_items,
    extract_code_tokens,
    extract_dimension_tags,
    extract_family_tags,
    extract_key_tokens,
    extract_material_tags_from_search_text,
    extract_parser_hint_tags,
    extract_root_tokens,
    extract_tokens,
    load_default_manual_selection_memory,
    load_order_lines,
    load_manual_selection_memory,
    load_reviewed_analog_decisions,
    load_stock,
    load_substitution_policy,
    match_orders,
    write_outputs,
)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
STOCK_DIR = Path(__file__).resolve().parent.parent / "inputs" / "stock"
JOBS_DIR = Path(__file__).resolve().parent.parent / "outputs" / "web_jobs"
DEFAULT_USER_DIRECTORY_PATH = DATA_DIR / "users.json"
DEFAULT_EXPORT_AUDIT_PATH = DATA_DIR / "export_activity_log.json"
DEFAULT_LEARNING_DB_PATH = DEFAULT_MANUAL_SELECTION_MEMORY_PATH

# Simple shared password — override via env var MATCHING_PASSWORD
SHARED_PASSWORD = os.environ.get("MATCHING_PASSWORD", "demo2026")
ADMIN_USERNAME = os.environ.get("MATCHING_ADMIN_USERNAME", "admin").strip().lower() or "admin"
ADMIN_DISPLAY_NAME = os.environ.get("MATCHING_ADMIN_DISPLAY_NAME", "Администратор").strip() or "Администратор"

# Active sessions (token -> session info)
_sessions: dict[str, "SessionInfo"] = {}

SESSION_TTL = 60 * 60 * 8  # 8 hours
MANUAL_SEARCH_LIMIT = 12
MANUAL_SEARCH_MIN_SCORE = 28.0
MANUAL_SEARCH_MIN_OVERLAP = 0.12
MANUAL_SEARCH_MIN_SOFT_OVERLAP = 0.18
MANUAL_SEARCH_MIN_DIMENSION_BONUS = 6.0
MANUAL_SEARCH_QUERY_TOKEN_LIMIT = 8
MANUAL_SEARCH_MAX_FRAGMENT_MATCHES = 320
MANUAL_SEARCH_MAX_SCORE_CANDIDATES = 1200

STOCK_EXTENSIONS = {".csv", ".xlsx", ".xlsm", ".xls"}
STOCK_LABELS_FILE = "stock_labels.json"

# Combined stock cache — keyed by fingerprint of all stock files + their mtimes
_stock_cache: dict[str, Any] = {
    "fingerprint": None,
    "base_stock": None,
    "stock_paths": [],
}
_stock_cache_lock = threading.Lock()

# Master StockMatcher (read-only indexes) — forked per job so index building
# happens only once per stock file version, not once per job.
_matcher_cache: dict[str, Any] = {
    "fingerprint": None,
    "master_matcher": None,
}
_matcher_cache_lock = threading.Lock()

_pipeline_result_cache: dict[str, Any] = {
    "scope": None,
    "entries": {},
}

_manual_search_cache: dict[str, Any] = {
    "scope": None,
    "entries": {},
}

_learning_store: LearningStore | None = None
_job_queue: "queue.Queue[tuple[str, Path, Path, dict[str, Any], str]]" = queue.Queue()
_job_worker_thread: threading.Thread | None = None
_job_queue_state_lock = threading.Lock()
_queued_job_ids: set[str] = set()
_active_job_id: str | None = None

# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------

app = FastAPI(title="1C Order Matching API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def warm_matcher_cache_on_startup() -> None:
    _start_job_worker_if_needed()
    _recover_processing_jobs_on_startup()
    threading.Thread(
        target=_warm_matcher_cache_background,
        daemon=True,
        name="matcher-cache-warmup",
    ).start()

# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class AuthUser:
    username: str
    display_name: str
    role: str


@dataclass
class SessionInfo:
    user: AuthUser
    expires_at: float


def _hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def _user_directory_path() -> Path:
    configured = os.environ.get("MATCHING_USERS_PATH", "").strip()
    if configured:
        return Path(configured).expanduser().resolve()
    return DEFAULT_USER_DIRECTORY_PATH


def _export_audit_path() -> Path:
    configured = os.environ.get("EXPORT_AUDIT_PATH", "").strip()
    if configured:
        return Path(configured).expanduser().resolve()
    return DEFAULT_EXPORT_AUDIT_PATH


def _is_db_path(path: Path) -> bool:
    return path.suffix.lower() in {".db", ".sqlite", ".sqlite3"}


def _learning_db_path() -> Path:
    configured = os.environ.get("LEARNING_DB_PATH", "").strip()
    if configured:
        return Path(configured).expanduser().resolve()
    railway_mount_path = os.environ.get("RAILWAY_VOLUME_MOUNT_PATH", "").strip()
    if railway_mount_path:
        return (Path(railway_mount_path).expanduser().resolve() / "learning_store.db")
    legacy_configured = os.environ.get("MANUAL_SELECTION_MEMORY_PATH", "").strip()
    if legacy_configured:
        candidate = Path(legacy_configured).expanduser().resolve()
        if _is_db_path(candidate):
            return candidate
    return DEFAULT_LEARNING_DB_PATH


def _legacy_manual_selection_path() -> Path:
    configured = os.environ.get("MANUAL_SELECTION_MEMORY_PATH", "").strip()
    if configured:
        candidate = Path(configured).expanduser().resolve()
        if not _is_db_path(candidate):
            return candidate
    return LEGACY_MANUAL_SELECTION_MEMORY_PATH


def _serialize_user(user: AuthUser) -> dict[str, Any]:
    return {
        "username": user.username,
        "display_name": user.display_name,
        "role": user.role,
        "is_admin": user.role == "admin",
    }


def _load_user_directory(path: Path | None = None) -> dict[str, dict[str, str]]:
    target = path or _user_directory_path()
    if not target.exists():
        return {}
    try:
        payload = json.loads(target.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    users_payload = payload.get("users", [])
    if not isinstance(users_payload, list):
        return {}
    users: dict[str, dict[str, str]] = {}
    for raw_user in users_payload:
        if not isinstance(raw_user, dict):
            continue
        username = str(raw_user.get("username") or "").strip().lower()
        display_name = str(raw_user.get("display_name") or username).strip()
        role = str(raw_user.get("role") or "manager").strip().lower() or "manager"
        password_sha256 = str(raw_user.get("password_sha256") or "").strip().lower()
        if not username or not display_name or not password_sha256:
            continue
        users[username] = {
            "username": username,
            "display_name": display_name,
            "role": role,
            "password_sha256": password_sha256,
        }
    return users


def _authenticate_user(username: str, password: str, user_directory: dict[str, dict[str, str]] | None = None) -> AuthUser:
    normalized_username = str(username or "").strip().lower()
    if (not normalized_username or normalized_username == ADMIN_USERNAME) and password == SHARED_PASSWORD:
        return AuthUser(username=ADMIN_USERNAME, display_name=ADMIN_DISPLAY_NAME, role="admin")

    users = user_directory if user_directory is not None else _load_user_directory()
    user = users.get(normalized_username)
    if not user:
        raise HTTPException(403, "Wrong username or password")
    if _hash_password(password) != user.get("password_sha256"):
        raise HTTPException(403, "Wrong username or password")
    return AuthUser(
        username=user["username"],
        display_name=user["display_name"],
        role=user.get("role", "manager"),
    )


def _extract_session_token(authorization: str | None, token: str | None = None) -> str:
    query_token = str(token or "").strip()
    if query_token:
        return query_token
    auth_header = str(authorization or "").strip()
    if auth_header:
        return auth_header.removeprefix("Bearer ").strip()
    return ""


def _check_auth(authorization: str | None, token: str | None = None) -> AuthUser:
    session_token = _extract_session_token(authorization, token)
    if not session_token:
        raise HTTPException(401, "Authorization token required")
    session = _sessions.get(session_token)
    if not session or session.expires_at < time.time():
        _sessions.pop(session_token, None)
        raise HTTPException(401, "Session expired or invalid")
    return session.user


async def _parse_request_payload(request: Request) -> dict[str, Any]:
    content_type = str(request.headers.get("content-type") or "").split(";", 1)[0].strip().lower()
    if content_type == "application/json":
        payload = await request.json()
        return payload if isinstance(payload, dict) else {}
    if content_type == "text/plain":
        raw = (await request.body()).decode("utf-8", errors="ignore").strip()
        if not raw:
            return {}
        payload = json.loads(raw)
        return payload if isinstance(payload, dict) else {}
    if content_type in {"application/x-www-form-urlencoded", "multipart/form-data"}:
        form = await request.form()
        return {str(key): value for key, value in form.items()}
    if not content_type:
        return {}
    raise HTTPException(415, f"Unsupported content type: {content_type}")


def _coerce_login_payload(payload: dict[str, Any]) -> dict[str, Any]:
    return {
        "username": str(payload.get("username") or "").strip(),
        "password": str(payload.get("password") or ""),
    }


def _require_admin(user: AuthUser) -> None:
    if user.role != "admin":
        raise HTTPException(403, "Admin access required")


def _iso_now() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())


def _job_owner_username(data: dict[str, Any]) -> str:
    return str(data.get("created_by") or "").strip().lower()


def _ensure_job_access(user: AuthUser, data: dict[str, Any]) -> None:
    if user.role == "admin":
        return
    owner = _job_owner_username(data)
    if not owner or owner != user.username:
        raise HTTPException(403, "No access to this job")


class LoginRequest(BaseModel):
    username: str = ""
    password: str


class UserResponse(BaseModel):
    username: str
    display_name: str
    role: str
    is_admin: bool


class LoginResponse(BaseModel):
    token: str
    expires_in: int
    user: UserResponse


@app.get("/api/health")
def health():
    return {"status": "ok"}


@app.get("/api/me", response_model=UserResponse)
def me(authorization: str | None = Header(None), token: str | None = Query(None)):
    user = _check_auth(authorization, token)
    return UserResponse(**_serialize_user(user))


@app.post("/api/login", response_model=LoginResponse)
async def login(request: Request):
    payload = _coerce_login_payload(await _parse_request_payload(request))
    body = LoginRequest(**payload)
    user = _authenticate_user(body.username, body.password)
    token = secrets.token_urlsafe(32)
    _sessions[token] = SessionInfo(user=user, expires_at=time.time() + SESSION_TTL)
    return LoginResponse(token=token, expires_in=SESSION_TTL, user=UserResponse(**_serialize_user(user)))


# ---------------------------------------------------------------------------
# Pipeline helpers
# ---------------------------------------------------------------------------


def _find_all_stock_files() -> list[tuple[Path, str]]:
    """Find all stock files in inputs/stock/ and their warehouse labels.

    Returns list of (path, label) tuples. Labels are loaded from
    stock_labels.json in the stock directory. Files without explicit
    labels get an empty string.
    """
    def infer_label(path: Path, explicit_label: str) -> str:
        label = str(explicit_label or "").strip()
        if label:
            return label
        name = path.stem.lower()
        if "сантех" in name or "santeh" in name:
            return "Сантехкомплект"
        if "остат" in name or "эк" in name or re.search(r"(^|[_\\-\\s])ek($|[_\\-\\s])", name):
            return "ЭК"
        return "ЭК"

    def label_sort_key(item: tuple[Path, str]) -> tuple[int, str]:
        _, label = item
        priority = {"ЭК": 0, "Сантехкомплект": 1}
        return (priority.get(label, 9), item[0].name.lower())

    labels_path = STOCK_DIR / STOCK_LABELS_FILE
    labels: dict[str, str] = {}
    if labels_path.exists():
        try:
            labels = json.loads(labels_path.read_text("utf-8"))
        except (json.JSONDecodeError, OSError):
            pass

    candidates = sorted(STOCK_DIR.glob("*.*"), key=lambda p: p.stat().st_mtime, reverse=True)
    latest_per_label: dict[str, tuple[Path, str]] = {}
    for c in candidates:
        if c.suffix.lower() in STOCK_EXTENSIONS and c.name != STOCK_LABELS_FILE:
            label = infer_label(c, labels.get(c.name, ""))
            current = latest_per_label.get(label)
            if current is None or c.stat().st_mtime > current[0].stat().st_mtime:
                latest_per_label[label] = (c, label)
    result = sorted(latest_per_label.values(), key=label_sort_key)
    if not result:
        raise HTTPException(500, "No stock file found in inputs/stock/")
    return result


def _stock_fingerprint(stock_files: list[tuple[Path, str]]) -> str:
    """Build a fingerprint from all stock file paths + mtimes."""
    parts = []
    for path, label in stock_files:
        parts.append(f"{path.name}:{path.stat().st_mtime}:{label}")
    return "|".join(sorted(parts))


def _manual_selection_memory_path() -> Path:
    return _learning_db_path()


def _path_fingerprint(path: Path) -> str:
    if not path.exists():
        return "missing"
    stat = path.stat()
    return f"{stat.st_mtime_ns}:{stat.st_size}"


def _cache_scope_fingerprint() -> str:
    stock_files = _find_all_stock_files()
    return "|".join(
        (
            _stock_fingerprint(stock_files),
            f"manual:{_path_fingerprint(_manual_selection_memory_path())}",
            f"reviewed:{_path_fingerprint(DEFAULT_REVIEWED_ANALOG_DECISIONS_PATH)}",
            f"policy:{_path_fingerprint(SUBSTITUTION_POLICY_PATH)}",
            f"hints:{_path_fingerprint(PARSER_HINTS_PATH)}",
        )
    )


def _ensure_request_caches() -> str:
    scope = _cache_scope_fingerprint()
    for cache in (_pipeline_result_cache, _manual_search_cache):
        if cache["scope"] != scope:
            cache["scope"] = scope
            cache["entries"] = {}
    return scope


def _invalidate_learning_caches() -> None:
    _matcher_cache["fingerprint"] = None
    _matcher_cache["master_matcher"] = None
    _pipeline_result_cache["scope"] = None
    _pipeline_result_cache["entries"] = {}
    _manual_search_cache["scope"] = None
    _manual_search_cache["entries"] = {}


def _snapshot_id(job_id: str, row: dict[str, Any]) -> str:
    return f"{job_id}:{row.get('id')}"


def _feedback_candidate_key(candidate: dict[str, Any]) -> str:
    return str(candidate.get("code_1c") or "").strip().upper()


def _normalize_feedback_candidate(candidate: dict[str, Any]) -> dict[str, Any] | None:
    if not isinstance(candidate, dict):
        return None
    code = str(candidate.get("code_1c") or "").strip()
    name = str(candidate.get("name") or "").strip()
    if not code or not name:
        return None
    payload = {
        "code_1c": code,
        "name": name,
        "source_label": str(candidate.get("source_label") or "").strip(),
        "score": float(candidate.get("score") or 0.0),
        "price": candidate.get("price") or "",
        "remaining": candidate.get("remaining"),
        "stock_qty": candidate.get("stock_qty"),
        "reasons": list(candidate.get("reasons") or []),
    }
    if candidate.get("manager_choice"):
        payload["manager_choice"] = True
    return payload


def _merge_candidate_pool(*candidate_groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_code: dict[str, dict[str, Any]] = {}
    for candidates in candidate_groups:
        if not isinstance(candidates, list):
            continue
        for raw_candidate in candidates:
            candidate = _normalize_feedback_candidate(raw_candidate)
            if candidate is None:
                continue
            code = _feedback_candidate_key(candidate)
            if not code:
                continue
            current = by_code.get(code)
            if current is None:
                by_code[code] = candidate
                continue
            if len(candidate.get("reasons") or []) > len(current.get("reasons") or []):
                current["reasons"] = list(candidate.get("reasons") or [])
            if not current.get("source_label") and candidate.get("source_label"):
                current["source_label"] = candidate["source_label"]
            if not current.get("price") and candidate.get("price"):
                current["price"] = candidate["price"]
            if current.get("remaining") is None and candidate.get("remaining") is not None:
                current["remaining"] = candidate["remaining"]
            if current.get("stock_qty") is None and candidate.get("stock_qty") is not None:
                current["stock_qty"] = candidate["stock_qty"]
            if float(candidate.get("score") or 0.0) > float(current.get("score") or 0.0):
                current["score"] = candidate["score"]
            if candidate.get("manager_choice"):
                current["manager_choice"] = True
    return list(by_code.values())


def _infer_selection_source(row: dict[str, Any], candidate: dict[str, Any]) -> str:
    selected_via = str(row.get("selected_via") or "").strip()
    if selected_via:
        return selected_via
    approved_code = _feedback_candidate_key(candidate)
    if any(_feedback_candidate_key(item) == approved_code for item in row.get("analogs", []) if isinstance(item, dict)):
        return "analog"
    if row.get("selection_search_query"):
        return "manual_search"
    return "analog"


def _build_feedback_snapshot_entries(
    *,
    job_id: str,
    row: dict[str, Any],
    approved_candidate: dict[str, Any],
    visible_candidates: list[dict[str, Any]],
    user: AuthUser,
    selected_via: str,
    search_query: str,
    selected_at: str,
) -> list[dict[str, Any]]:
    query = str(row.get("raw_query") or row.get("name") or "").strip()
    query_key = str(row.get("search_text") or build_search_text(query)).strip()
    key_tokens = list(row.get("key_tokens") or [])
    root_tokens = list(row.get("root_tokens") or [])
    dimension_tags = list(row.get("dimension_tags") or [])
    if not key_tokens or not root_tokens or not dimension_tags:
        query_tokens = extract_tokens(query_key)
        if not key_tokens:
            key_tokens = sorted(extract_key_tokens(query_tokens))
        if not root_tokens:
            root_tokens = sorted(extract_root_tokens(query_tokens))
        if not dimension_tags:
            derived_tags = (
                extract_dimension_tags(query)
                | extract_family_tags(query)
                | extract_parser_hint_tags(query)
                | extract_material_tags_from_search_text(query_key)
            )
            dimension_tags = sorted(derived_tags)
    structure_keys = list(
        build_structural_query_keys(
            search_text=query_key,
            key_tokens=key_tokens,
            dimension_tags=dimension_tags,
        )
    )
    clean_search_query = search_query.strip()
    snapshot_id = _snapshot_id(job_id, row)
    normalized_approved = _normalize_feedback_candidate(approved_candidate)
    if normalized_approved is None:
        return []
    approved_code = _feedback_candidate_key(normalized_approved)

    rejected_by_code: dict[str, dict[str, Any]] = {}
    for candidate in _merge_candidate_pool(visible_candidates):
        code = _feedback_candidate_key(candidate)
        if not code or code == approved_code:
            continue
        rejected_by_code[code] = candidate

    base_entry = {
        "snapshot_id": snapshot_id,
        "job_id": job_id,
        "row_id": str(row.get("id") or "").strip(),
        "query": query,
        "query_key": query_key,
        "structure_keys": structure_keys,
        "order_key_tokens": key_tokens,
        "order_root_tokens": root_tokens,
        "order_dimension_tags": dimension_tags,
        "manual_search_query": clean_search_query,
        "manual_search_query_key": build_search_text(clean_search_query) if clean_search_query else "",
        "selected_via": selected_via,
        "initial_status": str(row.get("initial_status") or "").strip(),
        "selected_by": user.username,
        "selected_by_display": user.display_name,
        "selected_at": selected_at,
        "updated_at": _iso_now(),
    }
    entries = [
        {
            **base_entry,
            "record_id": f"{snapshot_id}:approved:{approved_code}",
            "candidate_code": normalized_approved["code_1c"],
            "candidate_name": normalized_approved["name"],
            "candidate_source_label": normalized_approved.get("source_label") or "",
            "decision": "approved",
        }
    ]
    for code, candidate in rejected_by_code.items():
        entries.append(
            {
                **base_entry,
                "record_id": f"{snapshot_id}:rejected:{code}",
                "candidate_code": candidate["code_1c"],
                "candidate_name": candidate["name"],
                "candidate_source_label": candidate.get("source_label") or "",
                "decision": "rejected",
            }
        )
    return entries


def _feedback_user_for_row(row: dict[str, Any], data: dict[str, Any] | None = None) -> AuthUser:
    username = str(row.get("selected_by") or (data or {}).get("saved_by") or (data or {}).get("created_by") or "").strip().lower()
    display_name = str(row.get("selected_by_display") or (data or {}).get("saved_by_display") or (data or {}).get("created_by_display") or username).strip()
    role = str((data or {}).get("saved_by_role") or (data or {}).get("created_by_role") or "manager").strip().lower() or "manager"
    return AuthUser(username=username, display_name=display_name or username, role=role)


def _save_feedback_snapshot_for_row(
    *,
    job_id: str,
    row: dict[str, Any],
    approved_candidate: dict[str, Any],
    user: AuthUser,
    visible_candidates: list[dict[str, Any]] | None = None,
    previous_approved_candidate: dict[str, Any] | None = None,
    selected_via: str = "",
    search_query: str = "",
    selected_at: str = "",
) -> bool:
    normalized_approved = _normalize_feedback_candidate(approved_candidate)
    if normalized_approved is None:
        return False
    snapshot_id = _snapshot_id(job_id, row)
    effective_selected_at = selected_at or str(row.get("selected_at") or _iso_now())
    entries = _build_feedback_snapshot_entries(
        job_id=job_id,
        row=row,
        approved_candidate=normalized_approved,
        visible_candidates=_merge_candidate_pool(
            row.get("analogs", []),
            visible_candidates or [],
            [previous_approved_candidate] if isinstance(previous_approved_candidate, dict) else [],
        ),
        user=user,
        selected_via=selected_via or _infer_selection_source(row, normalized_approved),
        search_query=search_query or str(row.get("selection_search_query") or ""),
        selected_at=effective_selected_at,
    )
    if not entries:
        return False
    store = _get_learning_store()
    saved_count = store.replace_feedback_snapshot(snapshot_id, entries)
    row["learning_snapshot_id"] = snapshot_id
    row["learning_snapshot_saved"] = bool(saved_count)
    row["learning_snapshot_saved_at"] = _iso_now()
    row["learning_snapshot_entry_count"] = saved_count
    row["learning_snapshot_rejected_count"] = max(0, saved_count - 1)
    _invalidate_learning_caches()
    return bool(saved_count)


def _backfill_feedback_from_saved_jobs(store: LearningStore) -> None:
    if not JOBS_DIR.exists():
        return
    existing_snapshot_ids = {str(entry.get("snapshot_id") or "") for entry in store.load_feedback_entries() if entry.get("snapshot_id")}
    for result_path in JOBS_DIR.glob("*/result.json"):
        try:
            data = json.loads(result_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        job_id = result_path.parent.name
        for row in data.get("rows", []):
            approved = row.get("approved_analog")
            if not isinstance(approved, dict):
                continue
            snapshot_id = _snapshot_id(job_id, row)
            if snapshot_id in existing_snapshot_ids:
                continue
            user = _feedback_user_for_row(row, data)
            if not user.username:
                continue
            entries = _build_feedback_snapshot_entries(
                job_id=job_id,
                row=row,
                approved_candidate=approved,
                visible_candidates=row.get("analogs", []),
                user=user,
                selected_via=_infer_selection_source(row, approved),
                search_query=str(row.get("selection_search_query") or ""),
                selected_at=str(row.get("selected_at") or data.get("saved_at") or data.get("created_at") or _iso_now()),
            )
            if not entries:
                continue
            store.replace_feedback_snapshot(snapshot_id, entries)
            existing_snapshot_ids.add(snapshot_id)


def _get_learning_store() -> LearningStore:
    global _learning_store
    if _learning_store is not None:
        return _learning_store
    store = LearningStore(_learning_db_path())
    legacy_manual_path = _legacy_manual_selection_path()
    manual_payload = _load_manual_selection_payload(legacy_manual_path) if legacy_manual_path.exists() else None
    export_payload = _load_export_audit_payload()
    store.migrate_legacy_payloads(manual_payload=manual_payload, export_payload=export_payload)
    _backfill_feedback_from_saved_jobs(store)
    _learning_store = store
    return store


def _file_content_hash(path: Path) -> str:
    digest = hashlib.sha1()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def _pipeline_cache_key(upload_path: Path) -> str:
    scope = _ensure_request_caches()
    return f"{scope}|upload:{_file_content_hash(upload_path)}"


def _manual_search_cache_key(row: dict[str, Any], query: str, limit: int) -> str:
    scope = _ensure_request_caches()
    row_payload = {
        "name": row.get("name"),
        "mark": row.get("mark"),
        "vendor": row.get("vendor"),
        "unit": row.get("unit"),
        "requested_qty": row.get("requested_qty"),
        "raw_query": row.get("raw_query"),
        "search_text": row.get("search_text"),
        "key_tokens": row.get("key_tokens"),
        "root_tokens": row.get("root_tokens"),
        "dimension_tags": row.get("dimension_tags"),
    }
    row_hash = hashlib.sha1(
        json.dumps(row_payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()
    normalized_query = " ".join(query.lower().split())
    return f"{scope}|row:{row_hash}|query:{normalized_query}|limit:{limit}"


def _load_cached_stock() -> tuple[list[Path], list[Any]]:
    with _stock_cache_lock:
        stock_files = _find_all_stock_files()
        fp = _stock_fingerprint(stock_files)
        if _stock_cache["fingerprint"] == fp and _stock_cache["base_stock"] is not None:
            return _stock_cache["stock_paths"], _stock_cache["base_stock"]

        all_items: list[Any] = []
        paths: list[Path] = []
        for path, label in stock_files:
            items = load_stock(path, source_label=label)
            all_items.extend(items)
            paths.append(path)

        _stock_cache["fingerprint"] = fp
        _stock_cache["base_stock"] = all_items
        _stock_cache["stock_paths"] = paths
        # Invalidate matcher cache when stock changes
        _matcher_cache["master_matcher"] = None
        return paths, all_items


def _load_cached_matcher() -> "StockMatcher":
    """Return the master StockMatcher, building it once per stock version.

    Callers must call .fork() on the result to get a job-local copy with
    independent remaining-quantity tracking.
    """
    with _matcher_cache_lock:
        stock_files = _find_all_stock_files()
        _get_learning_store()
        manual_memory_path = _manual_selection_memory_path()
        fp = f"{_stock_fingerprint(stock_files)}|manual:{_path_fingerprint(manual_memory_path)}"
        if _matcher_cache["fingerprint"] == fp and _matcher_cache["master_matcher"] is not None:
            return _matcher_cache["master_matcher"]

        _, base_stock = _load_cached_stock()
        reviewed_decisions = load_reviewed_analog_decisions(DEFAULT_REVIEWED_ANALOG_DECISIONS_PATH)
        manual_selection_memory = load_manual_selection_memory(manual_memory_path)
        substitution_policy = load_substitution_policy()
        stock_items = clone_stock_items(base_stock)
        master = StockMatcher(
            stock_items,
            reviewed_analog_decisions=reviewed_decisions,
            manual_selection_memory=manual_selection_memory,
            substitution_policy=substitution_policy,
        )
        _matcher_cache["fingerprint"] = fp
        _matcher_cache["master_matcher"] = master
        return master


def _rebuild_status_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        status = row["status"]
        counts[status] = counts.get(status, 0) + 1
    return counts


def _serialize_candidate(analog) -> dict[str, Any]:
    result = {
        "code_1c": analog.stock.code_1c,
        "name": analog.stock.name,
        "score": round(analog.score, 1),
        "stock_qty": analog.stock.quantity,
        "remaining": analog.stock.remaining,
        "price": analog.stock.sale_price,
        "reasons": analog.reasons,
    }
    if getattr(analog, "retrieval_paths", ()):
        result["retrieval_paths"] = list(analog.retrieval_paths)
    if getattr(analog, "feature_scores", None):
        result["feature_scores"] = analog.feature_scores
    if analog.stock.source_label:
        result["source_label"] = analog.stock.source_label
    if getattr(analog, "manual_learning_boost", 0.0) > 0.0:
        result["manager_choice"] = True
    return result


def _build_manual_search_order(row: dict[str, Any], query: str) -> OrderLine:
    # Manual search is intentionally query-first: the manager is searching
    # the stock itself, not asking the matcher to keep the original row
    # constraints. This makes it possible to type any stock code/name and
    # choose it for export.
    name = (query or "").strip()
    mark = ""
    vendor = ""
    unit = (row.get("unit") or "").strip()
    requested_qty = float(row.get("requested_qty") or 0.0)
    query_parts = [name]
    raw_query = " | ".join(part for part in query_parts if part)
    search_text = build_search_text(*query_parts)
    dimension_tags = (
        extract_dimension_tags(*query_parts)
        | extract_family_tags(*query_parts)
        | extract_parser_hint_tags(*query_parts)
        | extract_material_tags_from_search_text(search_text)
    )
    search_text = augment_search_text_with_dimension_tags(search_text, dimension_tags)
    search_tokens = extract_tokens(search_text)
    return OrderLine(
        source_file="manual_search",
        sheet_name="manual_search",
        source_row=0,
        headers=[],
        row_values=[],
        position=str(row.get("position") or ""),
        name=name,
        mark=mark,
        supplier_code="",
        vendor=vendor,
        unit=unit,
        requested_qty=requested_qty,
        search_text=search_text,
        search_tokens=search_tokens,
        key_tokens=extract_key_tokens(search_tokens),
        root_tokens=extract_root_tokens(search_tokens),
        code_tokens=extract_code_tokens(*query_parts),
        dimension_tags=dimension_tags,
        raw_query=raw_query or name,
        classification=None,
    )


def _build_row_context_order(row: dict[str, Any]) -> OrderLine:
    name = str(row.get("name") or "").strip()
    mark = str(row.get("mark") or "").strip()
    vendor = str(row.get("vendor") or "").strip()
    unit = str(row.get("unit") or "").strip()
    requested_qty = float(row.get("requested_qty") or 0.0)
    raw_query = str(row.get("raw_query") or "").strip()
    search_text = str(row.get("search_text") or "").strip()
    if not search_text:
        search_text = build_search_text(name, mark, vendor)
    dimension_tags = {str(tag) for tag in row.get("dimension_tags") or [] if str(tag).strip()}
    if not dimension_tags:
        dimension_tags = (
            extract_dimension_tags(name, mark, vendor)
            | extract_family_tags(name, mark, vendor)
            | extract_parser_hint_tags(name, mark, vendor)
            | extract_material_tags_from_search_text(search_text)
        )
        search_text = augment_search_text_with_dimension_tags(search_text, dimension_tags)
    search_tokens = extract_tokens(search_text)
    key_tokens = {str(token) for token in row.get("key_tokens") or [] if str(token).strip()} or extract_key_tokens(search_tokens)
    root_tokens = {str(token) for token in row.get("root_tokens") or [] if str(token).strip()} or extract_root_tokens(search_tokens)
    return OrderLine(
        source_file="manual_search_context",
        sheet_name="manual_search_context",
        source_row=0,
        headers=[],
        row_values=[],
        position=str(row.get("position") or ""),
        name=name,
        mark=mark,
        supplier_code="",
        vendor=vendor,
        unit=unit,
        requested_qty=requested_qty,
        search_text=search_text,
        search_tokens=search_tokens,
        key_tokens=key_tokens,
        root_tokens=root_tokens,
        code_tokens=extract_code_tokens(name, mark, vendor),
        dimension_tags=dimension_tags,
        raw_query=raw_query or name,
        classification=None,
    )


def _normalize_manual_code(value: str) -> str:
    return re.sub(r"[^0-9a-zа-я]+", "", value.lower())


def _merge_manual_search_pool(
    candidate_sources: dict[int, set[str]],
    pool: dict[int, tuple[str, ...]],
    marker: str,
) -> None:
    for stock_index, paths in pool.items():
        current = candidate_sources.setdefault(int(stock_index), set())
        current.add(marker)
        current.update(str(path) for path in paths if str(path).strip())


def _build_manual_search_candidate_pool(
    matcher: StockMatcher,
    manual_order: OrderLine,
    context_order: OrderLine,
    query: str,
) -> dict[int, tuple[str, ...]]:
    candidate_sources: dict[int, set[str]] = {}
    _merge_manual_search_pool(candidate_sources, matcher.generate_candidate_pool(manual_order), "manual_query_pool")
    _merge_manual_search_pool(candidate_sources, matcher.generate_candidate_pool(context_order), "row_context_pool")

    query_lower = query.strip().lower()
    normalized_code_query = _normalize_manual_code(query)
    if normalized_code_query:
        for stock_index in matcher.normalized_code_index.get(normalized_code_query, []):
            candidate_sources.setdefault(int(stock_index), set()).add("manual_exact_code")

    query_tokens = extract_tokens(build_search_text(query))
    informative_tokens = sorted(
        query_tokens,
        key=lambda token: (len(matcher.search_token_index.get(token, [])) or 100000, -len(token), token),
    )
    for token in informative_tokens[:MANUAL_SEARCH_QUERY_TOKEN_LIMIT]:
        matches = matcher.search_token_index.get(token, [])
        if not matches:
            continue
        for stock_index in matches:
            candidate_sources.setdefault(int(stock_index), set()).add("manual_query_token")

    if query_lower and (len(query_lower) >= 4 or normalized_code_query):
        fragment_hits = 0
        for stock_index, stock in enumerate(matcher.stock_items):
            stock_code_lower = (stock.code_1c or "").lower()
            stock_name_lower = (stock.name or "").lower()
            normalized_stock_code = _normalize_manual_code(stock.code_1c or "")
            if (
                stock_code_lower == query_lower
                or (normalized_code_query and normalized_stock_code == normalized_code_query)
                or query_lower in stock_code_lower
                or query_lower in stock_name_lower
            ):
                candidate_sources.setdefault(int(stock_index), set()).add("manual_fragment")
                fragment_hits += 1
                if fragment_hits >= MANUAL_SEARCH_MAX_FRAGMENT_MATCHES:
                    break

    ranked_pool = sorted(
        candidate_sources.items(),
        key=lambda item: (
            "manual_exact_code" in item[1],
            "manual_fragment" in item[1],
            "memory" in item[1],
            "code" in item[1],
            "family" in item[1],
            "structure" in item[1],
            len(item[1]),
            matcher.stock_items[item[0]].remaining > 0,
            matcher.stock_items[item[0]].remaining,
        ),
        reverse=True,
    )
    return {
        stock_index: tuple(sorted(paths))
        for stock_index, paths in ranked_pool[:MANUAL_SEARCH_MAX_SCORE_CANDIDATES]
    }


def _manual_search_candidates(
    matcher: StockMatcher,
    row: dict[str, Any],
    query: str,
    limit: int,
) -> list[Candidate]:
    manual_order = _build_manual_search_order(row, query)
    context_order = _build_row_context_order(row)
    query_lower = query.strip().lower()
    normalized_code_query = _normalize_manual_code(query)
    candidate_pool = _build_manual_search_candidate_pool(matcher, manual_order, context_order, query)
    ranked: list[Candidate] = []

    for stock_index, retrieval_paths in candidate_pool.items():
        stock = matcher.stock_items[int(stock_index)]
        query_candidate = matcher.score_candidate(manual_order, stock, retrieval_paths=retrieval_paths)
        context_candidate = matcher.score_candidate(context_order, stock, retrieval_paths=retrieval_paths)
        reasons = list(query_candidate.reasons)
        score = query_candidate.score

        stock_code_lower = (stock.code_1c or "").lower()
        stock_name_lower = (stock.name or "").lower()
        normalized_stock_code = _normalize_manual_code(stock.code_1c or "")

        query_weight = 0.62 if len(query_lower) >= 6 else 0.54
        context_weight = 1.0 - query_weight
        score = max(score, query_candidate.score * query_weight + context_candidate.score * context_weight)

        context_family_match = float(context_candidate.feature_scores.get("family_match", 0.0))
        context_material_match = float(context_candidate.feature_scores.get("material_match", 0.0))
        context_structure_overlap = float(context_candidate.feature_scores.get("structure_overlap", 0.0))
        if context_family_match > 0.0 and context_material_match > 0.0:
            score += 8.0
            reasons.append("совпадает тип и материал с заявкой")
        elif context_family_match > 0.0:
            score += 5.0
            reasons.append("совпадает тип изделия с заявкой")
        if context_candidate.dimension_bonus >= 10.0:
            score += 10.0
            reasons.append("совпадают размеры с заявкой")
        elif context_candidate.dimension_bonus >= 6.0:
            score += 6.0
            reasons.append("частично совпадают размеры с заявкой")
        if context_structure_overlap >= 0.25:
            score += 4.0
            reasons.append("учтена структура исходной строки")

        if query_lower and stock_code_lower == query_lower:
            score = max(score, 100.0)
            reasons.append("точное совпадение по коду 1С")
        elif normalized_code_query and normalized_stock_code and normalized_code_query == normalized_stock_code:
            score = max(score, 98.0)
            reasons.append("совпадает нормализованный код")
        elif query_lower and (query_lower in stock_code_lower or query_lower in stock_name_lower):
            score = max(score, min(96.0, score + 20.0))
            reasons.append("совпадение по фрагменту кода/названия")

        retrieval_paths = tuple(
            dict.fromkeys(
                [
                    *getattr(query_candidate, "retrieval_paths", ()),
                    *getattr(context_candidate, "retrieval_paths", ()),
                    "manual_query",
                    "manual_context",
                ]
            )
        )
        feature_scores = dict(getattr(query_candidate, "feature_scores", {}) or {})
        feature_scores.update(
            {
                "manual_query_score": round(query_candidate.score, 3),
                "row_context_score": round(context_candidate.score, 3),
                "row_context_overlap": round(context_candidate.overlap, 4),
                "row_context_soft_overlap": round(context_candidate.soft_overlap, 4),
                "row_context_dimension_bonus": round(context_candidate.dimension_bonus, 3),
                "row_context_family_match": context_family_match,
                "row_context_material_match": context_material_match,
            }
        )
        ranked.append(
            Candidate(
                stock=query_candidate.stock,
                score=max(0.0, min(100.0, score)),
                overlap=max(query_candidate.overlap, context_candidate.overlap),
                reasons=tuple(dict.fromkeys(reasons)),
                code_hit=bool(query_candidate.code_hit or (normalized_code_query and normalized_stock_code == normalized_code_query)),
                dimension_bonus=max(query_candidate.dimension_bonus, context_candidate.dimension_bonus),
                dimension_penalty=min(query_candidate.dimension_penalty, context_candidate.dimension_penalty),
                soft_overlap=max(query_candidate.soft_overlap, context_candidate.soft_overlap),
                matched_dimension_keys=tuple(dict.fromkeys([*query_candidate.matched_dimension_keys, *context_candidate.matched_dimension_keys])),
                conflicting_dimension_keys=tuple(dict.fromkeys(query_candidate.conflicting_dimension_keys)),
                review_decision=query_candidate.review_decision,
                retrieval_paths=retrieval_paths,
                feature_scores=feature_scores,
            )
        )

    ranked.sort(
        key=lambda candidate: (
            candidate.score,
            candidate.code_hit,
            candidate.stock.remaining > 0,
            candidate.stock.remaining,
            -candidate.stock.row_index,
        ),
        reverse=True,
    )

    filtered: list[Candidate] = []
    seen_codes: set[str] = set()
    for candidate in ranked:
        code = candidate.stock.code_1c
        if not code or code in seen_codes:
            continue
        if (
            candidate.score < MANUAL_SEARCH_MIN_SCORE
            and not candidate.code_hit
            and candidate.overlap < MANUAL_SEARCH_MIN_OVERLAP
            and candidate.soft_overlap < MANUAL_SEARCH_MIN_SOFT_OVERLAP
            and candidate.dimension_bonus < MANUAL_SEARCH_MIN_DIMENSION_BONUS
            and float(candidate.feature_scores.get("row_context_score", 0.0)) < 44.0
            and float(candidate.feature_scores.get("row_context_dimension_bonus", 0.0)) < 6.0
        ):
            continue
        filtered.append(candidate)
        seen_codes.add(code)
        if len(filtered) >= limit:
            break
    return filtered


def _job_is_editable(data: dict[str, Any]) -> bool:
    return not data.get("saved_at") and int(data.get("export_count", 0) or 0) == 0


def _parse_requested_quantity(value: object) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        raise HTTPException(400, "Количество должно быть числом")
    if not math.isfinite(number) or number <= 0:
        raise HTTPException(400, "Количество должно быть больше нуля")
    return round(number, 3)


def _row_max_editable_quantity(row: dict[str, Any]) -> float | None:
    approved = row.get("approved_analog")
    if isinstance(approved, dict):
        for key in ("remaining", "stock_qty"):
            value = approved.get(key)
            try:
                number = float(value)
            except (TypeError, ValueError):
                continue
            if math.isfinite(number) and number > 0:
                return number
    if row.get("matched_code"):
        for key in ("matched_stock_qty", "available_qty"):
            value = row.get(key)
            try:
                number = float(value)
            except (TypeError, ValueError):
                continue
            if math.isfinite(number) and number > 0:
                return number
    return None


def _update_row_quantity_inplace(row: dict[str, Any], quantity: float) -> None:
    max_quantity = _row_max_editable_quantity(row)
    if max_quantity is not None and quantity > max_quantity + 1e-9:
        raise HTTPException(400, f"Нельзя поставить количество больше доступного остатка: {max_quantity:g}")

    row["requested_qty"] = quantity
    if row.get("approved_analog"):
        return
    if row.get("matched_code"):
        row["available_qty"] = quantity if max_quantity is None else min(quantity, max_quantity)
        if row.get("status") in {"Найдено полностью", "Найдено частично"} and max_quantity is not None and quantity <= max_quantity:
            row["status"] = "Найдено полностью"


def _find_row(data: dict[str, Any], row_id: str) -> dict[str, Any]:
    for row in data["rows"]:
        if row["id"] == row_id:
            return row
    raise HTTPException(404, f"Row {row_id} not found")


def _run_pipeline(upload_path: Path, job_dir: Path) -> dict[str, Any]:
    """Run the full normalization + matching pipeline."""
    normalized_dir = ensure_output_dir(job_dir / "normalized")
    matched_dir = ensure_output_dir(job_dir / "matched")

    # Stage 1: normalize
    normalized_path, parsed_count, issue_count = normalize_request_file(upload_path, normalized_dir)

    # Stage 2: match — fork the cached master matcher (indexes are pre-built)
    stock_paths, _ = _load_cached_stock()
    matcher = _load_cached_matcher().fork()
    order_lines = load_order_lines(normalized_path)
    order_results = match_orders(order_lines, matcher)

    output_paths = write_outputs(
        order_path=normalized_path,
        stock_path=stock_paths[0],
        reviewed_decisions_path=DEFAULT_REVIEWED_ANALOG_DECISIONS_PATH,
        order_results=order_results,
        stock_items=matcher.stock_items,
        out_dir=matched_dir,
    )

    # Build JSON result for the web UI
    rows: list[dict[str, Any]] = []
    for result in order_results:
        analogs_list = []
        for analog in result.analogs:
            payload = _serialize_candidate(analog)
            if analog.stock.code_1c == (result.matched_stock.code_1c if result.matched_stock else None):
                payload["remaining"] = payload["remaining"] + result.available_qty
            analogs_list.append(payload)

        row: dict[str, Any] = {
            "id": str(uuid.uuid4())[:8],
            "position": result.order.position,
            "name": result.order.name,
            "mark": result.order.mark,
            "vendor": result.order.vendor,
            "unit": result.order.unit,
            "requested_qty": result.order.requested_qty,
            "status": result.status,
            "confidence": round(result.confidence, 1),
            "comment": result.comment,
            "matched_code": result.matched_stock.code_1c if result.matched_stock else None,
            "matched_name": result.matched_stock.name if result.matched_stock else None,
            "matched_price": result.matched_stock.sale_price if result.matched_stock else None,
            "matched_source_label": result.matched_stock.source_label if result.matched_stock and result.matched_stock.source_label else None,
            "matched_stock_qty": result.matched_stock.quantity if result.matched_stock else None,
            "available_qty": result.available_qty,
            "analogs": analogs_list,
            "approved_analog": None,  # will be set by manager
            "initial_status": result.status,
            "raw_query": result.order.raw_query,
            "search_text": result.order.search_text,
            "key_tokens": sorted(result.order.key_tokens),
            "root_tokens": sorted(result.order.root_tokens),
            "dimension_tags": sorted(result.order.dimension_tags),
        }
        rows.append(row)

    # Status counts
    status_counts: dict[str, int] = {}
    for r in rows:
        s = r["status"]
        status_counts[s] = status_counts.get(s, 0) + 1

    result_data = {
        "total_rows": len(rows),
        "status_counts": status_counts,
        "rows": rows,
        "stock_files": [p.name for p in stock_paths],
        "parsed_count": parsed_count,
        "issue_count": issue_count,
        "output_files": [str(p) for p in output_paths],
    }

    # Save result JSON
    result_path = job_dir / "result.json"
    result_path.write_text(json.dumps(result_data, ensure_ascii=False, indent=2), encoding="utf-8")

    return result_data


# ---------------------------------------------------------------------------
# Job storage
# ---------------------------------------------------------------------------

# In-memory job index (job_id -> metadata)
_jobs: dict[str, dict[str, Any]] = {}

JOB_STATUS_PROCESSING = "processing"
JOB_STATUS_COMPLETED = "completed"
JOB_STATUS_ERROR = "error"


def _load_job(job_id: str) -> dict[str, Any]:
    job_dir = JOBS_DIR / job_id
    result_path = job_dir / "result.json"
    if not result_path.exists():
        raise HTTPException(404, f"Job {job_id} not found")
    return json.loads(result_path.read_text(encoding="utf-8"))


def _save_job(job_id: str, data: dict[str, Any]) -> None:
    job_dir = JOBS_DIR / job_id
    result_path = job_dir / "result.json"
    result_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _build_processing_job_data(job_id: str, filename: str, user: AuthUser, created_at: float) -> dict[str, Any]:
    return {
        "job_id": job_id,
        "filename": filename,
        "created_at": created_at,
        "created_by": user.username,
        "created_by_display": user.display_name,
        "created_by_role": user.role,
        "saved_at": None,
        "saved_by": "",
        "saved_by_display": "",
        "learning_saved_count": 0,
        "export_count": 0,
        "job_status": JOB_STATUS_PROCESSING,
        "progress_message": "Файл загружен. Идёт анализ, это может занять несколько минут...",
        "error_message": "",
        "total_rows": 0,
        "status_counts": {},
        "rows": [],
        "stock_files": [],
        "parsed_count": 0,
        "issue_count": 0,
        "output_files": [],
    }


def _queued_progress_message(recovered: bool = False) -> str:
    if recovered:
        return "Восстанавливаем обработку файла после перезапуска сервиса..."
    return "Файл загружен. Ожидает очередь на анализ..."


def _finalize_job_data(base_data: dict[str, Any], result: dict[str, Any]) -> dict[str, Any]:
    finalized = copy.deepcopy(result)
    finalized["job_id"] = base_data["job_id"]
    finalized["filename"] = base_data["filename"]
    finalized["created_at"] = base_data["created_at"]
    finalized["created_by"] = base_data["created_by"]
    finalized["created_by_display"] = base_data["created_by_display"]
    finalized["created_by_role"] = base_data["created_by_role"]
    finalized["saved_at"] = base_data.get("saved_at")
    finalized["saved_by"] = base_data.get("saved_by", "")
    finalized["saved_by_display"] = base_data.get("saved_by_display", "")
    finalized["learning_saved_count"] = int(base_data.get("learning_saved_count", 0) or 0)
    finalized["export_count"] = int(base_data.get("export_count", 0) or 0)
    finalized["job_status"] = JOB_STATUS_COMPLETED
    finalized["progress_message"] = ""
    finalized["error_message"] = ""
    return finalized


def _store_job_state(job_id: str, data: dict[str, Any]) -> None:
    _save_job(job_id, data)
    _jobs[job_id] = _build_job_summary(job_id, data, float(data.get("created_at") or time.time()))


def _build_job_summary(job_id: str, data: dict[str, Any], created_at: float) -> dict[str, Any]:
    replacements_count = sum(1 for row in data.get("rows", []) if row.get("approved_analog"))
    return {
        "job_id": job_id,
        "filename": data.get("filename") or job_id,
        "created_at": float(data.get("created_at") or created_at),
        "created_by": data.get("created_by") or "",
        "created_by_display": data.get("created_by_display") or "",
        "saved_at": data.get("saved_at"),
        "saved_by": data.get("saved_by") or "",
        "saved_by_display": data.get("saved_by_display") or "",
        "total_rows": int(data.get("total_rows", 0) or 0),
        "status_counts": data.get("status_counts", {}),
        "replacements_count": replacements_count,
        "learning_saved_count": int(data.get("learning_saved_count", 0) or 0),
        "export_count": int(data.get("export_count", 0) or 0),
        "job_status": data.get("job_status") or JOB_STATUS_COMPLETED,
        "progress_message": data.get("progress_message") or "",
        "error_message": data.get("error_message") or "",
    }


def _load_export_audit_payload(path: Path | None = None) -> dict[str, Any]:
    target = path or _export_audit_path()
    if not target.exists():
        return {
            "version": 1,
            "updated_at": "",
            "event_count": 0,
            "exports": [],
        }
    try:
        payload = json.loads(target.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {
            "version": 1,
            "updated_at": "",
            "event_count": 0,
            "exports": [],
        }
    if not isinstance(payload.get("exports"), list):
        payload["exports"] = []
    payload["version"] = int(payload.get("version", 1) or 1)
    payload["updated_at"] = str(payload.get("updated_at", ""))
    payload["event_count"] = len(payload["exports"])
    return payload


def _save_export_audit_payload(payload: dict[str, Any], path: Path | None = None) -> None:
    target = path or _export_audit_path()
    target.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = target.with_suffix(f"{target.suffix}.tmp")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(target)


def _build_export_audit_event(
    *,
    job_id: str,
    data: dict[str, Any],
    user: AuthUser,
    learning_saved_count: int,
) -> dict[str, Any]:
    export_sequence = int(data.get("export_count", 0) or 0)
    saved_at = str(data.get("saved_at") or _iso_now())
    replacements: list[dict[str, Any]] = []
    manual_search_replacements = 0
    learned_replacements = 0
    for row in data.get("rows", []):
        approved = row.get("approved_analog")
        if not isinstance(approved, dict):
            continue
        selected_via = str(row.get("selected_via") or "")
        initial_status = str(row.get("initial_status") or "")
        learned = bool(row.get("learning_snapshot_saved")) or (
            selected_via == "manual_search" and initial_status == STATUS_NOT_FOUND
        )
        if selected_via == "manual_search":
            manual_search_replacements += 1
        if learned:
            learned_replacements += 1
        replacements.append(
            {
                "row_id": row.get("id"),
                "position": row.get("position"),
                "name": row.get("name"),
                "mark": row.get("mark"),
                "vendor": row.get("vendor"),
                "requested_qty": row.get("requested_qty"),
                "initial_status": initial_status,
                "final_status": row.get("status"),
                "selected_via": selected_via,
                "selection_search_query": row.get("selection_search_query") or "",
                "selected_by": row.get("selected_by") or user.username,
                "selected_by_display": row.get("selected_by_display") or user.display_name,
                "selected_at": row.get("selected_at") or saved_at,
                "candidate_code": approved.get("code_1c") or "",
                "candidate_name": approved.get("name") or "",
                "candidate_source_label": approved.get("source_label") or "",
                "candidate_score": approved.get("score"),
                "manager_choice": bool(approved.get("manager_choice")),
                "reasons": list(approved.get("reasons") or []),
                "learning_snapshot_id": row.get("learning_snapshot_id") or "",
                "learned_on_export": learned,
            }
        )
    return {
        "export_event_id": f"{job_id}:export:{export_sequence}",
        "job_id": job_id,
        "filename": data.get("filename") or job_id,
        "saved_at": saved_at,
        "saved_by": user.username,
        "saved_by_display": user.display_name,
        "saved_by_role": user.role,
        "total_rows": int(data.get("total_rows", 0) or 0),
        "status_counts": data.get("status_counts", {}),
        "replacement_count": len(replacements),
        "manual_search_replacement_count": manual_search_replacements,
        "learned_replacement_count": learned_replacements,
        "learning_saved_count": learning_saved_count,
        "replacements": replacements,
    }


def _append_export_audit_event(job_id: str, data: dict[str, Any], user: AuthUser, learning_saved_count: int) -> dict[str, Any]:
    event = _build_export_audit_event(
        job_id=job_id,
        data=data,
        user=user,
        learning_saved_count=learning_saved_count,
    )
    _get_learning_store().replace_export_event(event)
    return event


def _build_admin_analytics(
    payload: dict[str, Any],
    users: dict[str, dict[str, str]],
    feedback_entries: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    exports = sorted(
        [entry for entry in payload.get("exports", []) if isinstance(entry, dict)],
        key=lambda item: str(item.get("saved_at") or ""),
        reverse=True,
    )
    feedback_entries = [entry for entry in (feedback_entries or []) if isinstance(entry, dict)]
    approved_snapshots: dict[str, dict[str, Any]] = {}
    for entry in feedback_entries:
        if str(entry.get("decision") or "").strip().lower() != "approved":
            continue
        snapshot_id = str(entry.get("snapshot_id") or entry.get("record_id") or "").strip()
        if not snapshot_id or snapshot_id in approved_snapshots:
            continue
        approved_snapshots[snapshot_id] = entry
    summary = {
        "saved_files": len(exports),
        "unique_jobs": len({entry.get("job_id") for entry in exports if entry.get("job_id")}),
        "total_rows": sum(int(entry.get("total_rows", 0) or 0) for entry in exports),
        "replacement_count": sum(int(entry.get("replacement_count", 0) or 0) for entry in exports),
        "manual_search_replacement_count": sum(int(entry.get("manual_search_replacement_count", 0) or 0) for entry in exports),
        "learned_replacement_count": (
            len(approved_snapshots)
            if approved_snapshots
            else sum(int(entry.get("learned_replacement_count", 0) or 0) for entry in exports)
        ),
        "users_with_exports": len({entry.get("saved_by") for entry in exports if entry.get("saved_by")}),
    }

    user_stats: dict[str, dict[str, Any]] = {}
    for username, raw_user in users.items():
        user_stats[username] = {
            "username": username,
            "display_name": raw_user.get("display_name") or username,
            "role": raw_user.get("role") or "manager",
            "saved_files": 0,
            "unique_jobs": 0,
            "total_rows": 0,
            "replacement_count": 0,
            "manual_search_replacement_count": 0,
            "learned_replacement_count": 0,
            "last_saved_at": "",
            "files": [],
            "_job_ids": set(),
        }

    replacement_popularity: dict[str, dict[str, Any]] = {}
    for entry in exports:
        username = str(entry.get("saved_by") or "").strip().lower()
        if not username:
            continue
        stat = user_stats.setdefault(
            username,
            {
                "username": username,
                "display_name": entry.get("saved_by_display") or username,
                "role": entry.get("saved_by_role") or "manager",
                "saved_files": 0,
                "unique_jobs": 0,
                "total_rows": 0,
                "replacement_count": 0,
                "manual_search_replacement_count": 0,
                "learned_replacement_count": 0,
                "last_saved_at": "",
                "files": [],
                "_job_ids": set(),
            },
        )
        stat["saved_files"] += 1
        stat["total_rows"] += int(entry.get("total_rows", 0) or 0)
        stat["replacement_count"] += int(entry.get("replacement_count", 0) or 0)
        stat["manual_search_replacement_count"] += int(entry.get("manual_search_replacement_count", 0) or 0)
        stat["last_saved_at"] = max(stat["last_saved_at"], str(entry.get("saved_at") or ""))
        job_id = str(entry.get("job_id") or "")
        if job_id:
            stat["_job_ids"].add(job_id)
        stat["files"].append(
            {
                "job_id": job_id,
                "filename": entry.get("filename") or "",
                "saved_at": entry.get("saved_at") or "",
                "total_rows": int(entry.get("total_rows", 0) or 0),
                "replacement_count": int(entry.get("replacement_count", 0) or 0),
                "learned_replacement_count": int(entry.get("learned_replacement_count", 0) or 0),
            }
        )

        for replacement in entry.get("replacements", []):
            if not isinstance(replacement, dict):
                continue
            code = str(replacement.get("candidate_code") or "").strip()
            name = str(replacement.get("candidate_name") or "").strip()
            if not code or not name:
                continue
            item = replacement_popularity.setdefault(
                code,
                {
                    "candidate_code": code,
                    "candidate_name": name,
                    "times_used": 0,
                    "times_learned": 0,
                    "last_used_at": "",
                    "used_by": set(),
                },
            )
            item["times_used"] += 1
            item["last_used_at"] = max(item["last_used_at"], str(entry.get("saved_at") or ""))
            item["used_by"].add(username)

    for entry in approved_snapshots.values():
        username = str(entry.get("selected_by") or "").strip().lower()
        if not username:
            continue
        stat = user_stats.setdefault(
            username,
            {
                "username": username,
                "display_name": entry.get("selected_by_display") or username,
                "role": "manager",
                "saved_files": 0,
                "unique_jobs": 0,
                "total_rows": 0,
                "replacement_count": 0,
                "manual_search_replacement_count": 0,
                "learned_replacement_count": 0,
                "last_saved_at": "",
                "files": [],
                "_job_ids": set(),
            },
        )
        stat["learned_replacement_count"] += 1
        code = str(entry.get("candidate_code") or "").strip()
        name = str(entry.get("candidate_name") or "").strip()
        if not code or not name:
            continue
        item = replacement_popularity.setdefault(
            code,
            {
                "candidate_code": code,
                "candidate_name": name,
                "times_used": 0,
                "times_learned": 0,
                "last_used_at": "",
                "used_by": set(),
            },
        )
        item["times_learned"] += 1
        item["last_used_at"] = max(item["last_used_at"], str(entry.get("selected_at") or ""))
        item["used_by"].add(username)

    if not approved_snapshots:
        for entry in exports:
            username = str(entry.get("saved_by") or "").strip().lower()
            if username and username in user_stats:
                user_stats[username]["learned_replacement_count"] += int(entry.get("learned_replacement_count", 0) or 0)
            for replacement in entry.get("replacements", []):
                if not isinstance(replacement, dict) or not replacement.get("learned_on_export"):
                    continue
                code = str(replacement.get("candidate_code") or "").strip()
                name = str(replacement.get("candidate_name") or "").strip()
                if not code or not name:
                    continue
                item = replacement_popularity.setdefault(
                    code,
                    {
                        "candidate_code": code,
                        "candidate_name": name,
                        "times_used": 0,
                        "times_learned": 0,
                        "last_used_at": "",
                        "used_by": set(),
                    },
                )
                item["times_learned"] += 1
                item["last_used_at"] = max(item["last_used_at"], str(entry.get("saved_at") or ""))
                if username:
                    item["used_by"].add(username)

    prepared_users: list[dict[str, Any]] = []
    for stat in user_stats.values():
        stat["unique_jobs"] = len(stat.pop("_job_ids"))
        stat["files"].sort(key=lambda item: item.get("saved_at") or "", reverse=True)
        prepared_users.append(stat)
    prepared_users.sort(
        key=lambda item: (
            int(item.get("saved_files", 0)),
            int(item.get("replacement_count", 0)),
            item.get("display_name") or item.get("username") or "",
        ),
        reverse=True,
    )

    top_replacements = sorted(
        (
            {
                "candidate_code": item["candidate_code"],
                "candidate_name": item["candidate_name"],
                "times_used": item["times_used"],
                "times_learned": item["times_learned"],
                "last_used_at": item["last_used_at"],
                "used_by_count": len(item["used_by"]),
            }
            for item in replacement_popularity.values()
        ),
        key=lambda item: (item["times_used"], item["times_learned"], item["candidate_name"]),
        reverse=True,
    )[:20]

    return {
        "summary": summary,
        "users": prepared_users,
        "exports": exports,
        "top_replacements": top_replacements,
    }


def _load_manual_selection_payload(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {
            "version": 1,
            "entry_count": 0,
            "updated_at": "",
            "entries": [],
        }
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {
            "version": 1,
            "entry_count": 0,
            "updated_at": "",
            "entries": [],
        }
    if not isinstance(payload.get("entries"), list):
        payload["entries"] = []
    payload["version"] = int(payload.get("version", 1) or 1)
    payload["entry_count"] = len(payload["entries"])
    payload["updated_at"] = str(payload.get("updated_at", ""))
    return payload


def _save_manual_selection_payload(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(f"{path.suffix}.tmp")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(path)


def _record_manual_selection_learning(
    *,
    job_id: str,
    row: dict[str, Any],
    candidate: dict[str, Any],
    search_query: str,
) -> bool:
    user = AuthUser(
        username=str(row.get("selected_by") or "").strip().lower(),
        display_name=str(row.get("selected_by_display") or "").strip(),
        role="manager",
    )
    if not user.username:
        return False
    return _save_feedback_snapshot_for_row(
        job_id=job_id,
        row=row,
        approved_candidate=candidate,
        user=user,
        selected_via="manual_search",
        search_query=search_query,
    )


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


def _process_uploaded_job(
    job_id: str,
    upload_path: Path,
    job_dir: Path,
    base_data: dict[str, Any],
    cache_key: str,
) -> None:
    try:
        cached_result = _pipeline_result_cache["entries"].get(cache_key)
        if cached_result is not None:
            result = copy.deepcopy(cached_result)
        else:
            result = _run_pipeline(upload_path, job_dir)
            _pipeline_result_cache["entries"][cache_key] = copy.deepcopy(result)
        _store_job_state(job_id, _finalize_job_data(base_data, result))
    except Exception as exc:
        error_data = copy.deepcopy(base_data)
        error_data["job_status"] = JOB_STATUS_ERROR
        error_data["progress_message"] = "Ошибка обработки"
        error_data["error_message"] = str(exc)
        _store_job_state(job_id, error_data)


def _job_worker_is_busy() -> bool:
    with _job_queue_state_lock:
        return _active_job_id is not None or bool(_queued_job_ids)


def _enqueue_processing_job(
    job_id: str,
    upload_path: Path,
    job_dir: Path,
    base_data: dict[str, Any],
    cache_key: str,
) -> None:
    with _job_queue_state_lock:
        if job_id == _active_job_id or job_id in _queued_job_ids:
            return
        _queued_job_ids.add(job_id)
    _job_queue.put((job_id, upload_path, job_dir, copy.deepcopy(base_data), cache_key))


def _job_worker_loop() -> None:
    global _active_job_id
    while True:
        job_id, upload_path, job_dir, base_data, cache_key = _job_queue.get()
        try:
            with _job_queue_state_lock:
                _queued_job_ids.discard(job_id)
                _active_job_id = job_id
            processing_data = copy.deepcopy(base_data)
            processing_data["job_status"] = JOB_STATUS_PROCESSING
            processing_data["progress_message"] = "Файл загружен. Идёт анализ, это может занять несколько минут..."
            processing_data["error_message"] = ""
            _store_job_state(job_id, processing_data)
            _process_uploaded_job(job_id, upload_path, job_dir, processing_data, cache_key)
        finally:
            with _job_queue_state_lock:
                if _active_job_id == job_id:
                    _active_job_id = None
            _job_queue.task_done()


def _start_job_worker_if_needed() -> None:
    global _job_worker_thread
    with _job_queue_state_lock:
        if _job_worker_thread is not None and _job_worker_thread.is_alive():
            return
        _job_worker_thread = threading.Thread(
            target=_job_worker_loop,
            daemon=True,
            name="upload-job-worker",
        )
        _job_worker_thread.start()


def _recover_processing_jobs_on_startup() -> None:
    if not JOBS_DIR.exists():
        return
    for job_dir in sorted((path for path in JOBS_DIR.iterdir() if path.is_dir()), key=lambda path: path.name):
        result_path = job_dir / "result.json"
        if not result_path.exists():
            continue
        try:
            data = json.loads(result_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if data.get("job_status") != JOB_STATUS_PROCESSING:
            continue
        filename = str(data.get("filename") or "").strip()
        if not filename:
            continue
        upload_path = job_dir / filename
        if not upload_path.exists():
            error_data = copy.deepcopy(data)
            error_data["job_status"] = JOB_STATUS_ERROR
            error_data["progress_message"] = "Ошибка обработки"
            error_data["error_message"] = "Файл загрузки не найден для восстановления задачи"
            _store_job_state(job_dir.name, error_data)
            continue
        recovery_data = copy.deepcopy(data)
        recovery_data["progress_message"] = _queued_progress_message(recovered=True)
        recovery_data["error_message"] = ""
        _store_job_state(job_dir.name, recovery_data)
        _enqueue_processing_job(
            job_dir.name,
            upload_path,
            job_dir,
            recovery_data,
            _pipeline_cache_key(upload_path),
        )


def _warm_matcher_cache_background() -> None:
    try:
        _load_cached_matcher()
    except Exception:
        pass


@app.post("/api/upload")
async def upload_file(
    file: UploadFile = File(...),
    authorization: str | None = Header(None),
    token: str | None = Query(None),
):
    user = _check_auth(authorization, token)
    created_at = time.time()
    job_id = f"{int(time.time())}_{uuid.uuid4().hex[:6]}"
    job_dir = ensure_output_dir(JOBS_DIR / job_id)

    # Save uploaded file
    upload_path = job_dir / file.filename
    with open(upload_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    cache_key = _pipeline_cache_key(upload_path)
    base_data = _build_processing_job_data(job_id, file.filename, user, created_at)
    if _job_worker_is_busy():
        base_data["progress_message"] = _queued_progress_message()
    _store_job_state(job_id, base_data)
    _start_job_worker_if_needed()
    _enqueue_processing_job(job_id, upload_path, job_dir, base_data, cache_key)

    return copy.deepcopy(base_data)


@app.get("/api/jobs")
def list_jobs(authorization: str | None = Header(None), token: str | None = Query(None)):
    user = _check_auth(authorization, token)
    # Also scan disk for jobs not in memory
    if JOBS_DIR.exists():
        for d in sorted(JOBS_DIR.iterdir(), reverse=True):
            if d.is_dir() and d.name not in _jobs:
                result_path = d / "result.json"
                if result_path.exists():
                    data = json.loads(result_path.read_text(encoding="utf-8"))
                    _jobs[d.name] = _build_job_summary(d.name, data, d.stat().st_mtime)
    visible_jobs = [
        job
        for job in _jobs.values()
        if user.role == "admin" or str(job.get("created_by") or "").strip().lower() == user.username
    ]
    return {"jobs": sorted(visible_jobs, key=lambda j: float(j.get("created_at") or 0), reverse=True)}


@app.get("/api/jobs/{job_id}")
def get_job(job_id: str, authorization: str | None = Header(None), token: str | None = Query(None)):
    user = _check_auth(authorization, token)
    data = _load_job(job_id)
    _ensure_job_access(user, data)
    return data


class ApproveRequest(BaseModel):
    approvals: dict[str, str]  # row_id -> approved analog code_1c
    candidate_pools: dict[str, list[dict[str, Any]]] = {}


@app.post("/api/jobs/{job_id}/approve")
async def approve_analogs(
    job_id: str,
    request: Request,
    authorization: str | None = Header(None),
    token: str | None = Query(None),
):
    user = _check_auth(authorization, token)
    body = ApproveRequest(**(await _parse_request_payload(request)))
    data = _load_job(job_id)
    _ensure_job_access(user, data)

    approved_count = 0
    for row in data["rows"]:
        if row["id"] in body.approvals:
            code = body.approvals[row["id"]]
            previous_approved = copy.deepcopy(row.get("approved_analog")) if isinstance(row.get("approved_analog"), dict) else None
            # Find the analog with this code
            for analog in row["analogs"]:
                if analog["code_1c"] == code:
                    row["approved_analog"] = analog
                    row["status"] = "Одобрена замена"
                    row["selected_via"] = "analog"
                    row["selection_search_query"] = ""
                    row["selected_by"] = user.username
                    row["selected_by_display"] = user.display_name
                    row["selected_at"] = _iso_now()
                    _save_feedback_snapshot_for_row(
                        job_id=job_id,
                        row=row,
                        approved_candidate=analog,
                        user=user,
                        visible_candidates=body.candidate_pools.get(row["id"], []),
                        previous_approved_candidate=previous_approved,
                        selected_via="analog",
                        search_query="",
                        selected_at=str(row["selected_at"]),
                    )
                    approved_count += 1
                    break

    # Recalculate status counts
    status_counts = _rebuild_status_counts(data["rows"])
    data["status_counts"] = status_counts

    _save_job(job_id, data)
    return {"approved_count": approved_count, "status_counts": status_counts}


class ManualSearchRequest(BaseModel):
    row_id: str
    query: str
    limit: int = MANUAL_SEARCH_LIMIT


@app.post("/api/jobs/{job_id}/search")
async def search_stock_for_row(
    job_id: str,
    request: Request,
    authorization: str | None = Header(None),
    token: str | None = Query(None),
):
    user = _check_auth(authorization, token)
    body = ManualSearchRequest(**(await _parse_request_payload(request)))
    data = _load_job(job_id)
    _ensure_job_access(user, data)
    row = _find_row(data, body.row_id)
    query = body.query.strip()
    if len(query) < 2:
        raise HTTPException(400, "Введите минимум 2 символа для поиска")

    cache_key = _manual_search_cache_key(row, query, body.limit)
    cached_results = _manual_search_cache["entries"].get(cache_key)
    if cached_results is not None:
        return {
            "job_id": job_id,
            "row_id": body.row_id,
            "query": query,
            "results": copy.deepcopy(cached_results),
        }

    # Manual search does not reserve stock, so the shared master matcher is safe
    # and avoids cloning 46k stock rows on every "Найти" click.
    matcher = _load_cached_matcher()
    candidates = _manual_search_candidates(matcher, row, query, limit=body.limit)
    results = [_serialize_candidate(candidate) for candidate in candidates]
    _manual_search_cache["entries"][cache_key] = copy.deepcopy(results)

    return {
        "job_id": job_id,
        "row_id": body.row_id,
        "query": query,
        "results": results,
    }


class ManualSelectRequest(BaseModel):
    row_id: str
    candidate: dict[str, Any]
    search_query: str = ""
    visible_candidates: list[dict[str, Any]] = []


@app.post("/api/jobs/{job_id}/select")
async def select_manual_candidate(
    job_id: str,
    request: Request,
    authorization: str | None = Header(None),
    token: str | None = Query(None),
):
    user = _check_auth(authorization, token)
    body = ManualSelectRequest(**(await _parse_request_payload(request)))
    data = _load_job(job_id)
    _ensure_job_access(user, data)
    row = _find_row(data, body.row_id)
    candidate = body.candidate
    code = str(candidate.get("code_1c") or "").strip()
    name = str(candidate.get("name") or "").strip()
    if not code or not name:
        raise HTTPException(400, "Для выбора нужен code_1c и name")

    approved = {
        "code_1c": code,
        "name": name,
        "score": float(candidate.get("score") or 0.0),
        "stock_qty": candidate.get("stock_qty"),
        "remaining": candidate.get("remaining"),
        "price": candidate.get("price") or "",
        "reasons": list(candidate.get("reasons") or []),
    }
    if candidate.get("source_label"):
        approved["source_label"] = candidate["source_label"]
    if candidate.get("manager_choice"):
        approved["manager_choice"] = True
    previous_approved = copy.deepcopy(row.get("approved_analog")) if isinstance(row.get("approved_analog"), dict) else None
    row["approved_analog"] = approved
    row["status"] = "Одобрена замена"
    row["selected_via"] = "manual_search"
    row["selection_search_query"] = body.search_query.strip()
    row["selected_by"] = user.username
    row["selected_by_display"] = user.display_name
    row["selected_at"] = _iso_now()

    analog_codes = {analog.get("code_1c") for analog in row.get("analogs", [])}
    if approved["code_1c"] not in analog_codes:
        row.setdefault("analogs", [])
        row["analogs"].insert(0, approved)

    _save_feedback_snapshot_for_row(
        job_id=job_id,
        row=row,
        approved_candidate=approved,
        user=user,
        visible_candidates=body.visible_candidates,
        previous_approved_candidate=previous_approved,
        selected_via="manual_search",
        search_query=body.search_query,
        selected_at=str(row["selected_at"]),
    )

    status_counts = _rebuild_status_counts(data["rows"])
    data["status_counts"] = status_counts
    _save_job(job_id, data)
    return {"row": row, "status_counts": status_counts}


class RowQuantityUpdateRequest(BaseModel):
    quantity: float


@app.post("/api/jobs/{job_id}/rows/{row_id}/quantity")
async def update_row_quantity(
    job_id: str,
    row_id: str,
    request: Request,
    authorization: str | None = Header(None),
    token: str | None = Query(None),
):
    user = _check_auth(authorization, token)
    body = RowQuantityUpdateRequest(**(await _parse_request_payload(request)))
    data = _load_job(job_id)
    _ensure_job_access(user, data)
    if not _job_is_editable(data):
        raise HTTPException(400, "Количество можно менять только до выгрузки файла")

    row = _find_row(data, row_id)
    quantity = _parse_requested_quantity(body.quantity)
    _update_row_quantity_inplace(row, quantity)

    status_counts = _rebuild_status_counts(data["rows"])
    data["status_counts"] = status_counts
    _save_job(job_id, data)
    _jobs[job_id] = _build_job_summary(job_id, data, float(data.get("created_at") or time.time()))
    return data


@app.post("/api/jobs/{job_id}/export")
def export_1c_file(job_id: str, authorization: str | None = Header(None), token: str | None = Query(None)):
    user = _check_auth(authorization, token)
    data = _load_job(job_id)
    _ensure_job_access(user, data)

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers matching 1C format
    headers = ["Штрихкод", "Код", "Артикул", "Номенклатура", "Количество", "Цена"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    row_num = 2

    # Exact matches
    exact_statuses = {"Найдено полностью", "Найдено частично"}
    for r in data["rows"]:
        if r["status"] in exact_statuses and r.get("matched_code"):
            ws.cell(row=row_num, column=1, value=None).border = thin_border
            ws.cell(row=row_num, column=2, value=r["matched_code"]).border = thin_border
            ws.cell(row=row_num, column=3, value=None).border = thin_border
            ws.cell(row=row_num, column=4, value=r["matched_name"]).border = thin_border
            ws.cell(row=row_num, column=5, value=r["available_qty"]).border = thin_border
            ws.cell(row=row_num, column=6, value=r.get("matched_price") or "").border = thin_border
            row_num += 1

    # Approved analogs
    for r in data["rows"]:
        if r.get("approved_analog"):
            analog = r["approved_analog"]
            qty = r["requested_qty"]
            ws.cell(row=row_num, column=1, value=None).border = thin_border
            ws.cell(row=row_num, column=2, value=analog["code_1c"]).border = thin_border
            ws.cell(row=row_num, column=3, value=None).border = thin_border
            ws.cell(row=row_num, column=4, value=analog["name"]).border = thin_border
            ws.cell(row=row_num, column=5, value=qty).border = thin_border
            ws.cell(row=row_num, column=6, value=analog.get("price") or "").border = thin_border
            row_num += 1

    # Auto-size columns
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 65
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12

    job_dir = JOBS_DIR / job_id
    export_path = job_dir / "export_for_1c.xlsx"
    wb.save(export_path)

    learning_saved_count = 0
    for row in data["rows"]:
        approved_analog = row.get("approved_analog")
        if not isinstance(approved_analog, dict):
            continue
        if _save_feedback_snapshot_for_row(
            job_id=job_id,
            row=row,
            approved_candidate=approved_analog,
            user=user,
            selected_via=str(row.get("selected_via") or ""),
            search_query=str(row.get("selection_search_query") or ""),
        ):
            learning_saved_count += 1
    data["saved_at"] = _iso_now()
    data["saved_by"] = user.username
    data["saved_by_display"] = user.display_name
    data["learning_saved_count"] = learning_saved_count
    data["export_count"] = int(data.get("export_count", 0) or 0) + 1
    _save_job(job_id, data)
    _jobs[job_id] = _build_job_summary(job_id, data, float(data.get("created_at") or time.time()))
    _append_export_audit_event(job_id, data, user, learning_saved_count)

    return FileResponse(
        path=str(export_path),
        filename="КП_для_1С.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/admin/analytics")
def admin_analytics(authorization: str | None = Header(None), token: str | None = Query(None)):
    user = _check_auth(authorization, token)
    _require_admin(user)
    store = _get_learning_store()
    payload = {
        "exports": store.load_export_events(),
    }
    users = _load_user_directory()
    analytics = _build_admin_analytics(payload, users, store.load_feedback_entries())
    analytics["generated_at"] = _iso_now()
    return analytics


# ---------------------------------------------------------------------------
# Static files fallback (serve frontend)
# ---------------------------------------------------------------------------

WEB_DIR = Path(__file__).resolve().parent.parent / "web"

if WEB_DIR.exists():
    from fastapi.staticfiles import StaticFiles
    app.mount("/", StaticFiles(directory=str(WEB_DIR), html=True), name="static")


if __name__ == "__main__":
    import uvicorn

    ensure_output_dir(JOBS_DIR)
    uvicorn.run(app, host="0.0.0.0", port=8000)
