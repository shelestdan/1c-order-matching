#!/usr/bin/env python3
"""Normalize messy client requests into a standard XLSX table.

Supported inputs:
- xlsx / xlsm
- csv / tsv
- txt / md
- docx
- pdf
- png / jpg / webp / tif

The output workbook is intentionally stable so the stock-matching algorithm can
consume it without caring where the original request came from.
"""

from __future__ import annotations

import argparse
import csv
import re
import zipfile
from dataclasses import dataclass
from io import StringIO
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from document_text_extractor import ExtractionResult, extract_document_text
from process_1c_orders import (
    TOKEN_RE,
    clean_text,
    normalize_parser_line_text,
    should_ignore_parser_row,
    strip_parser_body_noise,
    coalesce_material_fields,
    extract_code_tokens,
    format_number,
    infer_column_defaults,
    get_mapped_quantity,
    get_mapped_text,
    parse_quantity,
    score_header_row,
)

STANDARD_HEADERS = [
    "Позиция",
    "Наименование",
    "Тип/марка",
    "Код",
    "Производитель",
    "Единица измер.",
    "Количество",
    "Комментарий",
    "Исходный текст",
    "Статус парсинга",
    "Уверенность",
]

DELIMITERS = ("\t", ";", "|")
UNIT_PATTERN = r"(шт\.?|pcs|компл(?:ект)?\.?|комп\.?|м2|м3|мп|м\.п\.|м|кг|л|упак\.?|уп\.?|пара)"
UNIT_NORMALIZATION = {
    "шт": "шт",
    "шт.": "шт",
    "pcs": "шт",
    "компл": "компл",
    "компл.": "компл",
    "комплект": "компл",
    "комп.": "компл",
    "м": "м",
    "мп": "м",
    "м.п.": "м",
    "м2": "м2",
    "м3": "м3",
    "кг": "кг",
    "л": "л",
    "уп": "уп",
    "уп.": "уп",
    "упак": "уп",
    "упак.": "уп",
    "пара": "пара",
}
COMMENT_KEYWORDS = ("срочно", "аналог", "замена", "доставка", "обязательно", "желательно")
NON_ITEM_HINTS = ("добрый", "здравствуйте", "нужны", "нужен", "интересует", "прошу", "спасибо", "добрый день")
LEADING_REQUEST_PREFIXES = (
    "нужно ",
    "нужна ",
    "нужен ",
    "нужны ",
    "надо ",
    "прошу ",
    "пожалуйста ",
)


@dataclass
class ParsedClientLine:
    source_file: str
    source_ref: str
    position: str
    name: str
    mark: str
    supplier_code: str
    vendor: str
    unit: str
    quantity: float
    comment: str
    raw_text: str
    parse_status: str
    confidence: float


@dataclass
class ParseIssue:
    source_file: str
    source_ref: str
    raw_text: str
    reason: str


@dataclass
class NormalizationMetadata:
    source_kind: str
    extraction_mode: str = ""
    page_count: int = 0
    warnings: list[str] | None = None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Normalize client requests into a standard XLSX table.")
    parser.add_argument("--inputs", nargs="+", required=True, help="Paths to request files or text files.")
    parser.add_argument(
        "--out-dir",
        default="outputs/normalized_requests",
        help="Directory for normalized XLSX files.",
    )
    return parser.parse_args()


def ensure_output_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def normalize_unit(unit: str) -> str:
    cleaned = clean_text(unit).lower()
    return UNIT_NORMALIZATION.get(cleaned, cleaned)


def strip_embedded_position(text: str, explicit_position: str) -> str:
    cleaned = clean_text(text)
    marker = clean_text(explicit_position)
    if not cleaned:
        return ""
    if marker:
        escaped = re.escape(marker)
        cleaned = re.sub(rf"^\s*{escaped}(?:[.)]|\s+-)?\s+", "", cleaned)
    cleaned = re.sub(r"^\s*\d+(?:[./]\d+)*(?:[.)]|\s+-)?\s+(?=[A-Za-zА-Яа-яЁёІіЇїЄєҐґ])", "", cleaned)
    return cleaned.strip()


def strip_inline_quantity_suffix(text: str, expected_qty: float, expected_unit: str) -> str:
    cleaned = clean_text(text)
    if not cleaned:
        return ""
    patterns = [
        re.compile(
            rf"^(?P<body>.+?)\s+(?P<unit>{UNIT_PATTERN})\s*[:.]?\s*(?P<qty>\d+(?:[.,]\d+)?)\s*$",
            re.IGNORECASE,
        ),
        re.compile(
            rf"^(?P<body>.+?)\s+(?P<qty>\d+(?:[.,]\d+)?)\s*(?P<unit>{UNIT_PATTERN})\s*$",
            re.IGNORECASE,
        ),
    ]
    normalized_expected_unit = normalize_unit(expected_unit)
    for pattern in patterns:
        match = pattern.match(cleaned)
        if not match:
            continue
        qty = parse_quantity(match.group("qty"))
        unit = normalize_unit(match.group("unit"))
        if qty <= 0:
            continue
        qty_matches = expected_qty <= 0 or abs(qty - expected_qty) < 1e-6
        unit_matches = not normalized_expected_unit or normalized_expected_unit == unit
        if qty_matches and unit_matches:
            return clean_text(match.group("body"))
    return cleaned


def autosize_columns(ws) -> None:
    for column in ws.columns:
        letter = column[0].column_letter
        max_len = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def add_row(ws, values: Iterable[object]) -> None:
    ws.append(list(values))
    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def detect_header_from_rows(rows: list[list[object]]) -> tuple[int, dict[str, int], list[str]]:
    best_row = 0
    best_mapping: dict[str, int] = {}
    best_headers: list[str] = []
    best_score = -1
    for row_index, row in enumerate(rows[:12]):
        score, mapping = score_header_row(row)
        if score > best_score:
            best_score = score
            best_row = row_index
            best_mapping = mapping
            best_headers = [clean_text(value) for value in row]
    best_mapping = infer_column_defaults(rows, best_mapping, best_row)
    return best_row, best_mapping, best_headers


def read_text_with_fallbacks(path: Path) -> str:
    encodings = ("utf-8-sig", "utf-8", "cp1251", "windows-1251", "latin1")
    for encoding in encodings:
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")


def read_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        data = archive.read("word/document.xml")
    root = ET.fromstring(data)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    for paragraph in root.findall(".//w:p", ns):
        parts = [node.text or "" for node in paragraph.findall(".//w:t", ns)]
        text = "".join(parts).strip()
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def looks_like_header(line: str) -> bool:
    lowered = line.lower()
    has_name = "наимен" in lowered or "материал" in lowered
    has_qty = "колич" in lowered or "кол-во" in lowered or "кол во" in lowered or "qty" in lowered
    has_position = "позиц" in lowered
    return (
        has_name and has_qty
        or (has_position and has_name)
    )


def strip_leading_marker(text: str) -> tuple[str, str]:
    text = text.strip()
    match = re.match(r"^(?P<pos>\d+(?:[./]\d+)*)(?:[.)]|(?:\s+-))\s+(?P<body>.+)$", text)
    if match:
        return match.group("pos"), match.group("body").strip()
    match = re.match(r"^(?:[-*•]+)\s+(?P<body>.+)$", text)
    if match:
        return "", match.group("body").strip()
    return "", text


def split_comment(text: str) -> tuple[str, str]:
    match = re.search(r"(?:,|\s+-\s+)(?P<comment>(?:" + "|".join(COMMENT_KEYWORDS) + r").*)$", text, flags=re.IGNORECASE)
    if not match:
        return text, ""
    body = text[: match.start()].strip(" ,-")
    return body, match.group("comment").strip()


def split_trailing_keyword_comment(text: str) -> tuple[str, str]:
    match = re.search(r"\s+(?P<comment>(?:" + "|".join(COMMENT_KEYWORDS) + r").*)$", text, flags=re.IGNORECASE)
    if not match:
        return text, ""
    body = text[: match.start()].strip(" ,-")
    return body, match.group("comment").strip()


def is_non_item_line(text: str) -> bool:
    lowered = text.lower()
    if re.search(r"\d", lowered):
        return False
    return any(hint in lowered for hint in NON_ITEM_HINTS)


def strip_request_prefix(text: str) -> str:
    lowered = text.lower()
    for prefix in LEADING_REQUEST_PREFIXES:
        if lowered.startswith(prefix):
            return text[len(prefix) :].strip()
    return text


def extract_explicit_vendor(text: str) -> tuple[str, str]:
    patterns = [
        r"(?:производитель|бренд|vendor)\s*[:\-]?\s*(?P<vendor>[A-Za-zА-Яа-я0-9 .&+/-]+)$",
        r"\((?P<vendor>[A-Za-zА-Яа-я0-9 .&+/-]+)\)$",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if not match:
            continue
        vendor = clean_text(match.group("vendor"))
        if len(vendor) >= 2 and any(char.isalpha() for char in vendor):
            cleaned = (text[: match.start()] + " " + text[match.end() :]).strip(" ,-;")
            return cleaned, vendor
    return text, ""


def extract_mark(text: str) -> str:
    raw_tokens = TOKEN_RE.findall(text)
    candidates: list[str] = []
    code_skeletons = extract_code_tokens(text)
    if not code_skeletons:
        return ""
    for token in raw_tokens:
        if not re.search(r"\d", token):
            continue
        if not re.search(r"[A-Za-zА-Яа-яЁёІіЇїЄєҐґ]", token):
            continue
        lowered = token.lower()
        if lowered.startswith(("dn", "pn", "ру", "ду", "kvs")):
            continue
        candidates.append(token)
    if not candidates:
        return ""
    candidates.sort(key=lambda token: (len(token), token.count("-"), token.count(".")), reverse=True)
    return candidates[0]


def extract_qty_from_text(text: str) -> tuple[str, float, str, float] | None:
    patterns = [
        re.compile(
            rf"^(?P<body>.+?)\s+(?:кол(?:-?во|ичество)?|qty)\s*[:=]?\s*(?P<qty>\d+(?:[.,]\d+)?)\s*(?P<unit>{UNIT_PATTERN})?\s*$",
            re.IGNORECASE,
        ),
        re.compile(
            rf"^(?P<body>.+?)\s+(?P<unit>{UNIT_PATTERN})\s*[:.]?\s*(?P<qty>\d+(?:[.,]\d+)?)\s*$",
            re.IGNORECASE,
        ),
        re.compile(
            rf"^(?P<body>.+?)\s*(?:[-–—,:;]\s*)?(?P<qty>\d+(?:[.,]\d+)?)\s*(?P<unit>{UNIT_PATTERN})\s*$",
            re.IGNORECASE,
        ),
        re.compile(r"^(?P<body>.+?)\s*[xх*]\s*(?P<qty>\d+(?:[.,]\d+)?)\s*$", re.IGNORECASE),
        re.compile(
            rf"^(?P<qty>\d+(?:[.,]\d+)?)\s*(?P<unit>{UNIT_PATTERN})\s+(?P<body>.+)$",
            re.IGNORECASE,
        ),
    ]
    scores = (0.95, 0.93, 0.92, 0.84, 0.8)
    for pattern, confidence in zip(patterns, scores):
        match = pattern.match(text.strip())
        if not match:
            continue
        body = clean_text(match.groupdict().get("body"))
        qty = parse_quantity(match.group("qty"))
        unit = normalize_unit(match.groupdict().get("unit") or "")
        if qty <= 0 or not body:
            continue
        return body, qty, unit, confidence
    return None


def parse_freeform_line(line: str, source_file: str, source_ref: str) -> tuple[ParsedClientLine | None, ParseIssue | None]:
    raw = normalize_parser_line_text(line)
    if not raw:
        return None, None
    if looks_like_header(raw):
        return None, None
    if should_ignore_parser_row(raw):
        return None, None
    if is_non_item_line(raw):
        return None, None

    position, body = strip_leading_marker(raw)
    body = strip_parser_body_noise(body, strip_trailing=False)
    body = strip_request_prefix(body)
    body, comment = split_comment(body)
    body, trailing_comment = split_trailing_keyword_comment(body)
    if trailing_comment:
        comment = trailing_comment if not comment else f"{comment}; {trailing_comment}"
    qty_result = extract_qty_from_text(body)
    if qty_result is None:
        body_without_qty, vendor = extract_explicit_vendor(body)
        body_without_qty = strip_parser_body_noise(body_without_qty)
        mark = extract_mark(body_without_qty)
        name = body_without_qty.strip(" ,-;")
        if len(name) < 3:
            return None, ParseIssue(source_file=source_file, source_ref=source_ref, raw_text=raw, reason="не удалось выделить количество")
        return (
            ParsedClientLine(
                source_file=source_file,
                source_ref=source_ref,
                position=position,
                name=name,
                mark=mark,
                supplier_code="",
                vendor=vendor,
                unit="шт",
                quantity=1.0,
                comment=(comment + "; количество не указано, принято 1 шт").strip("; "),
                raw_text=raw,
                parse_status="parsed_text_default_qty",
                confidence=0.45,
            ),
            None,
        )

    body_without_qty, quantity, unit, confidence = qty_result
    body_without_qty, vendor = extract_explicit_vendor(body_without_qty)
    body_without_qty = strip_parser_body_noise(body_without_qty)
    mark = extract_mark(body_without_qty)
    name = body_without_qty.strip(" ,-;")
    if len(name) < 3:
        return None, ParseIssue(source_file=source_file, source_ref=source_ref, raw_text=raw, reason="после очистки осталось слишком мало текста")

    return (
        ParsedClientLine(
            source_file=source_file,
            source_ref=source_ref,
            position=position,
            name=name,
            mark=mark,
            supplier_code="",
            vendor=vendor,
            unit=unit or "шт",
            quantity=quantity,
            comment=comment,
            raw_text=raw,
            parse_status="parsed_text",
            confidence=confidence,
        ),
        None,
    )


def parse_freeform_text(text: str, source_file: str, source_prefix: str) -> tuple[list[ParsedClientLine], list[ParseIssue]]:
    lines: list[ParsedClientLine] = []
    issues: list[ParseIssue] = []
    for index, raw_line in enumerate(text.splitlines(), start=1):
        stripped = raw_line.strip()
        if not stripped:
            continue
        parsed, issue = parse_freeform_line(stripped, source_file, f"{source_prefix}:{index}")
        if parsed is not None:
            if not parsed.position:
                parsed.position = str(len(lines) + 1)
            lines.append(parsed)
        elif issue is not None:
            issues.append(issue)
    return lines, issues


def merge_wrapped_ocr_lines(text: str) -> str:
    raw_lines = [line.strip() for line in text.splitlines() if line.strip()]
    if len(raw_lines) < 2:
        return "\n".join(raw_lines)

    merged_lines: list[str] = []
    index = 0
    while index < len(raw_lines):
        current = raw_lines[index]
        if index + 1 >= len(raw_lines):
            merged_lines.append(current)
            break

        next_line = raw_lines[index + 1]
        _, next_body = strip_leading_marker(next_line)
        has_current_qty = extract_qty_from_text(normalize_parser_line_text(current)) is not None
        current_parsed, _ = parse_freeform_line(current, "__ocr__", f"ocr:{index}")
        merged_candidate = clean_text(f"{current} {next_body}")
        merged_parsed, _ = parse_freeform_line(merged_candidate, "__ocr__", f"ocr:{index}")

        should_merge = (
            not has_current_qty
            and merged_parsed is not None
            and merged_parsed.parse_status != "parsed_text_default_qty"
            and (
                current_parsed is None
                or current_parsed.parse_status == "parsed_text_default_qty"
                or merged_parsed.confidence > current_parsed.confidence
            )
        )

        if should_merge:
            merged_lines.append(merged_candidate)
            index += 2
            continue

        merged_lines.append(current)
        index += 1

    return "\n".join(merged_lines)


def parse_delimited_text(
    text: str,
    source_file: str,
    source_prefix: str,
    delimiter: str,
) -> tuple[list[ParsedClientLine], list[ParseIssue]]:
    rows = list(csv.reader(StringIO(text), delimiter=delimiter))
    rows = [[cell for cell in row] for row in rows if any(clean_text(cell) for cell in row)]
    if not rows:
        return [], []
    header_index, columns, _ = detect_header_from_rows(rows)
    parsed: list[ParsedClientLine] = []
    issues: list[ParseIssue] = []
    for index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        row_text = " | ".join(clean_text(value) for value in row if clean_text(value))
        if not row_text:
            continue
        if should_ignore_parser_row(row_text):
            continue
        position = get_mapped_text(row, columns["position"])
        name = get_mapped_text(row, columns["name"])
        mark = get_mapped_text(row, columns["mark"])
        supplier_code = get_mapped_text(row, columns["supplier_code"])
        qty = get_mapped_quantity(row, columns["qty"])
        unit = normalize_unit(get_mapped_text(row, columns["unit"])) or "шт"
        name = strip_embedded_position(name, position)
        name = strip_inline_quantity_suffix(name, qty, unit)
        name = strip_parser_body_noise(name)
        mark = strip_inline_quantity_suffix(mark, qty, unit)
        name, mark, supplier_code = coalesce_material_fields(name, mark, supplier_code)
        if not name or qty <= 0:
            issues.append(
                ParseIssue(
                    source_file=source_file,
                    source_ref=f"{source_prefix}:{index}",
                    raw_text=row_text,
                    reason="строка не похожа на товарную позицию",
                )
            )
            continue
        parsed.append(
            ParsedClientLine(
                source_file=source_file,
                source_ref=f"{source_prefix}:{index}",
                position=position or str(len(parsed) + 1),
                name=name,
                mark=mark,
                supplier_code=supplier_code,
                vendor=get_mapped_text(row, columns["vendor"]),
                unit=unit,
                quantity=qty,
                comment="",
                raw_text=row_text,
                parse_status="parsed_table_text",
                confidence=0.98,
            )
        )
    return parsed, issues


def normalize_table_rows(
    rows: list[list[object]],
    source_file: str,
    source_prefix: str,
) -> tuple[list[ParsedClientLine], list[ParseIssue]]:
    if not rows:
        return [], []
    header_index, columns, _ = detect_header_from_rows(rows)
    parsed: list[ParsedClientLine] = []
    issues: list[ParseIssue] = []
    for index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        row_text = " | ".join(clean_text(value) for value in row if clean_text(value))
        if not row_text:
            continue
        if should_ignore_parser_row(row_text):
            continue
        position = get_mapped_text(row, columns["position"])
        name = get_mapped_text(row, columns["name"])
        mark = get_mapped_text(row, columns["mark"])
        supplier_code = get_mapped_text(row, columns["supplier_code"])
        qty = get_mapped_quantity(row, columns["qty"])
        unit = normalize_unit(get_mapped_text(row, columns["unit"])) or "шт"
        name = strip_embedded_position(name, position)
        name = strip_inline_quantity_suffix(name, qty, unit)
        name = strip_parser_body_noise(name)
        mark = strip_inline_quantity_suffix(mark, qty, unit)
        name, mark, supplier_code = coalesce_material_fields(name, mark, supplier_code)
        if not name or qty <= 0:
            issues.append(
                ParseIssue(
                    source_file=source_file,
                    source_ref=f"{source_prefix}:{index}",
                    raw_text=row_text,
                    reason="в табличной строке не удалось выделить наименование и количество",
                )
            )
            continue
        parsed.append(
            ParsedClientLine(
                source_file=source_file,
                source_ref=f"{source_prefix}:{index}",
                position=position or str(len(parsed) + 1),
                name=name,
                mark=mark,
                supplier_code=supplier_code,
                vendor=get_mapped_text(row, columns["vendor"]),
                unit=unit,
                quantity=qty,
                comment="",
                raw_text=row_text,
                parse_status="parsed_sheet",
                confidence=0.99,
            )
        )
    return parsed, issues


def normalize_xlsx_request(path: Path) -> tuple[list[ParsedClientLine], list[ParseIssue]]:
    wb = load_workbook(path, data_only=True)
    parsed: list[ParsedClientLine] = []
    issues: list[ParseIssue] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [[ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)] for row in range(1, ws.max_row + 1)]
        sheet_parsed, sheet_issues = normalize_table_rows(rows, path.name, sheet_name)
        parsed.extend(sheet_parsed)
        issues.extend(sheet_issues)
        if sheet_parsed:
            continue
        flattened_lines = []
        for row_index, row in enumerate(rows, start=1):
            row_text = " ; ".join(clean_text(value) for value in row if clean_text(value))
            if row_text:
                flattened_lines.append((row_index, row_text))
        for row_index, row_text in flattened_lines:
            parsed_line, issue = parse_freeform_line(row_text, path.name, f"{sheet_name}:{row_index}")
            if parsed_line is not None:
                if not parsed_line.position:
                    parsed_line.position = str(len(parsed) + 1)
                parsed_line.parse_status = "parsed_sheet_fallback"
                parsed.append(parsed_line)
            elif issue is not None:
                issues.append(issue)
    return parsed, issues


def normalize_csv_request(path: Path) -> tuple[list[ParsedClientLine], list[ParseIssue]]:
    text = read_text_with_fallbacks(path)
    delimiter = "\t" if "\t" in text else ";"
    return parse_delimited_text(text, path.name, path.stem, delimiter)


def normalize_text_request(path: Path) -> tuple[list[ParsedClientLine], list[ParseIssue]]:
    text = read_text_with_fallbacks(path)
    non_empty_lines = [line for line in text.splitlines() if line.strip()]
    if non_empty_lines and any(delimiter in non_empty_lines[0] for delimiter in DELIMITERS):
        for delimiter in DELIMITERS:
            if delimiter in non_empty_lines[0]:
                parsed, issues = parse_delimited_text(text, path.name, path.stem, delimiter)
                if parsed:
                    return parsed, issues
    return parse_freeform_text(text, path.name, path.stem)


def normalize_docx_request(path: Path) -> tuple[list[ParsedClientLine], list[ParseIssue]]:
    text = read_docx_text(path)
    return parse_freeform_text(text, path.name, path.stem)


def normalize_document_request(path: Path) -> tuple[list[ParsedClientLine], list[ParseIssue], NormalizationMetadata]:
    extraction = extract_document_text(path)
    issues: list[ParseIssue] = []
    for warning in extraction.warnings:
        issues.append(
            ParseIssue(
                source_file=path.name,
                source_ref=path.stem,
                raw_text="",
                reason=warning,
            )
        )
    if not extraction.text.strip():
        return [], issues, NormalizationMetadata(
            source_kind=extraction.kind,
            extraction_mode=extraction.extraction_mode,
            page_count=extraction.page_count,
            warnings=extraction.warnings,
        )

    text = extraction.text
    if extraction.kind == "image":
        text = merge_wrapped_ocr_lines(text)
    non_empty_lines = [line for line in text.splitlines() if line.strip()]
    parsed: list[ParsedClientLine]
    parse_issues: list[ParseIssue]
    if non_empty_lines and any(delimiter in non_empty_lines[0] for delimiter in DELIMITERS):
        parsed = []
        parse_issues = []
        for delimiter in DELIMITERS:
            if delimiter in non_empty_lines[0]:
                parsed, parse_issues = parse_delimited_text(text, path.name, path.stem, delimiter)
                if parsed:
                    break
        freeform_text = text
        for delimiter in DELIMITERS:
            freeform_text = freeform_text.replace(delimiter, " ")
        freeform_parsed, freeform_issues = parse_freeform_text(freeform_text, path.name, path.stem)
        normalized_first_line = non_empty_lines[0]
        for delimiter in DELIMITERS:
            normalized_first_line = normalized_first_line.replace(delimiter, " ")
        header_like_first_line = looks_like_header(normalized_first_line)
        if len(parsed) == 0 or (not header_like_first_line and len(freeform_parsed) > len(parsed)):
            parsed, parse_issues = freeform_parsed, freeform_issues
    else:
        parsed, parse_issues = parse_freeform_text(text, path.name, path.stem)

    parse_status_prefix = extraction.extraction_mode or extraction.kind
    for line in parsed:
        if line.parse_status.startswith("parsed_"):
            line.parse_status = f"{line.parse_status}_{parse_status_prefix}"
        else:
            line.parse_status = f"parsed_{parse_status_prefix}"

    return parsed, issues + parse_issues, NormalizationMetadata(
        source_kind=extraction.kind,
        extraction_mode=extraction.extraction_mode,
        page_count=extraction.page_count,
        warnings=extraction.warnings,
    )


def normalize_request_with_metadata(path: Path) -> tuple[list[ParsedClientLine], list[ParseIssue], NormalizationMetadata]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        parsed, issues = normalize_xlsx_request(path)
        return parsed, issues, NormalizationMetadata(source_kind="spreadsheet")
    if suffix in {".csv", ".tsv"}:
        parsed, issues = normalize_csv_request(path)
        return parsed, issues, NormalizationMetadata(source_kind="delimited_text")
    if suffix in {".txt", ".md"}:
        parsed, issues = normalize_text_request(path)
        return parsed, issues, NormalizationMetadata(source_kind="text")
    if suffix == ".docx":
        parsed, issues = normalize_docx_request(path)
        return parsed, issues, NormalizationMetadata(source_kind="docx")
    if suffix == ".pdf" or suffix in {".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff", ".bmp", ".gif"}:
        return normalize_document_request(path)
    raise ValueError(f"Unsupported input format: {path.suffix}")


def normalize_request(path: Path) -> tuple[list[ParsedClientLine], list[ParseIssue]]:
    parsed, issues, _ = normalize_request_with_metadata(path)
    return parsed, issues


def write_normalized_workbook(
    source_path: Path,
    parsed_lines: list[ParsedClientLine],
    issues: list[ParseIssue],
    metadata: NormalizationMetadata,
    out_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявка"
    ws.append(STANDARD_HEADERS)
    for index, line in enumerate(parsed_lines, start=1):
        add_row(
            ws,
            [
                line.position or str(index),
                line.name,
                line.mark,
                line.supplier_code,
                line.vendor,
                line.unit or "шт",
                format_number(line.quantity),
                line.comment,
                line.raw_text,
                line.parse_status,
                round(line.confidence, 2),
            ],
        )
    style_header(ws)
    autosize_columns(ws)

    issues_ws = wb.create_sheet("Проблемы")
    issues_ws.append(["Файл", "Источник", "Исходный текст", "Причина"])
    for issue in issues:
        add_row(issues_ws, [issue.source_file, issue.source_ref, issue.raw_text, issue.reason])
    style_header(issues_ws)
    autosize_columns(issues_ws)

    summary_ws = wb.create_sheet("Сводка")
    summary_ws.append(["Параметр", "Значение"])
    summary_ws.append(["Исходный файл", source_path.name])
    summary_ws.append(["Тип источника", metadata.source_kind])
    summary_ws.append(["Режим извлечения", metadata.extraction_mode or "native"])
    summary_ws.append(["Страниц/кадров", metadata.page_count or ""])
    summary_ws.append(["Распознано строк", len(parsed_lines)])
    summary_ws.append(["Проблемных строк", len(issues)])
    summary_ws.append(["Предупреждений extractor", len(metadata.warnings or [])])
    style_header(summary_ws)
    autosize_columns(summary_ws)
    wb.save(out_path)


def normalize_request_file(input_path: Path, out_dir: Path) -> tuple[Path, int, int]:
    parsed_lines, issues, metadata = normalize_request_with_metadata(input_path)
    out_path = out_dir / f"{input_path.stem}__normalized.xlsx"
    write_normalized_workbook(input_path, parsed_lines, issues, metadata, out_path)
    return out_path, len(parsed_lines), len(issues)


def main() -> int:
    args = parse_args()
    out_dir = ensure_output_dir(Path(args.out_dir).expanduser().resolve())
    for input_name in args.inputs:
        input_path = Path(input_name).expanduser().resolve()
        normalized_path, parsed_count, issue_count = normalize_request_file(input_path, out_dir)
        print(f"\n{input_path.name}")
        print(f"  normalized: {normalized_path}")
        print(f"  parsed rows: {parsed_count}")
        print(f"  issues: {issue_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
