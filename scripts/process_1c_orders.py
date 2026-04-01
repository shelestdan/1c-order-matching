#!/usr/bin/env python3
"""Process order XLSX files against a 1C stock CSV.

The pipeline is intentionally deterministic:
1. Read the stock CSV as the source of truth.
2. Normalize mixed-script text from incoming XLSX requests.
3. Find exact or high-confidence matches with strict penalties for size/code mismatches.
4. Reserve stock sequentially so one остаток row is not spent twice.
5. Export:
   - a detailed workbook,
   - a workbook for 1C/import preparation,
   - analog suggestions,
   - not found / out-of-stock rows.

This prototype is tuned for the structures seen in the supplied files, but it
is generic enough to handle similar request forms from managers.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
from bisect import bisect_right
from collections import defaultdict
from dataclasses import dataclass, field, replace
from datetime import datetime
from pathlib import Path
from typing import Iterable, Mapping, Sequence

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from nomenclature_classifier import ClassificationStatus, QueryClassification, load_default_classifier

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_REVIEWED_ANALOG_DECISIONS_PATH = PROJECT_ROOT / "data" / "reviewed_analog_decisions.json"
CLASSIFIER_CONFIG_PATH = PROJECT_ROOT / "data" / "nomenclature_classifier_config.json"
DOMAIN_DICTIONARY_PATH = PROJECT_ROOT / "data" / "vgs2000_matching_dictionary.json"
SUBSTITUTION_POLICY_PATH = PROJECT_ROOT / "data" / "substitution_policy.json"
DEFAULT_MATCHING_GOLDEN_SET_PATH = PROJECT_ROOT / "data" / "matching_golden_set.json"
PARSER_HINTS_PATH = PROJECT_ROOT / "data" / "parser_hints.json"
REVIEW_DECISION_APPROVED = "approved"
REVIEW_DECISION_REJECTED = "rejected"

STATUS_FOUND_FULL = "Найдено полностью"
STATUS_FOUND_PARTIAL = "Найдено частично"
STATUS_SAFE_ANALOG = "Безопасный аналог"
STATUS_APPROVAL_ANALOG = "Допустимая замена по согласованию"
STATUS_NOT_FOUND = "Не найдено"
STATUS_STOCK_DEPLETED = "Найдено, но остаток уже распределен"
ANALOG_STATUSES = {STATUS_SAFE_ANALOG, STATUS_APPROVAL_ANALOG}

TOKEN_RE = re.compile(r"[0-9A-Za-zА-Яа-яЁёІіЇїЄєҐґ/+,.\"=:-]+")
CYRILLIC_RE = re.compile(r"[А-Яа-яЁёІіЇїЄєҐґ]")
LATIN_RE = re.compile(r"[A-Za-z]")
DIGIT_RE = re.compile(r"\d")

LATIN_TO_CYR_MIXED = {
    "A": "А",
    "a": "а",
    "B": "В",
    "C": "С",
    "c": "с",
    "E": "Е",
    "e": "е",
    "H": "Н",
    "K": "К",
    "k": "к",
    "M": "М",
    "O": "О",
    "o": "о",
    "P": "Р",
    "p": "р",
    "T": "Т",
    "X": "Х",
    "x": "х",
    "Y": "У",
    "y": "у",
    "V": "В",
    "v": "в",
    "N": "Н",
    "u": "и",
    "U": "И",
    "i": "і",
    "I": "І",
    "s": "ѕ",
    "S": "Ѕ",
}

VISUAL_CODE_MAP = {
    "А": "a",
    "а": "a",
    "A": "a",
    "a": "a",
    "В": "b",
    "в": "b",
    "B": "b",
    "b": "b",
    "С": "c",
    "с": "c",
    "C": "c",
    "c": "c",
    "Е": "e",
    "е": "e",
    "E": "e",
    "e": "e",
    "Н": "h",
    "н": "h",
    "H": "h",
    "h": "h",
    "І": "i",
    "і": "i",
    "I": "i",
    "i": "i",
    "К": "k",
    "к": "k",
    "K": "k",
    "k": "k",
    "М": "m",
    "м": "m",
    "M": "m",
    "m": "m",
    "О": "o",
    "о": "o",
    "O": "o",
    "o": "o",
    "Р": "p",
    "р": "p",
    "P": "p",
    "p": "p",
    "Ѕ": "s",
    "ѕ": "s",
    "S": "s",
    "s": "s",
    "Т": "t",
    "т": "t",
    "T": "t",
    "t": "t",
    "У": "y",
    "у": "y",
    "Y": "y",
    "y": "y",
    "Х": "x",
    "х": "x",
    "X": "x",
    "x": "x",
}

CYR_TO_LAT = {
    "а": "a",
    "б": "b",
    "в": "v",
    "г": "g",
    "д": "d",
    "е": "e",
    "ё": "e",
    "ж": "zh",
    "з": "z",
    "и": "i",
    "і": "i",
    "й": "i",
    "к": "k",
    "л": "l",
    "м": "m",
    "н": "n",
    "о": "o",
    "п": "p",
    "р": "r",
    "с": "s",
    "т": "t",
    "у": "u",
    "ф": "f",
    "х": "h",
    "ц": "c",
    "ч": "ch",
    "ш": "sh",
    "щ": "sch",
    "ъ": "",
    "ы": "y",
    "ь": "",
    "э": "e",
    "ю": "yu",
    "я": "ya",
    "ї": "i",
    "є": "e",
    "ґ": "g",
    "ѕ": "s",
}

HEADER_KEYWORDS = {
    "position": ("poz", "pozici", "pozits", "позиц", "поз", "nomer", "номер"),
    "name": ("naimen", "naim", "harakter", "наимен", "наим", "характер"),
    "mark": ("tip", "mark", "obozn", "модель", "тип", "марк", "обозн", "art", "арт"),
    "supplier_code": ("kod", "код", "artikul", "артикул"),
    "vendor": ("zavod", "izgotov", "postav", "brand", "бренд", "изготов", "постав"),
    "unit": ("edinic", "izmer", "единиц", "измер", "edizm", "ed izm", "ед.изм", "едизм"),
    "qty": ("kolich", "kolvo", "kol-vo", "kvo", "k-vo", "qty", "к-во", "колич", "колво", "кол-во"),
}

COMMON_TOKENS = {
    "sht",
    "sht.",
    "komplekt",
    "komplektom",
    "trub",
    "sistema",
    "sistem",
    "detal",
    "material",
    "izdelie",
    "oborudovanie",
    "postavka",
    "zakazchika",
    "шт",
    "шт.",
}

POSITION_CELL_RE = re.compile(r"^\d+(?:[./]\d+)*$")
UNIT_CELL_RE = re.compile(r"^(?:шт|шт\.|м|м2|м3|кг|л|компл|компл\.|комп|уп|уп\.|пара)$", re.IGNORECASE)

INCH_TO_DN = {
    "1/2": "15",
    "3/4": "20",
    "1": "25",
    "11/4": "32",
    "1-1/4": "32",
    "11/2": "40",
    "1-1/2": "40",
    "2": "50",
    "21/2": "65",
    "2-1/2": "65",
    "3": "80",
    "4": "100",
}

VGP_DN_TO_OD = {
    "15": "21.3",
    "20": "26.8",
    "25": "33.5",
    "32": "42.3",
    "40": "48.0",
    "50": "60.0",
    "65": "75.5",
    "80": "88.5",
    "100": "114.0",
}
VGP_OD_TO_DN = {value: key for key, value in VGP_DN_TO_OD.items()}
VGP_OD_TO_DN.update(
    {
        "48": "40",
        "60": "50",
        "114": "100",
    }
)

STATUS_COLORS = {
    STATUS_FOUND_FULL: "C6EFD1",
    STATUS_FOUND_PARTIAL: "FFF2CC",
    STATUS_SAFE_ANALOG: "D9EAD3",
    STATUS_APPROVAL_ANALOG: "FCE5CD",
    STATUS_NOT_FOUND: "F4CCCC",
    STATUS_STOCK_DEPLETED: "F4CCCC",
}

STATUS_PRIORITY = {
    STATUS_NOT_FOUND: 0,
    STATUS_APPROVAL_ANALOG: 1,
    STATUS_SAFE_ANALOG: 2,
    STATUS_STOCK_DEPLETED: 3,
    STATUS_FOUND_PARTIAL: 4,
    STATUS_FOUND_FULL: 5,
}

DIMENSION_WEIGHTS = {
    "dn": (10.0, 22.0),
    "inch": (10.0, 22.0),
    "od": (14.0, 40.0),
    "pairdn": (18.0, 32.0),
    "tripledn": (34.0, 48.0),
    "type": (10.0, 16.0),
    "exec": (8.0, 12.0),
    "face": (6.0, 10.0),
    "hole": (4.0, 8.0),
    "wall": (8.0, 12.0),
    "deg": (8.0, 18.0),
    "series": (6.0, 10.0),
    "mat": (8.0, 12.0),
    "spec": (8.0, 8.0),
    "pn": (6.0, 12.0),
    "kvs": (6.0, 12.0),
    "lmm": (6.0, 12.0),
    "subfamily": (10.0, 16.0),
    "subtype": (8.0, 12.0),
    "shape": (8.0, 12.0),
    "connection": (6.0, 10.0),
    "actuator": (6.0, 10.0),
    "handle": (4.0, 8.0),
    "extra_connection": (4.0, 8.0),
    "conn_gender": (6.0, 10.0),
    "meter_type": (10.0, 16.0),
    "interface": (4.0, 8.0),
    "pulse_output": (4.0, 8.0),
    "pressure_range_mpa": (10.0, 18.0),
    "operator": (6.0, 10.0),
    "for_family": (8.0, 12.0),
    "mounting_pad": (10.0, 14.0),
    "stem_square": (10.0, 14.0),
    "compatible_dn": (12.0, 18.0),
    "turn_type": (8.0, 12.0),
    "body_material": (4.0, 8.0),
    "disc_material": (4.0, 8.0),
    "connection_guess": (4.0, 8.0),
    "connection_position": (4.0, 8.0),
    "accuracy_class": (4.0, 8.0),
    "orientation": (4.0, 8.0),
    "face_to_face": (4.0, 8.0),
    "has_magnet": (8.0, 12.0),
    "filter_element": (6.0, 10.0),
}

FAMILY_PATTERNS = {
    "adapter": ("adapter",),
    "chugun_fason": ("vchshg", "tyton"),
    "elbow": ("otvod", "koleno", "ugolnik", "ugol"),
    "filter": ("filtr",),
    "flange": ("flanec",),
    "grooved": ("gruvlok",),
    "homut": ("homut",),
    "compensator": ("kompensator", "gibk", "vstavk"),
    "kran": ("kran", "kshc"),
    "klapan": ("klapan",),
    "manometer": ("manometr",),
    "mixer": ("smesitel",),
    "mufta": ("mufta", "amerikank"),
    "patrubok": ("patrub", "pds", "pps", "pfg", "pfrk"),
    "perehod": ("perehod", "pereh"),
    "pipe": ("truba",),
    "pozharka": ("pozharnyi",),
    "radiator": ("radiator", "registr"),
    "rezba": ("rezba",),
    "schetchik": ("schetchik", "vodomer", "rashodomer"),
    "sink": ("umyval", "moik"),
    "tee": ("troinik",),
    "zadvizhka": ("zadvizh",),
    "zatvor": ("zatvor",),
    "toilet": ("unitaz",),
    "zaglushka": ("zaglush",),
    "vozduhootvod": ("vozduhootvod", "vozdnhootvod", "vozdootvod", "vozdushnik"),
}

FAMILY_CONTAINS_TOKEN_PATTERNS = {"rezba"}

MATERIAL_CLASS_RULES = {
    "stainless": {
        "contains": ("nerzh", "inox"),
        "exact": {"aisi", "s30400", "s30403", "s31600", "s31603", "s32100", "sus304", "sus316", "sus321"},
    },
    "steel": {
        "contains": ("stal",),
        "exact": {"wcb", "09g2s", "vgp"},
        "prefix": ("mat",),
    },
    "cast_iron": {
        "contains": ("chugun", "vchshg", "kovk"),
        "exact": set(),
    },
    "brass": {
        "contains": ("latun",),
        "exact": set(),
    },
    "bronze": {
        "contains": ("bronz",),
        "exact": set(),
    },
    "copper": {
        "contains": ("medn",),
        "exact": {"med"},
    },
    "polypropylene": {
        "contains": ("poliprop",),
        "exact": {"ppr", "pp-r", "pilsa", "kalde"},
    },
    "pvc": {
        "contains": ("pvc", "pvh"),
        "exact": {"pvc-u", "u-pvc"},
    },
    "polyethylene": {
        "contains": ("polietilen",),
        "exact": {"pe", "pe80", "pe100", "pnd", "hdpe", "pex", "pex-a", "pex-b", "pe-rt", "pert"},
    },
}

GRADE_GROUPS = {
    "ss304": {"1.4301", "304", "s30400", "sus304", "08x18h10", "0cr18ni9"},
    "ss304l": {"1.4307", "304l", "s30403", "sus304l", "03x18h11", "00cr19ni10"},
    "ss316": {"1.4401", "316", "s31600", "sus316", "08x17h13m2", "0cr17ni12mo2"},
    "ss316l": {"1.4404", "316l", "s31603", "sus316l", "03x17h14m2", "00cr17ni14mo2"},
    "ss321": {"1.4541", "321", "s32100", "sus321", "08x18h10t", "0cr18ni10ti"},
    "cs20": {"mat20", "st20", "ct20"},
    "cs09g2s": {"09g2s", "mat09g2s"},
}

GRADE_TO_CLASS = {
    "ss304": "stainless",
    "ss304l": "stainless",
    "ss316": "stainless",
    "ss316l": "stainless",
    "ss321": "stainless",
    "cs20": "steel",
    "cs09g2s": "steel",
}

GRADE_SUBSTRING_ALIASES = {
    "ss304": ("08x18n10", "08x18h10", "1.4301"),
    "ss304l": ("03x18h11", "00cr19ni10", "1.4307"),
    "ss316": ("08x17h13m2", "0cr17ni12mo2", "1.4401"),
    "ss316l": ("03x17h14m2", "00cr17ni14mo2", "1.4404"),
    "ss321": ("08x18n10t", "08x18h10t", "12x18n10t", "12x18h10t", "0cr18ni10ti", "1.4541"),
    "cs20": ("st.20", "stalnoi", "stalnaya", "stalnye", "stalnyh", "stalnogo"),
}

PRIMARY_FAMILY_TAGS = {
    "adapter",
    "chugun_fason",
    "compensator",
    "elbow",
    "filter",
    "homut",
    "gearbox",
    "klapan",
    "kran",
    "manometer",
    "mixer",
    "mufta",
    "patrubok",
    "perehod",
    "pipe",
    "pozharka",
    "radiator",
    "schetchik",
    "sink",
    "tee",
    "toilet",
    "vozduhootvod",
    "zadvizhka",
    "zatvor",
    "zaglushka",
}

FAMILY_EXACT_REASONS = {
    "pipe": "совпали тип трубы, размер и ГОСТ",
    "tee": "совпали тип изделия и тройные размеры",
    "perehod": "совпали тип изделия и обе стороны перехода",
}

FAMILY_ANALOG_REASONS = {
    "compensator": "совпали семейство и DN, нужна ручная проверка исполнения",
    "filter": "совпали семейство и DN, нужна ручная проверка исполнения",
    "kran": "совпали семейство и размер, нужна ручная проверка исполнения",
    "manometer": "совпали семейство и ключевые параметры, нужна ручная проверка исполнения",
    "patrubok": "совпали семейство, подтип и DN",
    "pozharka": "совпали семейство и основные размеры, нужна ручная проверка",
    "schetchik": "совпали семейство и ключевые параметры счетчика, нужна ручная проверка",
    "chugun_fason": "совпали семейство и DN, нужна ручная проверка исполнения",
    "gearbox": "совпали семейство и ключевые параметры редуктора",
    "vozduhootvod": "совпали семейство и DN, нужна ручная проверка исполнения",
    "zatvor": "совпали семейство и ключевые размеры, нужна ручная проверка",
    "zadvizhka": "совпали семейство и DN, нужна ручная проверка исполнения",
}

DOMAIN_DICTIONARY_PATH = Path(__file__).resolve().parent.parent / "data" / "vgs2000_matching_dictionary.json"


def load_domain_dictionary() -> dict[str, object]:
    if not DOMAIN_DICTIONARY_PATH.exists():
        return {}
    try:
        return json.loads(DOMAIN_DICTIONARY_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def merge_family_patterns(base: dict[str, tuple[str, ...]], extra: object) -> dict[str, tuple[str, ...]]:
    merged = {key: tuple(values) for key, values in base.items()}
    if not isinstance(extra, dict):
        return merged
    for family, values in extra.items():
        if not isinstance(family, str) or not isinstance(values, list):
            continue
        normalized_values = [str(value).strip() for value in values if str(value).strip()]
        if not normalized_values:
            continue
        merged[family] = tuple(dict.fromkeys([*merged.get(family, ()), *normalized_values]))
    return merged


DOMAIN_DICTIONARY = load_domain_dictionary()
FAMILY_PATTERNS = merge_family_patterns(FAMILY_PATTERNS, DOMAIN_DICTIONARY.get("family_aliases"))
MANUAL_REVIEW_RULES = DOMAIN_DICTIONARY.get("manual_review_rules", [])
NOMENCLATURE_CLASSIFIER = load_default_classifier()
PARSER_HINTS: dict[str, object] = {}


def _normalize_hint_phrase(value: object) -> str:
    text = normalize_symbols(clean_text(value))
    text = text.lower().replace("ё", "е").replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _compile_parser_token_aliases(payload: Mapping[str, object]) -> dict[str, tuple[str, ...]]:
    compiled: dict[str, tuple[str, ...]] = {}
    if not isinstance(payload, Mapping):
        return compiled
    for canonical, aliases in payload.items():
        if not isinstance(canonical, str) or not isinstance(aliases, list):
            continue
        normalized = tuple(
            dict.fromkeys(
                _normalize_hint_phrase(alias)
                for alias in aliases
                if _normalize_hint_phrase(alias)
            )
        )
        if normalized:
            compiled[canonical] = normalized
    return compiled


def _compile_parser_family_triggers(payload: Mapping[str, object]) -> dict[str, dict[str, object]]:
    compiled: dict[str, dict[str, object]] = {}
    if not isinstance(payload, Mapping):
        return compiled
    for family, rule in payload.items():
        if not isinstance(family, str) or not isinstance(rule, Mapping):
            continue
        aliases_raw = rule.get("aliases", [])
        resolve_raw = rule.get("resolve_rules", [])
        aliases = tuple(
            dict.fromkeys(
                _normalize_hint_phrase(alias)
                for alias in aliases_raw
                if _normalize_hint_phrase(alias)
            )
        )
        resolve_rules: list[dict[str, object]] = []
        if isinstance(resolve_raw, list):
            for raw_rule in resolve_raw:
                if not isinstance(raw_rule, Mapping):
                    continue
                match_any = tuple(
                    dict.fromkeys(
                        _normalize_hint_phrase(alias)
                        for alias in raw_rule.get("match_any", [])
                        if _normalize_hint_phrase(alias)
                    )
                )
                set_payload = raw_rule.get("set", {})
                if not match_any or not isinstance(set_payload, Mapping):
                    continue
                resolve_rules.append({"match_any": match_any, "set": dict(set_payload)})
        compiled[family] = {"aliases": aliases, "resolve_rules": tuple(resolve_rules)}
    return compiled


PARSER_ALIAS_TAGS = {
    "female_female": ("conn_gender:ff",),
    "female_male": ("conn_gender:fm",),
    "male_male": ("conn_gender:mm",),
    "butterfly_handle": ("handle:butterfly",),
    "lever_handle": ("handle:lever",),
    "american_union": ("extra_connection:union",),
    "grooved": ("connection:grooved",),
    "magnetic_insert": ("has_magnet:true",),
    "mesh_strainer": ("filter_element:mesh",),
    "rubber_wedge": ("subtype:rubber_wedge",),
    "handwheel": ("actuator:handwheel",),
    "electric_actuator": ("actuator:electric",),
    "gearbox_operator": ("operator:gearbox",),
    "before_meter": ("subfamily:pds",),
    "after_meter": ("subfamily:pps",),
    "flange_plain_end": ("subfamily:pfg",),
    "combined_meter": ("meter_type:combined",),
    "turbine_meter": ("meter_type:turbine",),
    "impeller_meter": ("meter_type:impeller",),
    "pulse_output": ("pulse_output:true",),
    "rs485": ("interface:rs-485",),
    # Legacy flange GOSTs → current standard + subtype tags
    "flange_gost_33259": ("family:flange", "spec:33259"),
    "flange_flat_welded": ("family:flange", "type:01"),
    "flange_weld_neck": ("family:flange", "type:11"),
}

PARSER_FAMILY_CANONICAL = {
    "manometr": "manometer",
}

PARSER_QUERY_HINTS: dict[str, object] = {}
PARSER_PREPROCESSING_HINTS: dict[str, object] = {}
PARSER_MIXED_SCRIPT_HINTS: dict[str, object] = {}
PARSER_TOKEN_ALIASES: dict[str, tuple[str, ...]] = {}
PARSER_FAMILY_TRIGGERS: dict[str, dict[str, object]] = {}
PARSER_MIXED_CHAR_MAP: dict[str, str] = dict(LATIN_TO_CYR_MIXED)
PARSER_IGNORE_FULL_ROW_PATTERNS: tuple[str, ...] = ()
PARSER_STRIP_LEADING_PATTERNS: tuple[str, ...] = ()
PARSER_STRIP_TRAILING_PATTERNS: tuple[str, ...] = ()


@dataclass
class StockItem:
    row_index: int
    code_1c: str
    name: str
    print_name: str
    product_type: str
    sale_price: str
    stop_price: str
    plan_price: str
    quantity: float
    remaining: float
    search_text: str
    search_tokens: set[str]
    key_tokens: set[str]
    root_tokens: set[str]
    code_tokens: set[str]
    dimension_tags: set[str]


@dataclass
class OrderLine:
    source_file: str
    sheet_name: str
    source_row: int
    headers: list[str]
    row_values: list[object]
    position: str
    name: str
    mark: str
    supplier_code: str
    vendor: str
    unit: str
    requested_qty: float
    search_text: str
    search_tokens: set[str]
    key_tokens: set[str]
    root_tokens: set[str]
    code_tokens: set[str]
    dimension_tags: set[str]
    raw_query: str
    classification: QueryClassification | None = None


@dataclass
class Candidate:
    stock: StockItem
    score: float
    overlap: float
    soft_overlap: float
    reasons: list[str] = field(default_factory=list)
    code_hit: bool = False
    dimension_bonus: float = 0.0
    dimension_penalty: float = 0.0
    matched_dimension_keys: tuple[str, ...] = ()
    conflicting_dimension_keys: tuple[str, ...] = ()
    review_decision: str = ""


@dataclass
class MatchResult:
    order: OrderLine
    status: str
    matched_stock: StockItem | None
    available_qty: float
    confidence: float
    comment: str
    analogs: list[Candidate]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Match order XLSX files against 1C stock CSV.")
    parser.add_argument("--stock", required=True, help="Path to the stock CSV exported from 1C.")
    parser.add_argument("--orders", nargs="+", required=True, help="One or more order XLSX files.")
    parser.add_argument(
        "--out-dir",
        default="outputs/1c_matching",
        help="Directory where result workbooks will be written.",
    )
    parser.add_argument(
        "--reviewed-analog-decisions",
        default="",
        help=(
            "Optional JSON artifact with reviewed analog approve/reject decisions. "
            f"If omitted, {DEFAULT_REVIEWED_ANALOG_DECISIONS_PATH} will be used when it exists."
        ),
    )
    return parser.parse_args()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_candidate_code(value: object) -> str:
    return clean_text(value).upper()


def build_review_query_keys(*parts: object) -> tuple[str, ...]:
    keys: list[str] = []
    for part in parts:
        normalized = build_search_text(part)
        if normalized and normalized not in keys:
            keys.append(normalized)
    return tuple(keys)


def load_reviewed_analog_decisions(path: Path | None) -> dict[str, dict[str, str]]:
    if path is None or not path.exists():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    decisions_by_query: dict[str, dict[str, str]] = defaultdict(dict)
    for raw in payload.get("decisions", []):
        query_key = clean_text(raw.get("query_key"))
        if not query_key:
            query_key = build_search_text(raw.get("query", ""))
        candidate_code = normalize_candidate_code(raw.get("candidate_code"))
        decision = clean_text(raw.get("decision")).lower()
        if not query_key or not candidate_code or decision not in {REVIEW_DECISION_APPROVED, REVIEW_DECISION_REJECTED}:
            continue
        current = decisions_by_query[query_key].get(candidate_code)
        if current == REVIEW_DECISION_REJECTED:
            continue
        if current == REVIEW_DECISION_APPROVED and decision == REVIEW_DECISION_REJECTED:
            decisions_by_query[query_key][candidate_code] = decision
            continue
        if current is None:
            decisions_by_query[query_key][candidate_code] = decision
    return {query_key: dict(code_map) for query_key, code_map in decisions_by_query.items()}


def hash_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def collect_file_info(path: Path | None) -> dict[str, object] | None:
    if path is None or not path.exists():
        return None
    stat = path.stat()
    return {
        "path": str(path),
        "sha256": hash_file(path),
        "size_bytes": stat.st_size,
        "modified_at": datetime.fromtimestamp(stat.st_mtime).astimezone().isoformat(),
    }


def load_json_payload(path: Path) -> dict[str, object]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def load_substitution_policy(path: Path | None = None) -> dict[str, object]:
    target_path = path or SUBSTITUTION_POLICY_PATH
    payload = load_json_payload(target_path)
    if not isinstance(payload.get("families"), dict):
        return {"metadata": payload.get("metadata", {}), "families": {}}
    return payload


def load_parser_hints(path: Path | None = None) -> dict[str, object]:
    target_path = path or PARSER_HINTS_PATH
    payload = load_json_payload(target_path)
    if not isinstance(payload.get("query_understanding"), Mapping):
        return {"metadata": payload.get("metadata", {}), "query_understanding": {}}
    return payload


PARSER_HINTS = load_parser_hints()


def build_combined_input_hash(file_infos: Sequence[dict[str, object] | None]) -> str:
    digest = hashlib.sha256()
    for info in file_infos:
        if not info:
            continue
        digest.update(str(info.get("path", "")).encode("utf-8"))
        digest.update(str(info.get("sha256", "")).encode("utf-8"))
    return digest.hexdigest()


def parse_quantity(value: object) -> float:
    text = clean_text(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def get_mapped_text(values: Sequence[object], column_index: int) -> str:
    if column_index <= 0 or column_index > len(values):
        return ""
    return clean_text(values[column_index - 1])


def get_mapped_quantity(values: Sequence[object], column_index: int) -> float:
    if column_index <= 0 or column_index > len(values):
        return 0.0
    return parse_quantity(values[column_index - 1])


def coalesce_material_fields(name: str, mark: str, supplier_code: str) -> tuple[str, str, str]:
    normalized_name = clean_text(name)
    normalized_mark = clean_text(mark)
    normalized_supplier_code = clean_text(supplier_code)
    if normalized_name:
        return normalized_name, normalized_mark, normalized_supplier_code
    if normalized_mark:
        return normalized_mark, "", normalized_supplier_code
    if normalized_supplier_code and re.search(r"[A-Za-zА-Яа-яЁёІіЇїЄєҐґ]", normalized_supplier_code):
        return normalized_supplier_code, "", ""
    return normalized_name, normalized_mark, normalized_supplier_code


def format_number(value: float) -> int | float:
    if math.isclose(value, round(value)):
        return int(round(value))
    return round(value, 3)


def parse_price(value: str) -> int | float | str:
    numeric = parse_quantity(value)
    if numeric <= 0:
        return ""
    return format_number(numeric)


def fix_mixed_script_token(token: str) -> str:
    if not token:
        return token
    if CYRILLIC_RE.search(token) and LATIN_RE.search(token):
        return "".join(PARSER_MIXED_CHAR_MAP.get(char, char) for char in token)
    return token


def normalize_symbols(text: str) -> str:
    text = text.replace("\n", " ")
    text = text.replace("∅", " dn ")
    text = text.replace("ø", " dn ")
    text = text.replace("Ø", " dn ")
    text = text.replace("½", "1/2").replace("¼", "1/4").replace("¾", "3/4")
    text = text.replace("°", " deg ")
    text = text.replace("№", " no ")
    text = text.replace("”", '"').replace("“", '"').replace("″", '"')
    text = text.replace("−", "-").replace("–", "-").replace("—", "-")
    text = text.replace("×", "x")
    text = re.sub(r"(?<=\d),(?=\d)", ".", text)
    text = re.sub(r"(\d)\s*\*\s*(\d)", r"\1x\2", text)
    text = re.sub(r"(\d)\s*[xхXХ]\s*(\d)", r"\1x\2", text)
    return text


def initialize_parser_hints() -> None:
    global PARSER_QUERY_HINTS
    global PARSER_PREPROCESSING_HINTS
    global PARSER_MIXED_SCRIPT_HINTS
    global PARSER_TOKEN_ALIASES
    global PARSER_FAMILY_TRIGGERS
    global PARSER_MIXED_CHAR_MAP
    global PARSER_IGNORE_FULL_ROW_PATTERNS
    global PARSER_STRIP_LEADING_PATTERNS
    global PARSER_STRIP_TRAILING_PATTERNS

    PARSER_QUERY_HINTS = (
        PARSER_HINTS.get("query_understanding", {}) if isinstance(PARSER_HINTS.get("query_understanding"), Mapping) else {}
    )
    PARSER_PREPROCESSING_HINTS = (
        PARSER_QUERY_HINTS.get("preprocessing", {})
        if isinstance(PARSER_QUERY_HINTS.get("preprocessing"), Mapping)
        else {}
    )
    PARSER_MIXED_SCRIPT_HINTS = (
        PARSER_QUERY_HINTS.get("mixed_script_normalization", {})
        if isinstance(PARSER_QUERY_HINTS.get("mixed_script_normalization"), Mapping)
        else {}
    )
    PARSER_TOKEN_ALIASES = _compile_parser_token_aliases(
        PARSER_QUERY_HINTS.get("token_aliases", {})
        if isinstance(PARSER_QUERY_HINTS.get("token_aliases"), Mapping)
        else {}
    )
    PARSER_FAMILY_TRIGGERS = _compile_parser_family_triggers(
        PARSER_QUERY_HINTS.get("family_triggers", {})
        if isinstance(PARSER_QUERY_HINTS.get("family_triggers"), Mapping)
        else {}
    )
    PARSER_MIXED_CHAR_MAP = dict(LATIN_TO_CYR_MIXED)
    for key, value in (
        PARSER_MIXED_SCRIPT_HINTS.get("char_map", {})
        if isinstance(PARSER_MIXED_SCRIPT_HINTS.get("char_map"), Mapping)
        else {}
    ).items():
        if isinstance(key, str) and len(key) == 1 and isinstance(value, str) and len(value) == 1:
            PARSER_MIXED_CHAR_MAP[key] = value
    PARSER_IGNORE_FULL_ROW_PATTERNS = tuple(
        pattern
        for pattern in PARSER_PREPROCESSING_HINTS.get("ignore_full_row_patterns", [])
        if isinstance(pattern, str) and pattern.strip()
    )
    PARSER_STRIP_LEADING_PATTERNS = tuple(
        pattern
        for pattern in PARSER_PREPROCESSING_HINTS.get("strip_leading_patterns", [])
        if isinstance(pattern, str) and pattern.strip()
    )
    PARSER_STRIP_TRAILING_PATTERNS = tuple(
        pattern
        for pattern in PARSER_PREPROCESSING_HINTS.get("strip_trailing_patterns", [])
        if isinstance(pattern, str) and pattern.strip()
    )


initialize_parser_hints()


def transliterate_token(token: str) -> str:
    fixed = fix_mixed_script_token(token)
    fixed = normalize_symbols(fixed).lower()
    buffer: list[str] = []
    for char in fixed:
        if char in CYR_TO_LAT:
            buffer.append(CYR_TO_LAT[char])
        elif char.isascii():
            buffer.append(char.lower())
        elif char.isdigit():
            buffer.append(char)
        elif char in {"/", ".", "+", "-", '"', "="}:
            buffer.append(char)
        elif char.isspace():
            buffer.append(" ")
    joined = "".join(buffer)
    joined = re.sub(r"\bdu(?=\d)", "dn", joined)
    joined = re.sub(r"\bdу(?=\d)", "dn", joined)
    joined = re.sub(r"\bdн(?=\d)", "dn", joined)
    joined = re.sub(r"\bru(?=\d)", "pn", joined)
    joined = re.sub(r"\brу(?=\d)", "pn", joined)
    joined = joined.replace("мм", "mm")
    joined = re.sub(r"\s+", " ", joined)
    return joined.strip(" .-")


def normalize_parser_line_text(value: object) -> str:
    text = clean_text(value).replace("\xa0", " ")
    if not text:
        return ""
    if bool(PARSER_PREPROCESSING_HINTS.get("space_dash_normalization", True)):
        text = text.replace("−", "-").replace("–", "-").replace("—", "-")
        text = re.sub(r"\s*-\s*", "-", text)
    if bool(PARSER_PREPROCESSING_HINTS.get("normalize_degree_sign", True)):
        text = text.replace("º", "°")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def strip_parser_body_noise(text: str, *, strip_trailing: bool = True) -> str:
    normalized = normalize_parser_line_text(text)
    if not normalized:
        return ""
    for pattern in PARSER_STRIP_LEADING_PATTERNS:
        normalized = re.sub(pattern, "", normalized, flags=re.IGNORECASE).strip()
    if strip_trailing:
        for pattern in PARSER_STRIP_TRAILING_PATTERNS:
            normalized = re.sub(pattern, "", normalized, flags=re.IGNORECASE).strip(" ,-;")
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip()


def should_ignore_parser_row(text: str) -> bool:
    normalized = _normalize_hint_phrase(text)
    if not normalized:
        return True
    return any(re.search(pattern, normalized, flags=re.IGNORECASE) for pattern in PARSER_IGNORE_FULL_ROW_PATTERNS)


def normalize_structured_tag_value(value: object) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    text = normalize_symbols(clean_text(value)).lower().replace("ё", "е")
    chunks = [transliterate_token(chunk) for chunk in text.split() if transliterate_token(chunk)]
    normalized = "_".join(chunks)
    normalized = re.sub(r"[^a-z0-9_.:/-]+", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized


def extract_parser_alias_tokens(*parts: object) -> set[str]:
    normalized = _normalize_hint_phrase(" ".join(clean_text(part) for part in parts if clean_text(part)))
    if not normalized:
        return set()
    matched: set[str] = set()
    for canonical, aliases in PARSER_TOKEN_ALIASES.items():
        if any(alias in normalized for alias in aliases):
            matched.add(canonical)
    return matched


def extract_parser_hint_tags(*parts: object) -> set[str]:
    normalized = _normalize_hint_phrase(" ".join(clean_text(part) for part in parts if clean_text(part)))
    if not normalized:
        return set()

    tags: set[str] = set()
    family_tags = extract_family_tags(*parts)
    primary_families = extract_tag_values(family_tags, "family:")
    starts_with_gearbox = bool(re.match(r"^\s*(?:\(рвз\)\s*)?редуктор\b", normalized))
    armature_families = {"zatvor", "zadvizhka", "kran", "klapan"}

    for canonical in extract_parser_alias_tokens(normalized):
        if canonical == "gearbox_operator" and starts_with_gearbox:
            continue
        tags.update(PARSER_ALIAS_TAGS.get(canonical, ()))

    for family, payload in PARSER_FAMILY_TRIGGERS.items():
        aliases = payload.get("aliases", ())
        canonical_family = PARSER_FAMILY_CANONICAL.get(family, family)
        matched_family_alias = any(alias in normalized for alias in aliases)
        if matched_family_alias:
            if canonical_family == "gearbox":
                if starts_with_gearbox:
                    tags.add("family:gearbox")
            else:
                tags.add(f"family:{canonical_family}")
        for resolve_rule in payload.get("resolve_rules", ()):
            match_any = resolve_rule.get("match_any", ())
            if not any(alias in normalized for alias in match_any):
                continue
            for key, value in resolve_rule.get("set", {}).items():
                if value is None:
                    continue
                normalized_value = normalize_structured_tag_value(value)
                if not normalized_value:
                    continue
                if key == "family":
                    normalized_value = PARSER_FAMILY_CANONICAL.get(normalized_value, normalized_value)
                    if normalized_value == "gearbox" and not starts_with_gearbox:
                        continue
                tags.add(f"{key}:{normalized_value}")

    if "муфта" in normalized and re.search(r"\bперех(?:од\w*)?\b", normalized):
        tags.add("family:perehod")
        tags.add("subfamily:coupling")
        if re.search(
            r"\b(?:вн/?нар|нар/?вн|вн/?нр|вр/?нр|вн/?вр|внутр(?:енняя)?\s*[-/]?\s*наруж(?:ная)?)\b",
            normalized,
        ):
            tags.add("subtype:inner_outer")

    if "valfex" in normalized and re.search(r"\b(?:бел\w*|ppr|pp-r)\b", normalized):
        if tags & {"family:mufta", "family:perehod", "family:tee", "family:elbow"}:
            tags.add("mclass:polypropylene")

    if starts_with_gearbox:
        tags.add("family:gearbox")
    elif primary_families & armature_families and re.search(r"(?:\bс\s+редуктором\b|\bредуктор\b)", normalized):
        tags.add("operator:gearbox")

    for family in ("zadvizhka", "kran", "zatvor"):
        if f"family:{family}" not in family_tags:
            continue
        if re.search(r"\b(?:с\s+рукояткой|рукоятка|ручка)\b", normalized):
            tags.add("operator:handle")
        if re.search(r"\b(?:с\s+электроприводом|под\s+электропривод)\b", normalized):
            tags.add("operator:electric")
        if re.search(r"\b(?:голый\s+шток|под\s+голый\s+шток)\b", normalized):
            tags.add("operator:bare_stem")

    if "family:gearbox" in tags:
        range_members: set[str] = set()
        for match in re.findall(r"\bf\s*0?(\d{2})\b", normalized):
            tags.add(f"mounting_pad:f{match}")
        for first, second in re.findall(r"\b(\d{1,2})\s*[xх]\s*(\d{1,2})\b", normalized):
            tags.add(f"stem_square:{first}x{second}")
        for left, right in re.findall(
            r"\b(?:dn|du|dy|ду)\s*(\d{1,4})\s*[-/]\s*(\d{1,4})\b",
            normalized,
        ):
            tags.add(f"compatible_dn:{left}-{right}")
            range_members.update({left, right})
        for match in re.findall(r"\b(?:dn|du|dy|ду)\s*(\d{1,4})\b", normalized):
            if match in range_members:
                continue
            tags.add(f"compatible_dn:{match}")
        if re.search(r"\bчетвертьоборот", normalized):
            tags.add("turn_type:quarter_turn")
        if re.search(r"\bконическ", normalized):
            tags.add("subtype:bevel")
        if re.search(r"\bд/затв(?:ора)?\b|\bдля\s+затвор", normalized):
            tags.add("for_family:zatvor")
        if "benarmo euro" in normalized:
            tags.add("brand:benarmo_euro")
        if "ridval" in normalized:
            tags.add("brand:ridval")

    for pattern, key in (
        (r"\bкорпус[-\s:]*чугун\b", "body_material:cast_iron"),
        (r"\bдиск[-\s:]*чугун\b", "disc_material:cast_iron"),
    ):
        if re.search(pattern, normalized):
            tags.add(key)

    for start, end in re.findall(
        r"\b([0-9]+(?:[.,][0-9]+)?)\s*(?:\.\.|-)\s*([0-9]+(?:[.,][0-9]+)?)\s*(?:мпа|mpa)\b",
        normalized,
    ):
        tags.add(f"pressure_range_mpa:{normalize_measure_value(start)}-{normalize_measure_value(end)}")
        tags.add("family:manometer")

    return tags


def build_search_text(*parts: object) -> str:
    tokens: list[str] = []
    for part in parts:
        raw = normalize_symbols(clean_text(part))
        if not raw:
            continue
        for alias_token in sorted(extract_parser_alias_tokens(raw)):
            tokens.append(alias_token)
        for token in TOKEN_RE.findall(raw):
            normalized = transliterate_token(token)
            if normalized:
                tokens.extend(expand_search_token_variants(normalized))
    return " ".join(tokens)


def normalize_measure_value(value: str) -> str:
    cleaned = value.replace(",", ".").strip()
    if "." in cleaned:
        cleaned = cleaned.rstrip("0").rstrip(".")
    return cleaned


def normalize_flange_face(value: str) -> str:
    face = value.lower().strip()
    if face == "v":
        return "b"
    return face


def expand_search_token_variants(token: str) -> list[str]:
    variants: list[str] = [token]

    if token == "vgp":
        variants.append("vodogazoprovodnaya")
    elif token == "esh":
        variants.append("elektrosvarnaya")

    standard_match = re.fullmatch(r"(?P<std>\d{5})(?:-(?P<year>\d{2,4}))?", token)
    if standard_match:
        std_num = standard_match.group("std")
        variants.append(f"spec{std_num}")
        # Legacy flange GOSTs superseded by ГОСТ 33259-2015
        if std_num == "12820":  # плоские приварные → тип 01
            variants.extend(["spec33259", "type01", "subtypeflat"])
        elif std_num == "12821":  # воротниковые → тип 11
            variants.extend(["spec33259", "type11", "subtypeweld"])
    else:
        embedded_standard_match = re.match(r"(?P<std>\d{5})(?:[-/].*)?$", token)
        if embedded_standard_match:
            variants.append(f"spec{embedded_standard_match.group('std')}")

    material_match = re.fullmatch(r"(?:st|ct)\.?(?P<material>[0-9][0-9a-z]*)", token)
    if material_match:
        variants.append(f"mat{material_match.group('material')}")

    simple_size = re.fullmatch(r"(?P<od>\d{1,4}(?:\.\d+)?)x(?P<wall>\d+(?:\.\d+)?)", token)
    if simple_size:
        wall_value = float(simple_size.group("wall"))
        if wall_value <= 12.0:
            variants.append(f"od{simple_size.group('od')}")
            variants.append(f"wall{normalize_measure_value(simple_size.group('wall'))}")

    hyphen_size = re.fullmatch(r"(?P<od>\d{1,4}(?:\.\d+)?)-(?P<wall>\d+(?:\.\d+)?)", token)
    if hyphen_size:
        wall_value = float(hyphen_size.group("wall"))
        if wall_value <= 12.0:
            variants.append(f"{hyphen_size.group('od')}x{hyphen_size.group('wall')}")
            variants.append(f"od{hyphen_size.group('od')}")
            variants.append(f"wall{normalize_measure_value(hyphen_size.group('wall'))}")

    compound_size = re.fullmatch(
        r"(?P<angle>15|30|45|60|87|90|120|180)-(?P<series>\d+)-(?P<od>\d{2,4})x(?P<wall>\d+(?:\.\d+)?)(?:-(?P<material>[0-9a-z]+))?",
        token,
    )
    if compound_size:
        variants.append(f"deg{compound_size.group('angle')}")
        variants.append(f"series{compound_size.group('series')}")
        variants.append(f"od{compound_size.group('od')}")
        variants.append(f"wall{normalize_measure_value(compound_size.group('wall'))}")
        material = compound_size.group("material")
        if material:
            variants.append(f"mat{material}")

    full_flange_execution = re.fullmatch(
        r"(?P<dn>\d{2,4})-(?P<pn>\d{2,4})-(?P<type>01|11)(?:-(?P<execution>[12]))?(?:-(?P<face>[a-z]))?(?:-.*)?",
        token,
    )
    if full_flange_execution:
        variants.append(f"dn{full_flange_execution.group('dn')}")
        variants.append(f"pn{full_flange_execution.group('pn')}")
        variants.append(f"type{full_flange_execution.group('type')}")
        execution = full_flange_execution.group("execution")
        if execution:
            variants.append(f"exec{execution}")
        face = full_flange_execution.group("face")
        if face:
            variants.append(f"face{normalize_flange_face(face)}")

    short_flange_execution = re.fullmatch(
        r"(?P<pn>\d{2,4})-(?P<type>01|11)(?:-(?P<execution>[12]))?(?:-(?P<face>[a-z]))?(?:-.*)?",
        token,
    )
    if short_flange_execution:
        variants.append(f"pn{short_flange_execution.group('pn')}")
        variants.append(f"type{short_flange_execution.group('type')}")
        execution = short_flange_execution.group("execution")
        if execution:
            variants.append(f"exec{execution}")
        face = short_flange_execution.group("face")
        if face:
            variants.append(f"face{normalize_flange_face(face)}")

    if token.count("-") >= 2:
        fragments = [fragment for fragment in token.split("-") if fragment]
        for fragment in fragments:
            variants.append(fragment)
            material_match = re.fullmatch(r"(?:st|ct)\.?(?P<material>[0-9][0-9a-z]*)", fragment)
            if material_match:
                variants.append(f"mat{material_match.group('material')}")
            if fragment.isdigit() and fragment in {"8", "12"}:
                variants.append(f"hole{fragment}")
            if len(fragment) == 1 and fragment.isalpha():
                variants.append(f"face{normalize_flange_face(fragment)}")

    deduped: list[str] = []
    seen: set[str] = set()
    for value in variants:
        if value and value not in seen:
            deduped.append(value)
            seen.add(value)
    return deduped


def extract_structured_tags_from_search_text(search_text: str) -> set[str]:
    tags: set[str] = set()
    for token in search_text.split():
        if token.startswith("deg") and token[3:].isdigit():
            tags.add(f"deg:{token[3:]}")
        elif token.startswith("series") and token[6:].isdigit():
            tags.add(f"series:{token[6:]}")
        elif token.startswith("od") and token[2:].isdigit():
            tags.add(f"od:{token[2:]}")
        elif token.startswith("wall") and token[4:]:
            tags.add(f"wall:{normalize_measure_value(token[4:])}")
        elif token.startswith("mat") and token[3:]:
            tags.add(f"mat:{token[3:]}")
        elif token.startswith("type") and token[4:].isdigit():
            tags.add(f"type:{token[4:]}")
        elif token.startswith("exec") and token[4:].isdigit():
            tags.add(f"exec:{token[4:]}")
        elif token.startswith("face") and token[4:]:
            tags.add(f"face:{normalize_flange_face(token[4:])}")
        elif token.startswith("hole") and token[4:].isdigit():
            tags.add(f"hole:{token[4:]}")
        elif token.startswith("spec") and token[4:].isdigit():
            tags.add(f"spec:{token[4:]}")
    return tags


def augment_search_text_with_dimension_tags(search_text: str, dimension_tags: Iterable[str]) -> str:
    tokens = [token for token in search_text.split() if token]
    seen = set(tokens)
    extras: list[str] = []
    grouped: dict[str, set[str]] = defaultdict(set)
    for tag in dimension_tags:
        key, _, value = tag.partition(":")
        if value:
            grouped[key].add(value)

    def add(token: str) -> None:
        if token and token not in seen:
            extras.append(token)
            seen.add(token)

    for key in ("dn", "pn", "type", "exec", "face", "hole", "mat", "spec", "od", "wall", "deg", "series", "kvs", "lmm"):
        for value in sorted(grouped.get(key, set())):
            normalized = normalize_measure_value(value) if key in {"dn", "pn", "od", "wall", "mat", "kvs", "lmm"} else value
            add(f"{key}{normalized}")

    for key in ("pairdn", "tripledn"):
        for value in sorted(grouped.get(key, set())):
            add(f"{key}{value}")

    for value in sorted(grouped.get("family", set())):
        add(f"family_{value}")

    for key in (
        "subfamily",
        "subtype",
        "shape",
        "connection",
        "operator",
        "for_family",
        "mounting_pad",
        "stem_square",
        "compatible_dn",
        "turn_type",
        "actuator",
        "handle",
        "extra_connection",
        "brand",
        "body_material",
        "disc_material",
        "conn_gender",
        "meter_type",
        "interface",
        "pulse_output",
        "pressure_range_mpa",
        "connection_guess",
        "connection_position",
        "accuracy_class",
        "orientation",
        "face_to_face",
        "has_magnet",
        "filter_element",
    ):
        for value in sorted(grouped.get(key, set())):
            add(f"{key}_{value}")

    if "family:flange" in dimension_tags and "vorotnikovyi" in seen:
        add("privarnoi")
        add("vstyk")
    if "grade:cs20" in dimension_tags or "mat:20" in dimension_tags:
        add("stalnoi")

    return " ".join(tokens + extras)


def extract_tokens(search_text: str) -> set[str]:
    return {token for token in search_text.split() if token}


def extract_key_tokens(search_tokens: Iterable[str]) -> set[str]:
    result: set[str] = set()
    for token in search_tokens:
        if len(token) < 3:
            continue
        if token in COMMON_TOKENS:
            continue
        if token.isdigit():
            continue
        result.add(token)
    return result


def token_root(token: str) -> str:
    letters_only = re.sub(r"[^a-zа-я]", "", token.lower())
    if len(letters_only) < 4:
        return token
    return letters_only[:5]


def extract_root_tokens(tokens: Iterable[str]) -> set[str]:
    return {token_root(token) for token in tokens if token}


def extract_code_tokens(*parts: object) -> set[str]:
    tokens: set[str] = set()
    # Material grade patterns to exclude from code tokens — these are NOT product codes
    _MATERIAL_GRADE_RE = re.compile(
        r"^(?:ct\.?\d+|st\.?\d+|09g2s|12x18h10t|12kh18n10t|aisi\d+|a351|cf8[m]?|ggg\d+|"
        r"vch\d+|sch\d+|l?63|cw617n|cuzn\d+)$",
        re.IGNORECASE,
    )
    for part in parts:
        raw = normalize_symbols(clean_text(part))
        if not raw:
            continue
        for token in TOKEN_RE.findall(raw):
            has_digit = bool(DIGIT_RE.search(token))
            has_letter = bool(re.search(r"[A-Za-zА-Яа-яЁёІіЇїЄєҐґ]", token))
            # Allow pure-uppercase abbreviations >= 3 chars (КШЦП, КШЦМ, КШЦПР, etc.)
            is_uppercase_abbrev = bool(re.fullmatch(r"[A-ZА-ЯЁ]{3,}", token))
            if not has_letter:
                continue
            if not has_digit and not is_uppercase_abbrev:
                continue
            skeleton_chars: list[str] = []
            for char in fix_mixed_script_token(token):
                if char.isdigit():
                    skeleton_chars.append(char)
                    continue
                if char in {"/", ".", "+", "-", '"', "="}:
                    skeleton_chars.append(char)
                    continue
                if char in VISUAL_CODE_MAP:
                    skeleton_chars.append(VISUAL_CODE_MAP[char])
                    continue
                lower = char.lower()
                if lower in CYR_TO_LAT:
                    skeleton_chars.append(CYR_TO_LAT[lower])
                    continue
                if char.isascii():
                    skeleton_chars.append(lower)
            skeleton = re.sub(r"[^a-z0-9./+=\"-]", "", "".join(skeleton_chars))
            if re.fullmatch(r"\d+[xх]\d+(?:[xх]\d+)?", skeleton):
                continue
            if skeleton.startswith(("dn", "du", "pn", "ru", "kvs", "l=")):
                continue
            if re.fullmatch(r"[a-z]{1,2}\d+(?:[./-]\d+)*", skeleton) and skeleton[:2] in {"dn", "du", "pn", "ru"}:
                continue
            if len(skeleton) >= 3:
                # Filter out material grades — they are NOT product codes
                if _MATERIAL_GRADE_RE.fullmatch(skeleton):
                    continue
                tokens.add(skeleton)
    return tokens


def normalize_measure_text(*parts: object) -> str:
    text = normalize_symbols(" ".join(clean_text(part) for part in parts if clean_text(part)))
    text = " ".join(fix_mixed_script_token(token) for token in TOKEN_RE.findall(text))
    text = text.lower().replace("ё", "е")
    text = re.sub(r"\b(?:ду|dу|dн)\.?\s*(?=\d)", "dn ", text)
    text = re.sub(r"\b(?:ру|rу)\.?\s*(?=\d)", "pn ", text)
    text = text.replace("мм", "mm")
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_fraction(value: str) -> str:
    return re.sub(r"\s+", "", value.replace(",", "."))


def normalize_pair_tag(left: str, right: str) -> str:
    left_value = left.replace(",", ".")
    right_value = right.replace(",", ".")
    ordered = sorted((left_value, right_value), key=lambda value: float(value))
    return f"{ordered[0]}x{ordered[1]}"


def normalize_multi_measure_tag(*values: str) -> str:
    normalized = [normalize_measure_value(value) for value in values]
    ordered = sorted(normalized, key=lambda value: float(value))
    return "x".join(ordered)


def extract_dimension_tags(*parts: object) -> set[str]:
    raw_text = normalize_symbols(" ".join(clean_text(part) for part in parts if clean_text(part))).lower()
    text = normalize_measure_text(*parts)
    search_text = build_search_text(*parts)
    text_without_grade_sizes = re.sub(
        r"\b(?:0[38]|12)\d?\s*[xхh]\s*\d{2}\s*[hn]\s*\d{2}(?:\s*[tm]\s*\d?)?\b",
        " ",
        text,
    )
    tags: set[str] = set()
    for match in re.findall(r"\bdn(?:\.|-)?\s*([0-9]+(?:[.,][0-9]+)?)", text):
        tags.add(f"dn:{match.replace(',', '.')}")
    for match in re.findall(r"\b(?:du|dy|ду)\s*([0-9]+(?:[.,][0-9]+)?)\b", raw_text):
        tags.add(f"dn:{match.replace(',', '.')}")
    for match in re.findall(r"\bd(?!n)(?!u)(?!y)\s*([0-9]+(?:[.,][0-9]+)?)\b", text):
        normalized = match.replace(",", ".")
        tags.add(f"dn:{normalized}")
        if "." not in normalized:
            tags.add(f"od:{normalized}")
    for match in re.findall(r"\b(?:pn|ru)(?:\.|-)?\s*([0-9]+(?:[.,][0-9]+)?)", text):
        tags.add(f"pn:{match.replace(',', '.')}")
    for match in re.findall(r"\bkvs\s*=?\s*([0-9]+(?:[.,][0-9]+)?)", text):
        tags.add(f"kvs:{match.replace(',', '.')}")
    for match in re.findall(r"\bl\.?\s*=?\s*([0-9]+(?:[.,][0-9]+)?)\s*mm", text):
        tags.add(f"lmm:{match.replace(',', '.')}")
    # Capture L=NNN without explicit mm unit (common in valve specs like КШЦП)
    for match in re.findall(r"\bl\s*=\s*([0-9]+(?:[.,][0-9]+)?)(?!\s*mm)(?=\s|$)", text):
        tags.add(f"lmm:{match.replace(',', '.')}")
    for match in re.finditer(r"\b([0-9]+(?:[.,][0-9]+)?)\s*(?:mm|мм)\b", text):
        prefix = text[max(0, match.start() - 6) : match.start()]
        if re.search(r"(?:\bl\s*=?\s*|\bl=)$", prefix):
            continue
        normalized = match.group(1).replace(",", ".")
        tags.add(f"dn:{normalized}")
        tags.add(f"od:{normalized}")
    for match in re.findall(r"\b([0-9]+\s+[0-9]+/[0-9]+|[0-9]+/[0-9]+)\s*(?:\"|''|'|inch|дюйм)", text):
        tags.add(f'inch:{normalize_fraction(match)}')
    if any(marker in raw_text for marker in {'"', "''", "'", "дюйм"}):
        for match in re.findall(r"\b([0-9]+\s+[0-9]+/[0-9]+|[0-9]+/[0-9]+)\b", search_text):
            tags.add(f'inch:{normalize_fraction(match)}')
    for match in re.findall(r"\b(15|30|45|60|87|90|120|180)\s*(?:deg|град|grad|-)", text):
        tags.add(f"deg:{match}")
    for dn_value, deg_value in re.findall(
        r"\b(?:dn|du|dy|ду)(?:\.|-)?\s*(\d{2,4})\s*/\s*(15|30|45|60|87|90|120|180)\b",
        raw_text,
    ):
        tags.add(f"dn:{dn_value}")
        tags.add(f"deg:{deg_value}")
    for outer, wall in re.findall(
        r"(?<![a-zа-я0-9])(\d{1,4}(?:[.,]\d+)?)\s*[xхh]\s*(\d+(?:[.,]\d+)?)(?![a-zа-я0-9])",
        text_without_grade_sizes,
    ):
        normalized_outer = outer.replace(",", ".")
        normalized_wall = wall.replace(",", ".")
        if float(normalized_wall) > 12.0:
            continue
        tags.add(f"od:{normalized_outer}")
        tags.add(f"wall:{normalized_wall}")
    for outer, wall in re.findall(
        r"(?<![a-zа-я0-9])(\d{1,4}(?:[.,]\d+)?)\s*-\s*(\d+(?:[.,]\d+)?)(?![a-zа-я0-9])",
        text_without_grade_sizes,
    ):
        normalized_outer = outer.replace(",", ".")
        normalized_wall = wall.replace(",", ".")
        if float(normalized_wall) > 12.0:
            continue
        tags.add(f"od:{normalized_outer}")
        tags.add(f"wall:{normalized_wall}")
        if re.search(r"\b(?:vgp|vodogazoprovod|3262)\b", search_text):
            tags.add(f"dn:{normalized_outer}")
    for match in re.findall(r"\b(?:st|ct|ст)\.?\s*([0-9][0-9a-z]*)\b", text):
        tags.add(f"mat:{match}")
    for match in re.findall(r"\b(?:gost|гост)\s*([0-9]{5})\b", text):
        tags.add(f"spec:{match}")
    # Legacy flange GOSTs → map to current ГОСТ 33259-2015
    if re.search(r"\b(?:gost|гост)\s*12820\b", text):
        tags.update({"spec:33259", "family:flange", "type:01"})
    if re.search(r"\b(?:gost|гост)\s*12821\b", text):
        tags.update({"spec:33259", "family:flange", "type:11"})
    for match in re.findall(r"\b(15|30|45|60|87|90|120|180)\s*-\s*([12])\s*-\s*\d{2,4}\s*[xхh]", text):
        tags.add(f"deg:{match[0]}")
        tags.add(f"series:{match[1]}")
    for flange_type in re.findall(r"\b(?:tip|type|тип)\s*(01|11)\b", text):
        tags.add(f"type:{flange_type}")
    for dn, pn, flange_type in re.findall(r"\b(\d{2,4})\s*-\s*(\d{2,4})\s*-\s*(01|11)\b", text):
        tags.add(f"dn:{dn}")
        tags.add(f"pn:{pn}")
        tags.add(f"type:{flange_type}")
    for dn, pn, flange_type, execution in re.findall(
        r"\b(\d{2,4})\s*-\s*(\d{2,4})\s*-\s*(01|11)\s*-\s*([12])\b",
        text,
    ):
        tags.add(f"dn:{dn}")
        tags.add(f"pn:{pn}")
        tags.add(f"type:{flange_type}")
        tags.add(f"exec:{execution}")
    for pn, flange_type, execution, face in re.findall(
        r"\b(\d{2,4})\s*-\s*(01|11)\s*-\s*([12])\s*-\s*([a-zа-я])\b",
        text,
    ):
        tags.add(f"pn:{pn}")
        tags.add(f"type:{flange_type}")
        tags.add(f"exec:{execution}")
        tags.add(f"face:{normalize_flange_face(transliterate_token(face))}")
    if re.search(r"\b8\s+otverst", text):
        tags.add("hole:8")
    if re.search(r"\b(?:troinik|tee)\b", search_text):
        for first, second, third in re.findall(
            r"\b(?:dn|d)?\s*(15|20|25|32|40|50|65|80|100|125|150|200|250|300)\s*[xх/]\s*"
            r"(15|20|25|32|40|50|65|80|100|125|150|200|250|300)\s*[xх/]\s*"
            r"(15|20|25|32|40|50|65|80|100|125|150|200|250|300)\b",
            search_text,
        ):
            tags.add(f"tripledn:{normalize_multi_measure_tag(first, second, third)}")
    if re.search(r"\bpereh(?:od\w*)?\b", search_text):
        for left, right in re.findall(
            r"\b(?:dn|d)?\s*(15|20|25|32|40|50|65|80|100|125|150|200|250|300)\s*(?:[xх/]|-)\s*"
            r"(15|20|25|32|40|50|65|80|100|125|150|200|250|300)\b",
            search_text,
        ):
            if left != right:
                tags.add(f"pairdn:{normalize_pair_tag(left, right)}")
        for left, right in re.findall(
            r"\b(\d{2,4}(?:[.,]\d+)?)\s*[xхh]\s*\d+(?:[.,]\d+)?\s*[-/]\s*"
            r"(\d{2,4}(?:[.,]\d+)?)\s*[xхh]\s*\d+(?:[.,]\d+)?\b",
            text_without_grade_sizes,
        ):
            if left != right:
                tags.add(f"pairdn:{normalize_pair_tag(left, right)}")
    if re.search(r"\b(?:kran|klapan|filtr|rezba|vozduhootvod|vozdnhootvod|zatvor|zadvizh)\b", search_text):
        for match in re.findall(r"\b[a-z0-9./]+-(15|20|25|32|40|50|65|80|100|125|150)\b", search_text):
            tags.add(f"dn:{match}")
    if re.search(r"\b(?:отвод|колено|муфта|тройник|заглушка|переход|фланец|кран|клапан|затвор|фильтр|труба)\b", text):
        for match in re.findall(r"(?:^|\s)(\d{2,4})(?:\s*mm|\s*мм)?$", text):
            tags.add(f"dn:{match}")
            if re.search(r"\b(?:truba|отвод|колено)\b", search_text):
                tags.add(f"od:{match}")

    if re.search(r"\b(?:vgp|vodogazoprovod|3262)\b", search_text):
        dn_values = extract_tag_values(tags, "dn:")
        od_values = extract_tag_values(tags, "od:")
        for dn_value in list(dn_values):
            mapped_od = VGP_DN_TO_OD.get(normalize_measure_value(dn_value))
            if mapped_od:
                tags.add(f"od:{mapped_od}")
                normalized_mapped_od = normalize_measure_value(mapped_od)
                if normalized_mapped_od != mapped_od:
                    tags.add(f"od:{normalized_mapped_od}")
        for od_value in list(od_values):
            normalized_od = normalize_measure_value(od_value)
            mapped_dn = VGP_OD_TO_DN.get(normalized_od)
            if mapped_dn:
                tags.add(f"dn:{mapped_dn}")
    return tags


def extract_family_tags(*parts: object) -> set[str]:
    tokens = extract_tokens(build_search_text(*parts))
    tags: set[str] = set()
    for family, patterns in FAMILY_PATTERNS.items():
        if any(
            any(
                token.startswith(pattern)
                or (pattern in FAMILY_CONTAINS_TOKEN_PATTERNS and pattern in token)
                for token in tokens
            )
            for pattern in patterns
        ):
            tags.add(f"family:{family}")
    return tags


def extract_material_tags_from_search_text(search_text: str) -> set[str]:
    token_list = [token for token in search_text.split() if token]
    tokens = set(token_list)
    tags: set[str] = set()
    token_string = " ".join(token_list)
    explicit_grade_hits: set[str] = set()

    for canonical_grade, aliases in GRADE_GROUPS.items():
        if tokens & aliases:
            explicit_grade_hits.add(canonical_grade)
            tags.add(f"grade:{canonical_grade}")
            material_class = GRADE_TO_CLASS.get(canonical_grade)
            if material_class:
                tags.add(f"mclass:{material_class}")

    for canonical_grade, fragments in GRADE_SUBSTRING_ALIASES.items():
        if canonical_grade == "cs20" and any(
            explicit_grade != "cs20" and GRADE_TO_CLASS.get(explicit_grade) == "steel"
            for explicit_grade in explicit_grade_hits
        ):
            continue
        if any(fragment in token_string for fragment in fragments):
            tags.add(f"grade:{canonical_grade}")
            material_class = GRADE_TO_CLASS.get(canonical_grade)
            if material_class:
                tags.add(f"mclass:{material_class}")

    for material_class, rule in MATERIAL_CLASS_RULES.items():
        contains = rule.get("contains", ())
        exact = rule.get("exact", set())
        prefixes = rule.get("prefix", ())
        matched = False
        for index, token in enumerate(token_list):
            previous = token_list[index - 1] if index > 0 else ""
            if previous in {"dlya", "pod"} and material_class in {"cast_iron", "polypropylene", "pvc", "polyethylene", "copper"}:
                continue
            if token in exact:
                matched = True
                break
            if any(token.startswith(fragment) for fragment in contains):
                matched = True
                break
            if any(token.startswith(prefix) for prefix in prefixes):
                matched = True
                break
        if matched:
            tags.add(f"mclass:{material_class}")

    if "mclass:steel" not in tags and not (
        {"mclass:stainless", "mclass:cast_iron", "mclass:polypropylene", "mclass:pvc", "mclass:polyethylene"} & tags
    ):
        steel_pipe_markers = {"vgp", "vodogazoprovodnaya", "elektrosvarnaya", "spec10704", "spec10705"}
        if "truba" in tokens and tokens & steel_pipe_markers:
            tags.add("mclass:steel")

    if "mclass:stainless" in tags and "mclass:steel" in tags:
        tags.remove("mclass:steel")
    return tags


def extract_tag_values(tags: Iterable[str], prefix: str) -> set[str]:
    values: set[str] = set()
    for tag in tags:
        if tag.startswith(prefix):
            values.add(tag[len(prefix) :])
    return values


def primary_family_tag(tags: Iterable[str]) -> str:
    families = extract_tag_values(tags, "family:")
    primary = sorted(families & PRIMARY_FAMILY_TAGS)
    if "perehod" in primary and "mufta" in primary:
        return "perehod"
    return primary[0] if primary else ""


def find_manual_review_rule(order: OrderLine) -> dict[str, object] | None:
    if (
        order.classification
        and order.classification.status != ClassificationStatus.UNCLASSIFIED
        and order.classification.route == "manual_review"
    ):
        return {
            "id": f"classifier:{order.classification.category_key or 'manual_review'}",
            "comment": order.classification.manual_review_comment
            or order.classification.reason
            or "требует ручного подбора",
        }

    query = f"{order.name} {order.mark} {order.vendor} {order.raw_query}".lower()
    search_text = order.search_text
    for rule in MANUAL_REVIEW_RULES:
        if not isinstance(rule, dict):
            continue
        patterns = rule.get("patterns", [])
        if not isinstance(patterns, list):
            continue
        for pattern in patterns:
            normalized_pattern = str(pattern).strip().lower()
            if not normalized_pattern:
                continue
            transliterated = build_search_text(normalized_pattern)
            if normalized_pattern in query or (transliterated and transliterated in search_text):
                return rule
    return None


def group_dimension_tags(tags: Iterable[str]) -> dict[str, set[str]]:
    grouped: dict[str, set[str]] = defaultdict(set)
    for tag in tags:
        key, _, value = tag.partition(":")
        if value:
            grouped[key].add(value)
    return grouped


def expand_dimension_values(grouped: dict[str, set[str]]) -> dict[str, set[str]]:
    expanded = {key: set(values) for key, values in grouped.items()}
    dn_values = expanded.setdefault("dn", set())
    inch_values = expanded.setdefault("inch", set())
    for inch_value in list(inch_values):
        mapped_dn = INCH_TO_DN.get(inch_value)
        if mapped_dn:
            dn_values.add(mapped_dn)
    for dn_value in list(dn_values):
        for inch_value, mapped_dn in INCH_TO_DN.items():
            if mapped_dn == dn_value:
                inch_values.add(inch_value)
    return expanded


def get_substitution_family_policy(
    *,
    order: OrderLine | None = None,
    candidate: Candidate | None = None,
    substitution_policy: Mapping[str, object] | None = None,
) -> tuple[str, Mapping[str, object]]:
    families_payload = substitution_policy.get("families", {}) if isinstance(substitution_policy, Mapping) else {}
    if not isinstance(families_payload, Mapping):
        return "", {}
    family = primary_family_tag(order.dimension_tags) if order is not None else ""
    if not family and candidate is not None:
        family = primary_family_tag(candidate.stock.dimension_tags)
    if not family:
        return "", {}
    policy = families_payload.get(family, {})
    if not isinstance(policy, Mapping):
        return family, {}
    return family, policy


def candidate_matches_policy_dimensions(order: OrderLine, candidate: Candidate, required_keys: Sequence[str]) -> bool:
    if not required_keys:
        return False
    order_family = primary_family_tag(order.dimension_tags)
    stock_family = primary_family_tag(candidate.stock.dimension_tags)
    order_dimensions = expand_dimension_values(group_dimension_tags(order.dimension_tags))
    stock_dimensions = expand_dimension_values(group_dimension_tags(candidate.stock.dimension_tags))
    for key in required_keys:
        if key == "family":
            if not order_family or not stock_family or order_family != stock_family:
                return False
            continue
        order_values = order_dimensions.get(key, set())
        stock_values = stock_dimensions.get(key, set())
        if not order_values or not stock_values or not (order_values & stock_values):
            return False
        if key in candidate.conflicting_dimension_keys:
            return False
    return True


def candidate_hits_forbidden_policy_tokens(
    order: OrderLine,
    candidate: Candidate,
    forbidden_prefixes: Sequence[str],
) -> bool:
    for prefix in forbidden_prefixes:
        if not prefix:
            continue
        if any(token.startswith(prefix) for token in candidate.stock.search_tokens) and not any(
            token.startswith(prefix) for token in order.search_tokens
        ):
            return True
    return False


def iter_scored_dimensions(grouped: dict[str, set[str]]) -> Iterable[tuple[str, set[str]]]:
    for key, values in grouped.items():
        if key not in DIMENSION_WEIGHTS:
            continue
        if not values:
            continue
        yield key, values


def flange_preference_bonus(order: OrderLine, stock: StockItem) -> tuple[float, list[str]]:
    if "family:flange" not in order.dimension_tags or "family:flange" not in stock.dimension_tags:
        return 0.0, []

    bonus = 0.0
    reasons: list[str] = []
    order_faces = extract_tag_values(order.dimension_tags, "face:")
    stock_faces = extract_tag_values(stock.dimension_tags, "face:")
    order_execs = extract_tag_values(order.dimension_tags, "exec:")
    stock_execs = extract_tag_values(stock.dimension_tags, "exec:")
    order_types = extract_tag_values(order.dimension_tags, "type:")
    stock_holes = extract_tag_values(stock.dimension_tags, "hole:")
    order_grades = extract_tag_values(order.dimension_tags, "grade:")
    stock_grades = extract_tag_values(stock.dimension_tags, "grade:")

    if order_faces and stock_faces:
        if order_faces & stock_faces:
            bonus += 4.0
            reasons.append("совпадает исполнение уплотнительной поверхности")
        else:
            bonus -= 6.0
            reasons.append("исполнение поверхности отличается")
    elif not order_faces and "b" in stock_faces:
        bonus += 3.0
        reasons.append("предпочтительное исполнение B")

    if order_execs and stock_execs:
        if order_execs & stock_execs:
            bonus += 4.0
            reasons.append("совпадает ряд исполнения")
        else:
            bonus -= 8.0
            reasons.append("ряд исполнения отличается")
    elif not order_execs and order_types == {"11"}:
        if "2" in stock_execs:
            bonus += 6.0
            reasons.append("предпочтительный ряд 11-2")
        elif "1" in stock_execs:
            bonus -= 2.0
            reasons.append("ряд 11-1 менее предпочтителен")

    if not order_execs and order_types == {"11"} and "8" in stock_holes:
        bonus += 4.0
        reasons.append("исполнение на 8 отверстий")

    if not order_execs and order_types == {"11"} and "2" in stock_execs and "8" in stock_holes:
        bonus += 4.0
        reasons.append("предпочтительное исполнение по ГОСТ 33259")

    if not order_grades and "cs20" in stock_grades:
        bonus += 4.0
        reasons.append("сталь 20 как базовое исполнение")
    elif order_grades and stock_grades and order_grades & stock_grades:
        bonus += 2.0
        reasons.append("совпадает марка стали")
    elif order_grades and stock_grades and not (order_grades & stock_grades):
        bonus -= 6.0
        reasons.append("марка стали отличается")

    return bonus, reasons


def is_exact_pipe_candidate(order: OrderLine, candidate: Candidate) -> bool:
    if "family:pipe" not in order.dimension_tags or "family:pipe" not in candidate.stock.dimension_tags:
        return False
    if candidate.dimension_penalty != 0:
        return False
    matched_keys = set(candidate.matched_dimension_keys)
    if not {"od", "wall"} <= matched_keys:
        return False
    order_specs = extract_tag_values(order.dimension_tags, "spec:")
    stock_specs = extract_tag_values(candidate.stock.dimension_tags, "spec:")
    if order_specs and not (order_specs & stock_specs):
        return False
    if candidate.overlap < 0.5 and candidate.soft_overlap < 0.6:
        return False
    if any(token.startswith("ocink") for token in candidate.stock.search_tokens) and not any(
        token.startswith("ocink") for token in order.search_tokens
    ):
        return False
    return True


def safe_overlap(left: set[str], right: set[str]) -> float:
    if not left:
        return 0.0
    return len(left & right) / len(left)


def token_set_ratio(left: Sequence[str], right: Sequence[str]) -> float:
    if not left or not right:
        return 0.0
    left_set = set(left)
    right_set = set(right)
    intersection = left_set & right_set
    if not intersection:
        return 0.0
    return 100.0 * (2 * len(intersection)) / (len(left_set) + len(right_set))


def ordered_subsequence_ratio(query: str, candidate: str) -> float:
    if not query or not candidate:
        return 0.0
    if query in candidate or candidate in query:
        return 100.0
    query_tokens = query.split()
    candidate_tokens = candidate.split()
    if not query_tokens or not candidate_tokens:
        return 0.0
    candidate_positions: defaultdict[str, list[int]] = defaultdict(list)
    for index, token in enumerate(candidate_tokens):
        candidate_positions[token].append(index)
    matches = 0
    current_index = -1
    for token in query_tokens:
        positions = candidate_positions.get(token)
        if not positions:
            continue
        next_position = next((pos for pos in positions if pos > current_index), None)
        if next_position is None:
            continue
        current_index = next_position
        matches += 1
    return 100.0 * matches / len(query_tokens)


def load_stock(stock_path: Path) -> list[StockItem]:
    items: list[StockItem] = []
    for index, row in enumerate(iter_stock_rows(stock_path), start=1):
        quantity = parse_quantity(row.get("Остаток"))
        if quantity <= 0:
            continue
        search_text = build_search_text(
            row.get("Номенклатура"),
            row.get("НаименованиДляПечати"),
            row.get("ТипПродукта"),
        )
        dimension_tags = (
            extract_dimension_tags(
                row.get("Номенклатура"),
                row.get("НаименованиДляПечати"),
                row.get("ТипПродукта"),
            )
            | extract_family_tags(
                row.get("Номенклатура"),
                row.get("НаименованиДляПечати"),
                row.get("ТипПродукта"),
            )
            | extract_parser_hint_tags(
                row.get("Номенклатура"),
                row.get("НаименованиДляПечати"),
                row.get("ТипПродукта"),
            )
            | extract_material_tags_from_search_text(search_text)
            | extract_structured_tags_from_search_text(search_text)
        )
        search_text = augment_search_text_with_dimension_tags(search_text, dimension_tags)
        search_tokens = extract_tokens(search_text)
        items.append(
            StockItem(
                row_index=index,
                code_1c=clean_text(row.get("Код1с")),
                name=clean_text(row.get("Номенклатура")),
                print_name=clean_text(row.get("НаименованиДляПечати")),
                product_type=clean_text(row.get("ТипПродукта")),
                sale_price=clean_text(row.get("ПродажнаяЦена")),
                stop_price=clean_text(row.get("СтопЦена")),
                plan_price=clean_text(row.get("ПлановаяЦена")),
                quantity=quantity,
                remaining=quantity,
                search_text=search_text,
                search_tokens=search_tokens,
                key_tokens=extract_key_tokens(search_tokens),
                root_tokens=extract_root_tokens(search_tokens),
                code_tokens=extract_code_tokens(
                    row.get("Номенклатура"),
                    row.get("НаименованиДляПечати"),
                    row.get("ТипПродукта"),
                ),
                dimension_tags=dimension_tags,
            )
        )
    return items


def iter_stock_rows(stock_path: Path) -> Iterable[dict[str, object]]:
    suffix = stock_path.suffix.lower()
    if suffix == ".csv":
        yield from iter_stock_rows_from_csv(stock_path)
        return
    if suffix in {".xlsx", ".xlsm"}:
        yield from iter_stock_rows_from_workbook(stock_path)
        return
    raise ValueError(f"Unsupported stock format: {stock_path.suffix}")


def iter_stock_rows_from_csv(stock_path: Path) -> Iterable[dict[str, object]]:
    with stock_path.open("r", encoding="cp1251", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=";")
        for row in reader:
            yield dict(row)


def iter_stock_rows_from_workbook(stock_path: Path) -> Iterable[dict[str, object]]:
    workbook = load_workbook(stock_path, data_only=True, read_only=True)
    sheet_name = "остатки" if "остатки" in workbook.sheetnames else workbook.sheetnames[0]
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration:
        return
    headers = [clean_text(value) for value in raw_headers]
    header_map = {
        header: index
        for index, header in enumerate(headers)
        if header in {
            "Код1с",
            "Номенклатура",
            "НаименованиДляПечати",
            "ТипПродукта",
            "ПродажнаяЦена",
            "СтопЦена",
            "ПлановаяЦена",
            "Остаток",
        }
    }
    if "Номенклатура" not in header_map or "Остаток" not in header_map:
        raise ValueError(f"Workbook stock file {stock_path.name} does not contain required headers.")

    for row in rows:
        if not row or not any(clean_text(value) for value in row):
            continue
        yield {
            header: row[index] if index < len(row) else ""
            for header, index in header_map.items()
        }


def score_header_row(values: Sequence[object]) -> tuple[int, dict[str, int]]:
    normalized = [build_search_text(value) for value in values]
    mapping: dict[str, int] = {}
    score = 0
    for index, cell in enumerate(normalized, start=1):
        if not cell:
            continue
        for field_name, keywords in HEADER_KEYWORDS.items():
            if field_name in mapping:
                continue
            if any(keyword in cell for keyword in keywords):
                mapping[field_name] = index
                score += 3
    score += sum(1 for value in values if clean_text(value))
    return score, mapping


def infer_column_defaults(
    rows: Sequence[Sequence[object]],
    detected_mapping: dict[str, int],
    header_row_index: int,
) -> dict[str, int]:
    mapping = dict(detected_mapping)
    if not rows:
        return mapping

    max_columns = max((len(row) for row in rows), default=0)
    sample_rows = [
        list(row)
        for row in rows[header_row_index + 1 : header_row_index + 21]
        if any(clean_text(value) for value in row)
    ]
    if not sample_rows:
        sample_rows = [list(row) for row in rows if any(clean_text(value) for value in row)]

    stats: dict[int, dict[str, float]] = {}
    for column_index in range(1, max_columns + 1):
        values = [clean_text(row[column_index - 1]) for row in sample_rows if column_index <= len(row)]
        non_empty = [value for value in values if value]
        if not non_empty:
            stats[column_index] = {
                "non_empty": 0,
                "numeric": 0,
                "unit": 0,
                "position": 0,
                "text": 0,
                "code_like": 0,
            }
            continue
        numeric = sum(1 for value in non_empty if parse_quantity(value) > 0)
        unit = sum(1 for value in non_empty if UNIT_CELL_RE.fullmatch(value))
        position = sum(1 for value in non_empty if POSITION_CELL_RE.fullmatch(value))
        text = sum(1 for value in non_empty if re.search(r"[A-Za-zА-Яа-яЁёІіЇїЄєҐґ]", value))
        code_like = sum(1 for value in non_empty if extract_code_tokens(value))
        stats[column_index] = {
            "non_empty": float(len(non_empty)),
            "numeric": float(numeric),
            "unit": float(unit),
            "position": float(position),
            "text": float(text),
            "code_like": float(code_like),
        }

    used_columns = {index for index in mapping.values() if index > 0}

    if "name" not in mapping:
        name_candidates = [
            column_index
            for column_index, column_stats in stats.items()
            if column_stats["text"] > 0 and column_index not in used_columns
        ]
        if name_candidates:
            mapping["name"] = max(
                name_candidates,
                key=lambda column_index: (
                    stats[column_index]["text"],
                    stats[column_index]["non_empty"] - stats[column_index]["numeric"],
                    -column_index,
                ),
            )
            used_columns.add(mapping["name"])

    if "qty" not in mapping:
        qty_candidates = [
            column_index
            for column_index, column_stats in stats.items()
            if column_stats["numeric"] > 0 and column_index not in used_columns
        ]
        if qty_candidates:
            mapping["qty"] = max(
                qty_candidates,
                key=lambda column_index: (
                    stats[column_index]["numeric"],
                    column_index,
                ),
            )
            used_columns.add(mapping["qty"])

    if "unit" not in mapping:
        unit_candidates = [
            column_index
            for column_index, column_stats in stats.items()
            if column_stats["unit"] > 0 and column_index not in used_columns
        ]
        if unit_candidates:
            mapping["unit"] = max(
                unit_candidates,
                key=lambda column_index: (
                    stats[column_index]["unit"],
                    -abs(column_index - mapping.get("qty", column_index)),
                ),
            )
            used_columns.add(mapping["unit"])

    if "position" not in mapping:
        position_candidates = [
            column_index
            for column_index, column_stats in stats.items()
            if column_stats["position"] > 0
            and column_stats["text"] <= column_stats["position"]
            and column_index not in used_columns
        ]
        if position_candidates:
            mapping["position"] = min(position_candidates)

    if "mark" not in mapping:
        mark_candidates = [
            column_index
            for column_index, column_stats in stats.items()
            if column_stats["code_like"] > 0 and column_index not in used_columns
        ]
        if mark_candidates:
            mapping["mark"] = max(
                mark_candidates,
                key=lambda column_index: (
                    stats[column_index]["code_like"],
                    stats[column_index]["text"],
                    -column_index,
                ),
            )

    optional_defaults = {
        "position": 0,
        "mark": 0,
        "supplier_code": 0,
        "vendor": 0,
        "unit": 0,
        "qty": 0,
    }
    for key, value in optional_defaults.items():
        mapping.setdefault(key, value)
    mapping.setdefault("name", 1 if max_columns else 0)
    return mapping


def detect_header(ws) -> tuple[int, dict[str, int], list[str]]:
    best_row = 1
    best_mapping: dict[str, int] = {}
    best_score = -1
    best_headers: list[str] = []
    for row_index in range(1, min(ws.max_row, 12) + 1):
        values = [ws.cell(row=row_index, column=col).value for col in range(1, ws.max_column + 1)]
        score, mapping = score_header_row(values)
        if score > best_score:
            best_score = score
            best_row = row_index
            best_mapping = mapping
            best_headers = [clean_text(value) for value in values]
    rows = [
        [ws.cell(row=row_index, column=col).value for col in range(1, ws.max_column + 1)]
        for row_index in range(1, min(ws.max_row, 50) + 1)
    ]
    best_mapping = infer_column_defaults(rows, best_mapping, best_row - 1)
    return best_row, best_mapping, best_headers


def row_has_content(values: Sequence[object]) -> bool:
    return any(clean_text(value) for value in values)


def load_order_lines(order_path: Path) -> list[OrderLine]:
    wb = load_workbook(order_path, data_only=True)
    all_lines: list[OrderLine] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, columns, headers = detect_header(ws)
        for row_index in range(header_row + 1, ws.max_row + 1):
            row_values = [ws.cell(row=row_index, column=col).value for col in range(1, ws.max_column + 1)]
            if not row_has_content(row_values):
                continue
            name = get_mapped_text(row_values, columns["name"])
            qty = get_mapped_quantity(row_values, columns["qty"])
            position = get_mapped_text(row_values, columns["position"])
            mark = get_mapped_text(row_values, columns["mark"])
            supplier_code = get_mapped_text(row_values, columns["supplier_code"])
            name, mark, supplier_code = coalesce_material_fields(name, mark, supplier_code)
            if not name or qty <= 0:
                continue
            vendor = get_mapped_text(row_values, columns["vendor"])
            unit = get_mapped_text(row_values, columns["unit"])
            query_parts = [name, mark, supplier_code, vendor]
            raw_query = " | ".join(part for part in query_parts if part)
            classification = NOMENCLATURE_CLASSIFIER.classify(raw_query or name)
            classifier_hint = " ".join(classification.normalized.canonical_tokens)
            search_text = build_search_text(*query_parts, classifier_hint)
            classifier_family_tags = {f"family:{tag}" for tag in classification.family_tags}
            dimension_tags = (
                extract_dimension_tags(*query_parts)
                | extract_family_tags(*query_parts, classifier_hint)
                | extract_parser_hint_tags(*query_parts, classifier_hint)
                | extract_material_tags_from_search_text(search_text)
                | extract_structured_tags_from_search_text(search_text)
                | classifier_family_tags
            )
            search_text = augment_search_text_with_dimension_tags(search_text, dimension_tags)
            search_tokens = extract_tokens(search_text)
            all_lines.append(
                OrderLine(
                    source_file=order_path.name,
                    sheet_name=sheet_name,
                    source_row=row_index,
                    headers=headers,
                    row_values=row_values,
                    position=position,
                    name=name,
                    mark=mark,
                    supplier_code=supplier_code,
                    vendor=vendor,
                    unit=unit,
                    requested_qty=qty,
                    search_text=search_text,
                    search_tokens=search_tokens,
                    key_tokens=extract_key_tokens(search_tokens),
                    root_tokens=extract_root_tokens(search_tokens),
                    code_tokens=extract_code_tokens(*query_parts),
                    dimension_tags=dimension_tags,
                    raw_query=raw_query,
                    classification=classification,
                )
            )
    return all_lines


class StockMatcher:
    def __init__(
        self,
        stock_items: Sequence[StockItem],
        reviewed_analog_decisions: Mapping[str, Mapping[str, str]] | None = None,
        substitution_policy: Mapping[str, object] | None = None,
    ):
        self.stock_items = list(stock_items)
        self.reviewed_analog_decisions = {
            query_key: {normalize_candidate_code(code): str(decision) for code, decision in decisions.items()}
            for query_key, decisions in (reviewed_analog_decisions or {}).items()
        }
        self.substitution_policy = dict(substitution_policy or {})
        self.search_token_lists: list[tuple[str, ...]] = []
        self.search_token_positions: list[dict[str, tuple[int, ...]]] = []
        search_token_index: defaultdict[str, list[int]] = defaultdict(list)
        token_index: defaultdict[str, list[int]] = defaultdict(list)
        root_index: defaultdict[str, list[int]] = defaultdict(list)
        code_index: defaultdict[str, list[int]] = defaultdict(list)
        dimension_index: defaultdict[str, list[int]] = defaultdict(list)
        dimension_group_index: defaultdict[str, list[int]] = defaultdict(list)
        self._candidate_cache: dict[tuple[tuple[str, ...], tuple[str, ...], tuple[str, ...]], list[int]] = {}
        for index, item in enumerate(self.stock_items):
            token_list = tuple(item.search_text.split())
            token_positions: defaultdict[str, list[int]] = defaultdict(list)
            for position, token in enumerate(token_list):
                token_positions[token].append(position)
            self.search_token_lists.append(token_list)
            self.search_token_positions.append({token: tuple(values) for token, values in token_positions.items()})
            for token in item.search_tokens:
                search_token_index[token].append(index)
            for token in item.key_tokens:
                token_index[token].append(index)
            for token in item.root_tokens:
                root_index[token].append(index)
            for token in item.code_tokens:
                code_index[token].append(index)

            expanded_dimensions = expand_dimension_values(group_dimension_tags(item.dimension_tags))
            for key, values in expanded_dimensions.items():
                if not values:
                    continue
                dimension_group_index[key].append(index)
                for value in values:
                    dimension_index[f"{key}:{value}"].append(index)

        self.search_token_index = dict(search_token_index)
        self.token_index = dict(token_index)
        self.root_index = dict(root_index)
        self.code_index = dict(code_index)
        self.dimension_index = dict(dimension_index)
        self.dimension_group_index = dict(dimension_group_index)
        self.search_token_counts = np.array([len(item.search_tokens) for item in self.stock_items], dtype=np.float64)

    def fork(self) -> "StockMatcher":
        forked = object.__new__(StockMatcher)
        forked.stock_items = clone_stock_items(self.stock_items)
        forked.reviewed_analog_decisions = self.reviewed_analog_decisions
        forked.substitution_policy = self.substitution_policy
        forked.search_token_lists = self.search_token_lists
        forked.search_token_positions = self.search_token_positions
        forked.search_token_index = self.search_token_index
        forked.token_index = self.token_index
        forked.root_index = self.root_index
        forked.code_index = self.code_index
        forked.dimension_index = self.dimension_index
        forked.dimension_group_index = self.dimension_group_index
        forked.search_token_counts = self.search_token_counts
        forked._candidate_cache = {}
        return forked

    def _review_decision_for_candidate(self, order: OrderLine, candidate_code: str) -> str:
        normalized_code = normalize_candidate_code(candidate_code)
        for query_key in build_review_query_keys(order.raw_query or order.name, order.name):
            decision = self.reviewed_analog_decisions.get(query_key, {}).get(normalized_code)
            if decision:
                return decision
        return ""

    def _sort_candidates(self, candidates: Sequence[Candidate]) -> list[Candidate]:
        ranked = list(candidates)
        ranked.sort(
            key=lambda candidate: (
                candidate.score,
                candidate.review_decision == REVIEW_DECISION_APPROVED,
                candidate.code_hit,
                candidate.stock.remaining > 0,
                candidate.stock.remaining,
                -candidate.stock.row_index,
            ),
            reverse=True,
        )
        return ranked

    def _apply_reviewed_candidate_decisions(self, order: OrderLine, candidates: Sequence[Candidate]) -> list[Candidate]:
        if not self.reviewed_analog_decisions:
            return list(candidates)

        adjusted: list[Candidate] = []
        for candidate in candidates:
            decision = self._review_decision_for_candidate(order, candidate.stock.code_1c)
            if decision == REVIEW_DECISION_REJECTED:
                continue
            if decision == REVIEW_DECISION_APPROVED:
                reasons = list(candidate.reasons)
                if "подтверждено ручной разметкой" not in reasons:
                    reasons.append("подтверждено ручной разметкой")
                candidate = replace(candidate, reasons=reasons, review_decision=decision)
            adjusted.append(candidate)
        return adjusted

    def _policy_for_order(self, order: OrderLine, candidate: Candidate | None = None) -> tuple[str, Mapping[str, object]]:
        return get_substitution_family_policy(
            order=order,
            candidate=candidate,
            substitution_policy=self.substitution_policy,
        )

    def _matches_policy_requirements(self, order: OrderLine, candidate: Candidate, required_keys: Sequence[str]) -> bool:
        return candidate_matches_policy_dimensions(order, candidate, required_keys)

    def _hits_policy_forbidden_tokens(self, order: OrderLine, candidate: Candidate, policy: Mapping[str, object]) -> bool:
        forbidden_prefixes = policy.get("forbid_candidate_token_prefixes", [])
        if not isinstance(forbidden_prefixes, list):
            return False
        return candidate_hits_forbidden_policy_tokens(order, candidate, [str(prefix) for prefix in forbidden_prefixes])

    def _hits_policy_exact_only_forbidden_tokens(
        self,
        order: OrderLine,
        candidate: Candidate,
        policy: Mapping[str, object],
    ) -> bool:
        forbidden_prefixes = policy.get("forbid_exact_candidate_token_prefixes", [])
        if not isinstance(forbidden_prefixes, list):
            return False
        return candidate_hits_forbidden_policy_tokens(order, candidate, [str(prefix) for prefix in forbidden_prefixes])

    def _exact_reason_by_policy(self, order: OrderLine, candidate: Candidate) -> str:
        family, policy = self._policy_for_order(order, candidate)
        if not policy or not bool(policy.get("allow_exact")):
            return ""
        required_keys = [str(key) for key in policy.get("exact_required", []) if str(key)]
        min_score = float(policy.get("exact_min_score", 0.0) or 0.0)
        if candidate.dimension_penalty != 0:
            return ""
        if not self._matches_policy_requirements(order, candidate, required_keys):
            return ""
        if self._hits_policy_forbidden_tokens(order, candidate, policy):
            return ""
        if self._hits_policy_exact_only_forbidden_tokens(order, candidate, policy):
            return ""
        if candidate.score < min_score:
            order_materials = extract_tag_values(order.dimension_tags, "mclass:")
            stock_materials = extract_tag_values(candidate.stock.dimension_tags, "mclass:")
            shared_materials = order_materials & stock_materials
            ppr_reducer_exact = (
                family == "perehod"
                and "polypropylene" in shared_materials
                and "pairdn" in candidate.matched_dimension_keys
            )
            if not ppr_reducer_exact:
                return ""
        return FAMILY_EXACT_REASONS.get(family, "совпали обязательные технические параметры")

    def _analog_reason_by_policy(self, order: OrderLine, candidate: Candidate) -> str:
        family, policy = self._policy_for_order(order, candidate)
        if not policy:
            return ""
        required_keys = [str(key) for key in policy.get("safe_required", []) if str(key)]
        if not required_keys:
            return ""
        analog_min_score = float(
            policy.get(
                "analog_min_score",
                max(40.0, float(policy.get("safe_min_score", 0.0) or 0.0) - 12.0),
            )
            or 0.0
        )
        if candidate.score < analog_min_score:
            return ""
        if not self._matches_policy_requirements(order, candidate, required_keys):
            return ""
        if self._hits_policy_forbidden_tokens(order, candidate, policy):
            return ""
        return FAMILY_ANALOG_REASONS.get(family, "совпали обязательные параметры семейства, нужна ручная проверка")

    def _order_signature(self, order: OrderLine) -> tuple[tuple[str, ...], tuple[str, ...], tuple[str, ...]]:
        return (
            tuple(sorted(order.key_tokens)),
            tuple(sorted(order.code_tokens)),
            tuple(sorted(order.dimension_tags)),
        )

    def _has_compatible_families(self, order: OrderLine, stock: StockItem) -> bool:
        order_families = extract_tag_values(order.dimension_tags, "family:")
        if not order_families:
            return True
        stock_families = extract_tag_values(stock.dimension_tags, "family:")
        if not stock_families:
            return False
        order_primary = order_families & PRIMARY_FAMILY_TAGS
        stock_primary = stock_families & PRIMARY_FAMILY_TAGS
        if order_primary:
            if not stock_primary or not order_primary.issubset(stock_primary):
                return False
            order_modifiers = order_families - PRIMARY_FAMILY_TAGS
            return order_modifiers.issubset(stock_families)
        return order_families.issubset(stock_families)

    def _has_compatible_materials(self, order: OrderLine, stock: StockItem) -> bool:
        order_materials = extract_tag_values(order.dimension_tags, "mclass:")
        if not order_materials:
            return True
        stock_materials = extract_tag_values(stock.dimension_tags, "mclass:")
        if not stock_materials:
            return False
        if not (order_materials & stock_materials):
            return False

        order_grades = extract_tag_values(order.dimension_tags, "grade:")
        if not order_grades:
            return True
        stock_grades = extract_tag_values(stock.dimension_tags, "grade:")
        if not stock_grades:
            return "steel" in stock_materials and order_grades == {"cs20"}
        return bool(order_grades & stock_grades)

    def _has_required_dimension_matches(self, order: OrderLine, stock: StockItem) -> bool:
        order_dimensions = expand_dimension_values(group_dimension_tags(order.dimension_tags))
        stock_dimensions = expand_dimension_values(group_dimension_tags(stock.dimension_tags))

        for required_key in ("tripledn", "pairdn"):
            order_values = order_dimensions.get(required_key, set())
            if not order_values:
                continue
            stock_values = stock_dimensions.get(required_key, set())
            if not stock_values or not (order_values & stock_values):
                return False
        return True

    def is_candidate_compatible(self, order: OrderLine, stock: StockItem) -> bool:
        return (
            self._has_compatible_families(order, stock)
            and self._has_compatible_materials(order, stock)
            and self._has_required_dimension_matches(order, stock)
        )

    def filter_analog_candidates(self, candidates: Sequence[Candidate]) -> list[Candidate]:
        filtered: list[Candidate] = []
        for candidate in candidates:
            approved = candidate.review_decision == REVIEW_DECISION_APPROVED
            if candidate.score < 50.0 and not approved:
                continue
            if candidate.dimension_penalty > 0 and not approved:
                continue
            if (
                candidate.overlap < 0.25
                and candidate.soft_overlap < 0.35
                and candidate.dimension_bonus < 10.0
                and not candidate.code_hit
                and not approved
            ):
                continue
            filtered.append(candidate)
        return self._sort_candidates(filtered)

    def generate_candidates(self, order: OrderLine) -> list[int]:
        signature = self._order_signature(order)
        if signature in self._candidate_cache:
            return self._candidate_cache[signature]

        candidate_ids: set[int] = set()
        for code in order.code_tokens:
            candidate_ids.update(self.code_index.get(code, []))
        informative_tokens = sorted(
            order.key_tokens,
            key=lambda token: (len(self.token_index.get(token, [])) or 100000, -len(token), token),
        )
        for token in informative_tokens[:8]:
            matches = self.token_index.get(token, [])
            if len(matches) == 0:
                continue
            candidate_ids.update(matches)
            if len(candidate_ids) >= 350:
                break
        if len(candidate_ids) < 40:
            informative_roots = sorted(
                order.root_tokens,
                key=lambda token: (len(self.root_index.get(token, [])) or 100000, -len(token), token),
            )
            for token in informative_roots[:6]:
                matches = self.root_index.get(token, [])
                if len(matches) == 0:
                    continue
                candidate_ids.update(matches)
                if len(candidate_ids) >= 450:
                    break
        if not candidate_ids:
            for tag in order.dimension_tags:
                candidate_ids.update(self.dimension_index.get(tag, []))
        candidates = list(candidate_ids)
        self._candidate_cache[signature] = candidates
        return candidates

    def rank_candidates(
        self,
        order: OrderLine,
        candidate_ids: Sequence[int] | None = None,
        limit: int | None = 5,
    ) -> list[Candidate]:
        if candidate_ids is None:
            candidate_ids = self.generate_candidates(order)
        scored = [
            self.score_candidate(order, self.stock_items[index])
            for index in candidate_ids
            if self.is_candidate_compatible(order, self.stock_items[index])
        ]
        scored = self._sort_candidates(self._apply_reviewed_candidate_decisions(order, scored))
        if limit is None:
            return scored
        return scored[:limit]

    def score_candidate(self, order: OrderLine, stock: StockItem) -> Candidate:
        query_tokens = order.search_text.split()
        stock_tokens = stock.search_text.split()
        set_ratio = token_set_ratio(query_tokens, stock_tokens)
        ordered_ratio = ordered_subsequence_ratio(order.search_text, stock.search_text)
        overlap = safe_overlap(order.key_tokens, stock.key_tokens)
        soft_overlap = safe_overlap(order.root_tokens, stock.root_tokens)
        score = 0.45 * set_ratio + 0.15 * ordered_ratio + 20.0 * overlap + 18.0 * soft_overlap
        reasons: list[str] = []

        code_hit = False
        if order.code_tokens and stock.code_tokens:
            shared_codes = order.code_tokens & stock.code_tokens
            if shared_codes:
                # Require at least one shared code that is 4+ chars or contains a dot/slash
                # (filters out trivially short matches like "l63")
                meaningful = any(
                    len(c) >= 4 or "." in c or "/" in c
                    for c in shared_codes
                )
                if meaningful:
                    code_hit = True
                    score += 18.0
                    reasons.append("совпадает код/марка")

        order_dimensions = expand_dimension_values(group_dimension_tags(order.dimension_tags))
        stock_dimensions = expand_dimension_values(group_dimension_tags(stock.dimension_tags))
        dimension_bonus = 0.0
        dimension_penalty = 0.0
        matched_dimension_keys: list[str] = []
        conflicting_dimension_keys: list[str] = []
        for key, order_values in iter_scored_dimensions(order_dimensions):
            stock_values = stock_dimensions.get(key)
            if not stock_values:
                continue
            bonus, penalty = DIMENSION_WEIGHTS.get(key, (6.0, 12.0))
            if order_values & stock_values:
                dimension_bonus += bonus
                matched_dimension_keys.append(key)
            else:
                dimension_penalty += penalty
                conflicting_dimension_keys.append(key)
        score += dimension_bonus
        score -= dimension_penalty
        if matched_dimension_keys:
            reasons.append("совпадают размеры")
        if conflicting_dimension_keys:
            reasons.append("конфликт по размерам")

        flange_bonus, flange_reasons = flange_preference_bonus(order, stock)
        score += flange_bonus
        if flange_reasons:
            reasons.extend(flange_reasons)

        missing_major_tokens = [token for token in order.key_tokens if len(token) >= 5 and token not in stock.key_tokens]
        if order.key_tokens:
            penalty = 12.0 * len(missing_major_tokens) / len(order.key_tokens)
            score -= penalty
            if penalty >= 6:
                reasons.append("не хватает ключевых слов")

        if stock.remaining <= 0:
            reasons.append("остаток уже исчерпан")

        return Candidate(
            stock=stock,
            score=max(0.0, min(100.0, score)),
            overlap=overlap,
            reasons=reasons,
            code_hit=code_hit,
            dimension_bonus=dimension_bonus,
            dimension_penalty=dimension_penalty,
            soft_overlap=soft_overlap,
            matched_dimension_keys=tuple(dict.fromkeys(matched_dimension_keys)),
            conflicting_dimension_keys=tuple(dict.fromkeys(conflicting_dimension_keys)),
        )

    def find_candidates(self, order: OrderLine, limit: int = 5) -> list[Candidate]:
        return self.rank_candidates(order, self.generate_candidates(order), limit)

    def find_candidates_exhaustive_reference(self, order: OrderLine, limit: int | None = 5) -> list[Candidate]:
        return self.rank_candidates(order, range(len(self.stock_items)), limit)

    def _accumulate_token_hits(self, tokens: Iterable[str], index_map: dict[str, list[int]]) -> np.ndarray:
        counts = np.zeros(len(self.stock_items), dtype=np.float64)
        for token in tokens:
            matches = index_map.get(token, [])
            if matches:
                counts[matches] += 1.0
        return counts

    def _build_exhaustive_arrays(
        self,
        order: OrderLine,
    ) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
        stock_count = len(self.stock_items)

        search_hits = np.zeros(stock_count, dtype=np.float64)
        set_ratio = np.zeros(stock_count, dtype=np.float64)
        if order.search_tokens:
            search_hits = self._accumulate_token_hits(order.search_tokens, self.search_token_index)
            mask = search_hits > 0
            if mask.any():
                denominator = float(len(order.search_tokens)) + self.search_token_counts[mask]
                set_ratio[mask] = 200.0 * search_hits[mask] / denominator

        overlap = np.zeros(stock_count, dtype=np.float64)
        missing_penalty = np.zeros(stock_count, dtype=np.float64)
        if order.key_tokens:
            key_hits = self._accumulate_token_hits(order.key_tokens, self.token_index)
            overlap = key_hits / float(len(order.key_tokens))
            major_tokens = [token for token in order.key_tokens if len(token) >= 5]
            if major_tokens:
                major_hits = self._accumulate_token_hits(major_tokens, self.token_index)
                missing_penalty = 12.0 * (len(major_tokens) - major_hits) / float(len(order.key_tokens))

        soft_overlap = np.zeros(stock_count, dtype=np.float64)
        if order.root_tokens:
            root_hits = self._accumulate_token_hits(order.root_tokens, self.root_index)
            soft_overlap = root_hits / float(len(order.root_tokens))

        # Only count meaningful code tokens (len >= 4 or contains "." or "/")
        # to match the same filter applied in score_candidate().
        meaningful_order_codes = {
            c for c in order.code_tokens
            if len(c) >= 4 or "." in c or "/" in c
        }
        code_hits = np.zeros(stock_count, dtype=bool)
        for token in meaningful_order_codes:
            matches = self.code_index.get(token, [])
            if matches:
                code_hits[matches] = True

        dimension_bonus = np.zeros(stock_count, dtype=np.float64)
        dimension_penalty = np.zeros(stock_count, dtype=np.float64)
        order_dimensions = expand_dimension_values(group_dimension_tags(order.dimension_tags))
        for key, order_values in iter_scored_dimensions(order_dimensions):
            group_matches = self.dimension_group_index.get(key, [])
            if not group_matches:
                continue
            matched_ids: set[int] = set()
            for value in order_values:
                matched_ids.update(self.dimension_index.get(f"{key}:{value}", []))
            bonus, penalty = DIMENSION_WEIGHTS.get(key, (6.0, 12.0))
            dimension_penalty[group_matches] += penalty
            if matched_ids:
                matched_list = list(matched_ids)
                dimension_bonus[matched_list] += bonus
                dimension_penalty[matched_list] -= penalty

        base_scores = 0.45 * set_ratio + 20.0 * overlap + 18.0 * soft_overlap + dimension_bonus - dimension_penalty
        if "family:flange" in order.dimension_tags:
            flange_bonus = np.zeros(stock_count, dtype=np.float64)
            for stock_index, stock in enumerate(self.stock_items):
                bonus, _ = flange_preference_bonus(order, stock)
                flange_bonus[stock_index] = bonus
            base_scores += flange_bonus
        base_scores -= missing_penalty
        base_scores += code_hits.astype(np.float64) * 18.0
        return (
            base_scores,
            overlap,
            soft_overlap,
            code_hits,
            dimension_bonus,
            dimension_penalty,
            missing_penalty,
            search_hits,
        )

    def _candidate_sort_key_from_values(self, stock_index: int, score: float, code_hit: bool) -> tuple[float, bool, bool, float, int]:
        stock = self.stock_items[stock_index]
        return (
            score,
            code_hit,
            stock.remaining > 0,
            stock.remaining,
            -stock.row_index,
        )

    def _ordered_subsequence_ratio_cached(
        self,
        order_query: str,
        query_tokens: Sequence[str],
        stock_index: int,
    ) -> float:
        candidate_text = self.stock_items[stock_index].search_text
        if not order_query or not candidate_text:
            return 0.0
        if order_query in candidate_text or candidate_text in order_query:
            return 100.0
        if not query_tokens:
            return 0.0

        candidate_positions = self.search_token_positions[stock_index]
        matches = 0
        current_index = -1
        for token in query_tokens:
            positions = candidate_positions.get(token)
            if not positions:
                continue
            next_index = bisect_right(positions, current_index)
            if next_index >= len(positions):
                continue
            current_index = positions[next_index]
            matches += 1
        return 100.0 * matches / len(query_tokens)

    def _build_candidate_from_arrays(
        self,
        order: OrderLine,
        stock_index: int,
        score: float,
        overlap: np.ndarray,
        soft_overlap: np.ndarray,
        code_hits: np.ndarray,
        dimension_bonus: np.ndarray,
        dimension_penalty: np.ndarray,
        missing_penalty: np.ndarray,
    ) -> Candidate:
        stock = self.stock_items[stock_index]
        reasons: list[str] = []
        matched_dimension_keys: list[str] = []
        conflicting_dimension_keys: list[str] = []
        order_dimensions = expand_dimension_values(group_dimension_tags(order.dimension_tags))
        stock_dimensions = expand_dimension_values(group_dimension_tags(stock.dimension_tags))
        if code_hits[stock_index]:
            reasons.append("совпадает код/марка")
        for key, order_values in iter_scored_dimensions(order_dimensions):
            stock_values = stock_dimensions.get(key)
            if not stock_values:
                continue
            if order_values & stock_values:
                matched_dimension_keys.append(key)
            else:
                conflicting_dimension_keys.append(key)
        if dimension_bonus[stock_index] > 0:
            reasons.append("совпадают размеры")
        if dimension_penalty[stock_index] > 0:
            reasons.append("конфликт по размерам")
        flange_bonus, flange_reasons = flange_preference_bonus(order, stock)
        if flange_bonus and flange_reasons:
            reasons.extend(flange_reasons)
        if missing_penalty[stock_index] >= 6.0:
            reasons.append("не хватает ключевых слов")
        if stock.remaining <= 0:
            reasons.append("остаток уже исчерпан")
        return Candidate(
            stock=stock,
            score=score,
            overlap=float(overlap[stock_index]),
            soft_overlap=float(soft_overlap[stock_index]),
            reasons=reasons,
            code_hit=bool(code_hits[stock_index]),
            dimension_bonus=float(dimension_bonus[stock_index]),
            dimension_penalty=float(dimension_penalty[stock_index]),
            matched_dimension_keys=tuple(dict.fromkeys(matched_dimension_keys)),
            conflicting_dimension_keys=tuple(dict.fromkeys(conflicting_dimension_keys)),
        )

    def find_candidates_exhaustive(self, order: OrderLine, limit: int | None = 5) -> list[Candidate]:
        if limit is None:
            return self.find_candidates_exhaustive_reference(order, limit=None)

        base_scores, overlap, soft_overlap, code_hits, dimension_bonus, dimension_penalty, missing_penalty, search_hits = (
            self._build_exhaustive_arrays(order)
        )
        query_tokens = order.search_text.split()
        lower_bounds = np.clip(base_scores, 0.0, 100.0)
        if query_tokens:
            if len(query_tokens) == len(order.search_tokens):
                ordered_bonus_upper = 15.0 * search_hits / float(len(query_tokens))
            else:
                ordered_bonus_upper = np.where(search_hits > 0, 15.0, 0.0)
        else:
            ordered_bonus_upper = np.zeros(len(self.stock_items), dtype=np.float64)
        upper_bounds = np.clip(base_scores + ordered_bonus_upper, 0.0, 100.0)
        sorted_indices = np.argsort(upper_bounds)[::-1]

        target = max(limit, 1)
        if self.reviewed_analog_decisions:
            target += 12
        top_entries: list[tuple[float, bool, bool, float, int, int]] = []

        for raw_index in sorted_indices:
            stock_index = int(raw_index)
            if not self.is_candidate_compatible(order, self.stock_items[stock_index]):
                continue
            if len(top_entries) >= target and upper_bounds[stock_index] + 1e-9 < top_entries[-1][0]:
                break

            if upper_bounds[stock_index] <= lower_bounds[stock_index] + 1e-9:
                final_score = float(lower_bounds[stock_index])
            else:
                ordered_ratio = self._ordered_subsequence_ratio_cached(order.search_text, query_tokens, stock_index)
                final_score = float(np.clip(base_scores[stock_index] + 0.15 * ordered_ratio, 0.0, 100.0))

            entry = self._candidate_sort_key_from_values(
                stock_index,
                final_score,
                bool(code_hits[stock_index]),
            ) + (stock_index,)
            if len(top_entries) < target:
                top_entries.append(entry)
                top_entries.sort(reverse=True)
            elif entry > top_entries[-1]:
                top_entries[-1] = entry
                top_entries.sort(reverse=True)

        candidates = [
            self._build_candidate_from_arrays(
                order=order,
                stock_index=entry[-1],
                score=entry[0],
                overlap=overlap,
                soft_overlap=soft_overlap,
                code_hits=code_hits,
                dimension_bonus=dimension_bonus,
                dimension_penalty=dimension_penalty,
                missing_penalty=missing_penalty,
            )
            for entry in top_entries
        ]
        candidates = self._sort_candidates(self._apply_reviewed_candidate_decisions(order, candidates))
        return candidates[:limit]

    def classify(self, order: OrderLine, candidates: Sequence[Candidate]) -> tuple[str, Candidate | None, str]:
        if not candidates:
            return "not_found", None, "ничего не нашлось"
        best = candidates[0]
        second = candidates[1] if len(candidates) > 1 else None
        gap = best.score - second.score if second else best.score
        _, policy = self._policy_for_order(order, best)
        generic_exact_allowed = not policy or bool(policy.get("allow_exact"))

        if best.code_hit and best.dimension_penalty == 0 and best.score >= 78 and best.overlap >= 0.30:
            return "exact", best, "совпал код/марка"
        for candidate in candidates:
            policy_exact_reason = self._exact_reason_by_policy(order, candidate)
            if policy_exact_reason:
                return "exact", candidate, policy_exact_reason
        if is_exact_pipe_candidate(order, best) and best.score >= 84:
            return "exact", best, "совпали тип трубы, размер и ГОСТ"
        if generic_exact_allowed and best.score >= 90 and best.overlap >= 0.55 and best.dimension_penalty == 0 and gap >= 4:
            return "exact", best, "высокая уверенность"
        if generic_exact_allowed and "tripledn" in best.matched_dimension_keys and best.dimension_penalty == 0 and best.score >= 60 and gap >= 5:
            return "exact", best, "совпали тип изделия и тройные размеры"
        if generic_exact_allowed and best.score >= 82 and best.overlap >= 0.65 and best.dimension_bonus >= 10 and gap >= 5:
            return "exact", best, "совпали ключевые слова и размеры"
        policy_analog_reason = self._analog_reason_by_policy(order, best)
        if policy_analog_reason:
            return "analog", best, policy_analog_reason
        if best.score >= 72:
            return "analog", best, "нужна ручная проверка"
        if best.dimension_penalty == 0 and best.dimension_bonus >= 30 and best.score >= 58:
            return "analog", best, "сильное совпадение по исполнению и размерам, нужна ручная проверка"
        if best.dimension_penalty == 0 and best.dimension_bonus >= 18 and best.score >= 50 and best.overlap >= 0.45:
            return "analog", best, "совпадают ключевые слова и размеры, нужна ручная проверка"
        if best.dimension_bonus >= 18 and best.score >= 52 and (best.overlap >= 0.35 or best.soft_overlap >= 0.45):
            return "analog", best, "совпадают тип изделия и размеры, нужна ручная проверка"
        if best.dimension_bonus >= 10 and best.score >= 58 and (best.overlap >= 0.20 or best.soft_overlap >= 0.35):
            return "analog", best, "совпадают размеры, нужна ручная проверка"
        if best.code_hit and best.score >= 58:
            return "analog", best, "совпал код/марка, нужна ручная проверка"
        return "not_found", best, "подходящего совпадения нет"


def determine_analog_status(
    candidates: Sequence[Candidate],
    order: OrderLine | None = None,
    substitution_policy: Mapping[str, object] | None = None,
) -> str:
    if not candidates:
        return STATUS_APPROVAL_ANALOG
    best = candidates[0]
    second = candidates[1] if len(candidates) > 1 else None
    gap = best.score - second.score if second else best.score
    _, policy = get_substitution_family_policy(order=order, candidate=best, substitution_policy=substitution_policy)
    safe_required = [str(key) for key in policy.get("safe_required", []) if str(key)] if policy else []
    approval_only = bool(policy.get("approval_only")) if policy else False
    safe_min_score = float(policy.get("safe_min_score", 0.0) or 0.0) if policy else 0.0
    policy_forbidden = (
        candidate_hits_forbidden_policy_tokens(order, best, [str(prefix) for prefix in policy.get("forbid_candidate_token_prefixes", [])])
        if order is not None and isinstance(policy.get("forbid_candidate_token_prefixes", []), list)
        else False
    )
    reviewed_safe = (
        best.review_decision == REVIEW_DECISION_APPROVED
        and best.dimension_penalty == 0
        and best.dimension_bonus >= 10.0
    )
    if reviewed_safe and order is not None and safe_required:
        reviewed_safe = candidate_matches_policy_dimensions(order, best, safe_required)
    rule_safe = (
        best.dimension_penalty == 0
        and best.dimension_bonus >= 18.0
        and best.score >= 58.0
        and gap >= 3.0
        and (best.overlap >= 0.35 or best.soft_overlap >= 0.45 or best.code_hit)
    )
    high_confidence_safe = (
        best.dimension_penalty == 0
        and best.score >= 72.0
        and gap >= 4.0
        and best.overlap >= 0.5
    )
    policy_safe = (
        bool(policy)
        and not approval_only
        and order is not None
        and safe_required
        and best.dimension_penalty == 0
        and best.score >= safe_min_score
        and gap >= 3.0
        and candidate_matches_policy_dimensions(order, best, safe_required)
        and not policy_forbidden
        and (best.overlap >= 0.25 or best.soft_overlap >= 0.35 or best.code_hit or best.dimension_bonus >= 18.0)
    )
    if approval_only:
        return STATUS_APPROVAL_ANALOG
    if reviewed_safe or policy_safe or rule_safe or high_confidence_safe:
        return STATUS_SAFE_ANALOG
    return STATUS_APPROVAL_ANALOG


def match_orders(
    lines: Sequence[OrderLine],
    matcher: StockMatcher,
    use_full_scan_fallback: bool = True,
) -> list[MatchResult]:
    results: list[MatchResult] = []
    for line in lines:
        candidates = matcher.find_candidates(line)
        match_kind, best, reason = matcher.classify(line, candidates)
        # Use full NumPy scan as fallback only when token-based search is
        # inconclusive: no candidates, low confidence, or nothing found at all.
        # Skip when token-based already returned a confident analog (score ≥ 68)
        # to avoid the per-item compatibility check overhead on every row.
        needs_exhaustive = use_full_scan_fallback and match_kind != "exact" and (
            not candidates
            or match_kind == "not_found"
            or best is None
            or best.score < 68.0
        )
        if needs_exhaustive:
            exhaustive_candidates = matcher.find_candidates_exhaustive(line, limit=5)
            exhaustive_kind, exhaustive_best, exhaustive_reason = matcher.classify(line, exhaustive_candidates)
            candidates = exhaustive_candidates
            match_kind = exhaustive_kind
            best = exhaustive_best
            reason = exhaustive_reason
        if match_kind == "exact" and best is not None:
            available_qty = min(line.requested_qty, best.stock.remaining)
            if available_qty > 0:
                best.stock.remaining -= available_qty
                analogs = matcher.filter_analog_candidates(candidates[1:4])
                if available_qty >= line.requested_qty:
                    status = STATUS_FOUND_FULL
                    comment = reason
                else:
                    status = STATUS_FOUND_PARTIAL
                    comment = f"{reason}; на остатке меньше, чем запрошено"
                results.append(
                    MatchResult(
                        order=line,
                        status=status,
                        matched_stock=best.stock,
                        available_qty=available_qty,
                        confidence=best.score,
                        comment=comment,
                        analogs=analogs,
                    )
                )
                continue
            results.append(
                MatchResult(
                    order=line,
                    status=STATUS_STOCK_DEPLETED,
                    matched_stock=best.stock,
                    available_qty=0.0,
                    confidence=best.score,
                    comment="позиция нашлась, но её остаток уже ушёл на предыдущие строки",
                    analogs=matcher.filter_analog_candidates(candidates[1:4]),
                )
            )
            continue
        if match_kind == "analog" and best is not None:
            analog_status = determine_analog_status(
                candidates,
                order=line,
                substitution_policy=matcher.substitution_policy,
            )
            analog_candidates = matcher.filter_analog_candidates(candidates[:4])
            if not analog_candidates:
                analog_candidates = [best]
            results.append(
                MatchResult(
                    order=line,
                    status=analog_status,
                    matched_stock=None,
                    available_qty=0.0,
                    confidence=best.score,
                    comment=reason,
                    analogs=analog_candidates,
                )
            )
            continue
        filtered_analogs = matcher.filter_analog_candidates(candidates[:4])
        manual_review_rule = find_manual_review_rule(line)
        if manual_review_rule and not filtered_analogs:
            reason = str(
                manual_review_rule.get(
                    "comment",
                    "требует ручного подбора: в запросе не хватает параметров для безопасного аналога",
                )
            )
        elif line.classification and line.classification.status == ClassificationStatus.AMBIGUOUS and not filtered_analogs:
            alternative_labels = [
                candidate.category.label
                for candidate in line.classification.alternatives[1:3]
                if candidate.category.label
            ]
            if alternative_labels:
                reason = f"нужна ручная классификация: рядом {', '.join(alternative_labels)}"
            else:
                reason = "нужна ручная классификация: позиция слишком неоднозначна"
        elif line.classification and line.classification.status == ClassificationStatus.UNCLASSIFIED and not filtered_analogs:
            reason = "классификатор не распознал категорию: нужна ручная проверка строки"
        results.append(
            MatchResult(
                order=line,
                status=STATUS_NOT_FOUND,
                matched_stock=None,
                available_qty=0.0,
                confidence=best.score if best else 0.0,
                comment=reason,
                analogs=filtered_analogs,
            )
        )
    return results


def clone_stock_items(stock_items: Sequence[StockItem]) -> list[StockItem]:
    return [replace(item, remaining=item.quantity) for item in stock_items]


def ensure_output_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def autosize_columns(ws) -> None:
    for column in ws.columns:
        letter = column[0].column_letter
        max_len = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 40)


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def add_row(ws, values: Sequence[object]) -> None:
    ws.append(list(values))
    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def format_candidate_text(candidate: Candidate, include_score: bool = True) -> str:
    base = (
        f"{candidate.stock.code_1c} | {candidate.stock.name} | "
        f"остаток {format_number(candidate.stock.remaining)}"
    )
    if include_score:
        base = f"{base} | score {candidate.score:.1f}"
    if candidate.review_decision == REVIEW_DECISION_APPROVED:
        return f"{base} | подтверждено"
    return base


def manager_action_for_result(result: MatchResult) -> str:
    if result.status in {STATUS_FOUND_FULL, STATUS_FOUND_PARTIAL}:
        return "Загрузить в 1С"
    if result.status == STATUS_SAFE_ANALOG:
        return "Подтвердить аналог"
    if result.status == STATUS_APPROVAL_ANALOG:
        return "Согласовать замену"
    if "ручного подбора" in result.comment.lower():
        return "Ручной подбор"
    if result.status == STATUS_STOCK_DEPLETED:
        return "Проверить остаток"
    return "Нет в остатке"


def build_summary_sheet(wb: Workbook, results: Sequence[MatchResult], stock_items: Sequence[StockItem]) -> None:
    ws = wb.create_sheet("Сводка")
    counters: defaultdict[str, int] = defaultdict(int)
    for result in results:
        counters[result.status] += 1
    total_requested = sum(result.order.requested_qty for result in results)
    total_allocated = sum(result.available_qty for result in results)
    rows = [
        ("Показатель", "Значение"),
        ("Всего строк заявки", len(results)),
        (STATUS_FOUND_FULL, counters[STATUS_FOUND_FULL]),
        (STATUS_FOUND_PARTIAL, counters[STATUS_FOUND_PARTIAL]),
        (STATUS_SAFE_ANALOG, counters[STATUS_SAFE_ANALOG]),
        (STATUS_APPROVAL_ANALOG, counters[STATUS_APPROVAL_ANALOG]),
        (STATUS_NOT_FOUND, counters[STATUS_NOT_FOUND]),
        (STATUS_STOCK_DEPLETED, counters[STATUS_STOCK_DEPLETED]),
        ("Запрошено всего", format_number(total_requested)),
        ("Выделено из остатка", format_number(total_allocated)),
        ("Свободный остаток после распределения", format_number(sum(item.remaining for item in stock_items))),
    ]
    for row in rows:
        ws.append(list(row))
    style_header(ws)
    autosize_columns(ws)


def create_detailed_workbook(
    results: Sequence[MatchResult],
    stock_items: Sequence[StockItem],
    out_path: Path,
) -> None:
    wb = Workbook()
    manager_ws = wb.active
    manager_ws.title = "Менеджеру"
    manager_ws.append(
        [
            "Действие",
            "Заявка",
            "Кол-во",
            "Что берем / проверяем",
            "Остаток",
            "Комментарий",
        ]
    )
    detail_ws = wb.create_sheet("Результат")
    detail_ws.append(
        [
            "Действие менеджера",
            "Статус",
            "Позиция",
            "Наименование заявки",
            "Категория",
            "Статус классификатора",
            "Маршрут",
            "Класс. уверенность",
            "Тип/марка",
            "Производитель",
            "Ед.",
            "Запрошено",
            "Доступно в заявку",
            "Уверенность",
            "Код1с",
            "Номенклатура 1С",
            "Цена",
            "Комментарий",
            "Аналог 1",
            "Аналог 2",
            "Аналог 3",
            "Файл",
            "Лист",
            "Строка",
        ]
    )

    sorted_results = sorted(
        results,
        key=lambda result: (
            result.order.source_file,
            result.order.sheet_name,
            result.order.source_row,
        ),
    )

    for result in sorted_results:
        matched = result.matched_stock
        manager_target = ""
        manager_stock = ""
        if matched is not None:
            manager_target = f"{matched.code_1c} | {matched.name}"
            manager_stock = format_number(result.available_qty)
        elif result.analogs:
            best_analog = result.analogs[0]
            manager_target = f"{best_analog.stock.code_1c} | {best_analog.stock.name}"
            manager_stock = format_number(best_analog.stock.remaining)
        add_row(
            manager_ws,
            [
                manager_action_for_result(result),
                result.order.name,
                format_number(result.order.requested_qty),
                manager_target,
                manager_stock,
                result.comment,
            ],
        )
        status_fill = STATUS_COLORS.get(result.status)
        if status_fill:
            manager_ws.cell(row=manager_ws.max_row, column=1).fill = PatternFill("solid", fgColor=status_fill)

        analog_names = []
        for analog in result.analogs[:3]:
            analog_names.append(format_candidate_text(analog, include_score=False))
        while len(analog_names) < 3:
            analog_names.append("")
        price = ""
        if matched is not None:
            price = parse_price(matched.sale_price) or parse_price(matched.stop_price) or parse_price(matched.plan_price)
        classification = result.order.classification
        row = [
            manager_action_for_result(result),
            result.status,
            result.order.position,
            result.order.name,
            classification.category_label if classification and classification.category_label else "",
            classification.status.value if classification else "",
            classification.route if classification else "",
            classification.confidence if classification else "",
            result.order.mark,
            result.order.vendor,
            result.order.unit,
            format_number(result.order.requested_qty),
            format_number(result.available_qty),
            round(result.confidence, 1),
            matched.code_1c if matched else "",
            matched.name if matched else "",
            price,
            result.comment,
            analog_names[0],
            analog_names[1],
            analog_names[2],
            result.order.source_file,
            result.order.sheet_name,
            result.order.source_row,
        ]
        detail_ws.append(row)
        status_cell = detail_ws.cell(row=detail_ws.max_row, column=2)
        fill_color = STATUS_COLORS.get(result.status)
        if fill_color:
            status_cell.fill = PatternFill("solid", fgColor=fill_color)
            detail_ws.cell(row=detail_ws.max_row, column=1).fill = PatternFill("solid", fgColor=fill_color)

    style_header(manager_ws)
    manager_ws.auto_filter.ref = manager_ws.dimensions
    manager_ws.column_dimensions["A"].width = 18
    manager_ws.column_dimensions["B"].width = 42
    manager_ws.column_dimensions["C"].width = 10
    manager_ws.column_dimensions["D"].width = 46
    manager_ws.column_dimensions["E"].width = 12
    manager_ws.column_dimensions["F"].width = 36
    for row in manager_ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    style_header(detail_ws)
    autosize_columns(detail_ws)
    detail_ws.auto_filter.ref = detail_ws.dimensions
    detail_ws.column_dimensions["D"].width = 44
    detail_ws.column_dimensions["E"].width = 24
    detail_ws.column_dimensions["F"].width = 18
    detail_ws.column_dimensions["G"].width = 18
    detail_ws.column_dimensions["P"].width = 44
    detail_ws.column_dimensions["R"].width = 50
    detail_ws.column_dimensions["S"].width = 44
    detail_ws.column_dimensions["T"].width = 44
    detail_ws.column_dimensions["U"].width = 44

    raw_ws = wb.create_sheet("Исходные строки")
    max_original_columns = max((len(result.order.row_values) for result in results), default=0)
    raw_headers = [f"Исходная колонка {index}" for index in range(1, max_original_columns + 1)]
    raw_ws.append(
        raw_headers
        + [
            "Файл",
            "Лист",
            "Строка",
            "Позиция",
            "Наименование",
            "Исходный запрос",
            "Категория",
            "Статус классификатора",
            "Маршрут",
            "Класс. уверенность",
            "Причина",
            "Объяснение",
        ]
    )
    for result in results:
        classification = result.order.classification
        row = list(result.order.row_values) + [None] * (max_original_columns - len(result.order.row_values))
        row.extend(
            [
                result.order.source_file,
                result.order.sheet_name,
                result.order.source_row,
                result.order.position,
                result.order.name,
                result.order.raw_query,
                classification.category_label if classification and classification.category_label else "",
                classification.status.value if classification else "",
                classification.route if classification else "",
                classification.confidence if classification else "",
                classification.reason if classification else "",
                " | ".join(classification.explanation) if classification else "",
            ]
        )
        raw_ws.append(row)
    style_header(raw_ws)
    autosize_columns(raw_ws)
    raw_ws.auto_filter.ref = raw_ws.dimensions

    build_summary_sheet(wb, results, stock_items)
    wb.save(out_path)


def create_for_1c_workbook(results: Sequence[MatchResult], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Штрихкод", "Код", "Артикул", "Номенклатура", "Количество", "Цена"])
    for result in results:
        if result.status not in {STATUS_FOUND_FULL, STATUS_FOUND_PARTIAL}:
            continue
        stock = result.matched_stock
        if stock is None or result.available_qty <= 0:
            continue
        price = parse_price(stock.sale_price) or parse_price(stock.stop_price) or parse_price(stock.plan_price)
        add_row(
            ws,
            [
                "",
                stock.code_1c,
                "",
                stock.name,
                format_number(result.available_qty),
                price,
            ],
        )
    style_header(ws)
    autosize_columns(ws)
    wb.save(out_path)


def create_analogs_workbook(results: Sequence[MatchResult], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Аналоги"
    ws.append(
        [
            "Файл",
            "Лист",
            "Строка",
            "Позиция",
            "Наименование заявки",
            "Статус",
            "Тип/марка",
            "Запрошено",
            "Комментарий",
            "Аналог 1",
            "Аналог 2",
            "Аналог 3",
            "Аналог 4",
        ]
    )
    for result in results:
        if result.status not in ANALOG_STATUSES:
            continue
        analogs = []
        for candidate in result.analogs[:4]:
            analogs.append(format_candidate_text(candidate, include_score=True))
        while len(analogs) < 4:
            analogs.append("")
        add_row(
            ws,
            [
                result.order.source_file,
                result.order.sheet_name,
                result.order.source_row,
                result.order.position,
                result.order.name,
                result.status,
                result.order.mark,
                format_number(result.order.requested_qty),
                result.comment,
                analogs[0],
                analogs[1],
                analogs[2],
                analogs[3],
            ],
        )
    style_header(ws)
    autosize_columns(ws)
    wb.save(out_path)


def create_not_found_workbook(results: Sequence[MatchResult], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Не найдено"
    ws.append(
        [
            "Файл",
            "Лист",
            "Строка",
            "Позиция",
            "Наименование заявки",
            "Тип/марка",
            "Поставщик/производитель",
            "Запрошено",
            "Статус",
            "Комментарий",
            "Лучший кандидат",
        ]
    )
    for result in results:
        if result.status not in {STATUS_NOT_FOUND, STATUS_STOCK_DEPLETED}:
            continue
        best_candidate = result.analogs[0] if result.analogs else None
        best_text = ""
        if best_candidate is not None:
            best_text = (
                f"{best_candidate.stock.code_1c} | {best_candidate.stock.name} | "
                f"остаток {format_number(best_candidate.stock.remaining)} | score {best_candidate.score:.1f}"
            )
        add_row(
            ws,
            [
                result.order.source_file,
                result.order.sheet_name,
                result.order.source_row,
                result.order.position,
                result.order.name,
                result.order.mark,
                result.order.vendor,
                format_number(result.order.requested_qty),
                result.status,
                result.comment,
                best_text,
            ],
        )
    style_header(ws)
    autosize_columns(ws)
    wb.save(out_path)


def build_status_counts(results: Sequence[MatchResult]) -> dict[str, int]:
    counters: defaultdict[str, int] = defaultdict(int)
    for result in results:
        counters[result.status] += 1
    return {
        STATUS_FOUND_FULL: counters[STATUS_FOUND_FULL],
        STATUS_FOUND_PARTIAL: counters[STATUS_FOUND_PARTIAL],
        STATUS_SAFE_ANALOG: counters[STATUS_SAFE_ANALOG],
        STATUS_APPROVAL_ANALOG: counters[STATUS_APPROVAL_ANALOG],
        STATUS_NOT_FOUND: counters[STATUS_NOT_FOUND],
        STATUS_STOCK_DEPLETED: counters[STATUS_STOCK_DEPLETED],
    }


def build_matching_metrics(results: Sequence[MatchResult], stock_items: Sequence[StockItem]) -> dict[str, object]:
    status_counts = build_status_counts(results)
    classifier_statuses: defaultdict[str, int] = defaultdict(int)
    for result in results:
        if result.order.classification:
            classifier_statuses[result.order.classification.status.value] += 1
    requested_total = sum(result.order.requested_qty for result in results)
    allocated_total = sum(result.available_qty for result in results)
    analog_results = [result for result in results if result.status in ANALOG_STATUSES]
    analog_confidences = [result.confidence for result in analog_results]
    top_reviewed = [
        result
        for result in analog_results
        if result.analogs and result.analogs[0].review_decision == REVIEW_DECISION_APPROVED
    ]
    return {
        "total_rows": len(results),
        "status_counts": status_counts,
        "classifier_status_counts": dict(classifier_statuses),
        "requested_qty_total": round(requested_total, 3),
        "allocated_qty_total": round(allocated_total, 3),
        "allocation_ratio": round(allocated_total / requested_total, 4) if requested_total else 0.0,
        "free_stock_total": round(sum(item.remaining for item in stock_items), 3),
        "stock_row_count": len(stock_items),
        "analog_rows_total": len(analog_results),
        "reviewed_top_analog_rows": len(top_reviewed),
        "avg_analog_confidence": round(sum(analog_confidences) / len(analog_confidences), 3) if analog_confidences else 0.0,
    }


def build_rules_metadata(reviewed_decisions_path: Path | None) -> dict[str, object]:
    classifier_payload = load_json_payload(CLASSIFIER_CONFIG_PATH)
    dictionary_payload = load_json_payload(DOMAIN_DICTIONARY_PATH)
    substitution_payload = load_substitution_policy()
    parser_hints_payload = load_parser_hints()
    golden_payload = load_json_payload(DEFAULT_MATCHING_GOLDEN_SET_PATH)
    reviewed_info = collect_file_info(reviewed_decisions_path)
    classifier_info = collect_file_info(CLASSIFIER_CONFIG_PATH)
    dictionary_info = collect_file_info(DOMAIN_DICTIONARY_PATH)
    substitution_info = collect_file_info(SUBSTITUTION_POLICY_PATH)
    parser_hints_info = collect_file_info(PARSER_HINTS_PATH)
    golden_info = collect_file_info(DEFAULT_MATCHING_GOLDEN_SET_PATH)
    script_info = collect_file_info(Path(__file__).resolve())
    return {
        "classifier_config_version": clean_text(classifier_payload.get("metadata", {}).get("version")),
        "classifier_config_sha256": classifier_info["sha256"] if classifier_info else "",
        "dictionary_generated_on": clean_text(dictionary_payload.get("metadata", {}).get("generated_on")),
        "dictionary_sha256": dictionary_info["sha256"] if dictionary_info else "",
        "substitution_policy_version": clean_text(substitution_payload.get("metadata", {}).get("version")),
        "substitution_policy_sha256": substitution_info["sha256"] if substitution_info else "",
        "parser_hints_version": clean_text(parser_hints_payload.get("metadata", {}).get("version")),
        "parser_hints_sha256": parser_hints_info["sha256"] if parser_hints_info else "",
        "golden_set_version": clean_text(golden_payload.get("metadata", {}).get("version") or golden_payload.get("version")),
        "golden_set_sha256": golden_info["sha256"] if golden_info else "",
        "golden_set_case_count": int(golden_payload.get("case_count", 0) or 0),
        "process_script_sha256": script_info["sha256"] if script_info else "",
        "reviewed_analog_decisions_sha256": reviewed_info["sha256"] if reviewed_info else "",
    }


def write_matching_artifact(
    order_path: Path,
    stock_path: Path,
    reviewed_decisions_path: Path | None,
    order_results: Sequence[MatchResult],
    stock_items: Sequence[StockItem],
    output_paths: Sequence[Path],
    out_path: Path,
) -> Path:
    stock_info = collect_file_info(stock_path)
    order_info = collect_file_info(order_path)
    reviewed_info = collect_file_info(reviewed_decisions_path)
    parser_hints_info = collect_file_info(PARSER_HINTS_PATH)
    payload = {
        "artifact_version": 1,
        "generated_at": datetime.now().astimezone().isoformat(),
        "order_name": order_path.name,
        "combined_input_hash": build_combined_input_hash([stock_info, order_info, reviewed_info, parser_hints_info]),
        "inputs": {
            "stock": stock_info,
            "order": order_info,
            "reviewed_analog_decisions": reviewed_info,
            "parser_hints": parser_hints_info,
        },
        "rules": build_rules_metadata(reviewed_decisions_path),
        "metrics": build_matching_metrics(order_results, stock_items),
        "outputs": [str(path) for path in output_paths],
    }
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_path


def write_outputs(
    order_path: Path,
    stock_path: Path,
    reviewed_decisions_path: Path | None,
    order_results: Sequence[MatchResult],
    stock_items: Sequence[StockItem],
    out_dir: Path,
) -> list[Path]:
    stem = order_path.stem
    detail_path = out_dir / f"{stem}__результат.xlsx"
    for_1c_path = out_dir / f"{stem}__для_1с.xlsx"
    analogs_path = out_dir / f"{stem}__аналоги.xlsx"
    not_found_path = out_dir / f"{stem}__не_найдено.xlsx"
    artifact_path = out_dir / f"{stem}__matching_artifact.json"
    create_detailed_workbook(order_results, stock_items, detail_path)
    create_for_1c_workbook(order_results, for_1c_path)
    create_analogs_workbook(order_results, analogs_path)
    create_not_found_workbook(order_results, not_found_path)
    output_paths = [detail_path, for_1c_path, analogs_path, not_found_path]
    write_matching_artifact(
        order_path=order_path,
        stock_path=stock_path,
        reviewed_decisions_path=reviewed_decisions_path,
        order_results=order_results,
        stock_items=stock_items,
        output_paths=output_paths,
        out_path=artifact_path,
    )
    return output_paths + [artifact_path]


def print_summary(order_path: Path, results: Sequence[MatchResult], output_paths: Sequence[Path]) -> None:
    statuses: defaultdict[str, int] = defaultdict(int)
    for result in results:
        statuses[result.status] += 1
    print(f"\n{order_path.name}")
    print(f"  строк: {len(results)}")
    for status in (
        STATUS_FOUND_FULL,
        STATUS_FOUND_PARTIAL,
        STATUS_SAFE_ANALOG,
        STATUS_APPROVAL_ANALOG,
        STATUS_NOT_FOUND,
        STATUS_STOCK_DEPLETED,
    ):
        print(f"  {status}: {statuses[status]}")
    print(f"  файлов создано: {len(output_paths)}")
    for path in output_paths:
        print(f"    - {path}")


def main() -> int:
    args = parse_args()
    stock_path = Path(args.stock).expanduser().resolve()
    order_paths = [Path(path).expanduser().resolve() for path in args.orders]
    out_dir = ensure_output_dir(Path(args.out_dir).expanduser().resolve())
    reviewed_decisions_path = (
        Path(args.reviewed_analog_decisions).expanduser().resolve()
        if args.reviewed_analog_decisions
        else DEFAULT_REVIEWED_ANALOG_DECISIONS_PATH
    )
    reviewed_decisions = load_reviewed_analog_decisions(reviewed_decisions_path)
    substitution_policy = load_substitution_policy()

    base_stock_items = load_stock(stock_path)
    if reviewed_decisions:
        decisions_count = sum(len(code_map) for code_map in reviewed_decisions.values())
        print(
            f"Загружены ручные решения по аналогам: {decisions_count} "
            f"из {reviewed_decisions_path}"
        )

    for order_path in order_paths:
        stock_items = clone_stock_items(base_stock_items)
        matcher = StockMatcher(
            stock_items,
            reviewed_analog_decisions=reviewed_decisions,
            substitution_policy=substitution_policy,
        )
        order_lines = load_order_lines(order_path)
        order_results = match_orders(order_lines, matcher)
        output_paths = write_outputs(
            order_path=order_path,
            stock_path=stock_path,
            reviewed_decisions_path=reviewed_decisions_path if reviewed_decisions else None,
            order_results=order_results,
            stock_items=stock_items,
            out_dir=out_dir,
        )
        print_summary(order_path, order_results, output_paths)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
